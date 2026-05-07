from dotenv import load_dotenv
from pathlib import Path
ROOT_DIR = Path(__file__).parent
load_dotenv(ROOT_DIR / '.env')

import os
import random
import logging
import time
import uuid
from datetime import datetime, timezone, timedelta
from typing import Optional, List, Dict, Any

import bcrypt
import jwt
import re
from fastapi import FastAPI, APIRouter, HTTPException, Request, Response, Depends, Query
from fastapi.responses import StreamingResponse
from starlette.middleware.cors import CORSMiddleware
from motor.motor_asyncio import AsyncIOMotorClient
from pydantic import BaseModel, Field, validator

from integrations.asaas import AsaasClient, AsaasError, map_asaas_status
from integrations.paypal import PayPalClient, PayPalError, map_paypal_status
from integrations.fusionpbx import (
    FusionPBXClient, FusionPBXError,
    normalize_extension, normalize_agent, normalize_queue, normalize_cdr,
)
from integrations.fusionpbx_db import FusionPBXDBClient, FusionPBXDBError
from integrations.fusionpbx_sftp import (
    find_recording, stream_recording, RecordingFetchError,
)
from integrations.freeswitch_esl import (
    FreeSwitchESL, FreeSwitchESLError, normalize_esl_channel,
)


async def _pbx_resync_agent_from_db(tid: str, agent_external_id: str) -> Dict[str, Any]:
    """Lê agent_contact/agent_status do PostgreSQL do PBX e aplica em memória do
    mod_callcenter via ESL. Use depois de QUALQUER mudança direta no banco
    (via GUI do FusionPBX, psql, ou UPDATE no próprio Voxyra). O
    `reload mod_callcenter` não atualiza agentes já carregados — só
    `agent set contact/status` via ESL garante reflexo imediato."""
    out: Dict[str, Any] = {"contact": None, "status": None, "errors": []}
    if not agent_external_id:
        return out
    try:
        s = await db.fusionpbx_settings.find_one({"tenant_id": tid}) or {}
        if not (s.get("enabled") and s.get("esl_host")
                and s.get("connection_type") == "db" and s.get("db_host")):
            return out
        pg = FusionPBXDBClient(
            host=s["db_host"], port=int(s.get("db_port") or 5432),
            database=s.get("db_name") or "fusionpbx",
            username=s["db_username"], password=s.get("db_password") or "",
            domain_uuid=s.get("domain_uuid"), ssl=bool(s.get("db_ssl")),
        )
        try:
            conn = await pg._connect()
            row = await conn.fetchrow(
                """SELECT agent_contact, agent_status
                   FROM v_call_center_agents
                   WHERE call_center_agent_uuid = $1::uuid""",
                agent_external_id,
            )
            await conn.close()
        except Exception as e:
            out["errors"].append(f"read db: {e}")
            return out
        if not row:
            out["errors"].append("agent not found in PBX db")
            return out
        await _pbx_apply_agent_live(
            tid, agent_external_id,
            contact=row["agent_contact"],
            status=row["agent_status"] or "Logged Out",
        )
        out["contact"] = row["agent_contact"]
        out["status"] = row["agent_status"]
    except Exception as e:
        out["errors"].append(str(e))
    return out


async def _pbx_reload_callcenter(tid: str) -> None:
    """After changing tiers/agents in DB, tell mod_callcenter to reload so the
    distributor picks up the changes. Best-effort: silent on failure."""
    try:
        s = await db.fusionpbx_settings.find_one({"tenant_id": tid}) or {}
        if not (s.get("enabled") and s.get("esl_host")):
            return
        esl = FreeSwitchESL(
            host=s["esl_host"], port=int(s.get("esl_port") or 8021),
            password=s.get("esl_password") or "ClueCon",
            timeout=float(s.get("esl_timeout") or 5.0),
        )
        await esl.callcenter_reload()
    except Exception as e:
        logger.warning("callcenter_reload falhou: %s", e)


async def _pbx_resolve_agent_esl_name(tid: str, agent_external_id: str) -> Optional[str]:
    """No mod_callcenter, o agente é identificado por `agent_id` em memória
    (que costuma ser `<extension>@<domain>`), NÃO pelo `call_center_agent_uuid`.
    Esta função consulta o PostgreSQL do FusionPBX para resolver o agent_id real.
    Cai pro UUID em último caso.
    """
    if not agent_external_id:
        return None
    try:
        s = await db.fusionpbx_settings.find_one({"tenant_id": tid}) or {}
        if not (s.get("connection_type") == "db" and s.get("db_host")):
            return agent_external_id
        pg = FusionPBXDBClient(
            host=s["db_host"], port=int(s.get("db_port") or 5432),
            database=s.get("db_name") or "fusionpbx",
            username=s["db_username"], password=s.get("db_password") or "",
            domain_uuid=s.get("domain_uuid"), ssl=bool(s.get("db_ssl")),
        )
        conn = await pg._connect()
        try:
            row = await conn.fetchrow(
                """SELECT agent_id FROM v_call_center_agents
                   WHERE call_center_agent_uuid = $1::uuid""",
                agent_external_id,
            )
            if row and row["agent_id"]:
                return row["agent_id"]
        finally:
            await conn.close()
    except Exception as e:
        logger.warning("resolve_agent_esl_name falhou: %s", e)
    return agent_external_id


async def _pbx_apply_agent_live(tid: str, agent_name: str, *,
                                  status: Optional[str] = None,
                                  state: Optional[str] = None,
                                  contact: Optional[str] = None,
                                  clear_tiers: bool = False,
                                  add_tier_queues: Optional[List[str]] = None,
                                  ) -> Dict[str, Any]:
    """Apply changes directly into mod_callcenter MEMORY (live distributor).
    `agent_name` should be the FusionPBX `agent_id` (typically
    `<ext>@<domain>`). If a UUID is passed, we resolve it to the real
    agent_id automatically.
    """
    out: Dict[str, Any] = {"status": None, "state": None, "contact": None,
                            "tiers_removed": [], "tiers_added": [], "errors": []}
    try:
        # Resolve UUID -> agent_id if it looks like a UUID (8-4-4-4-12)
        if agent_name and len(agent_name) == 36 and agent_name.count("-") == 4:
            real = await _pbx_resolve_agent_esl_name(tid, agent_name)
            if real and real != agent_name:
                logger.info("apply_agent_live: resolved UUID %s -> agent_id %s", agent_name, real)
                agent_name = real
        s = await db.fusionpbx_settings.find_one({"tenant_id": tid}) or {}
        if not (s.get("enabled") and s.get("esl_host") and agent_name):
            logger.info("apply_agent_live skip: enabled=%s host=%s name=%s",
                        s.get("enabled"), bool(s.get("esl_host")), bool(agent_name))
            return out
        esl = FreeSwitchESL(
            host=s["esl_host"], port=int(s.get("esl_port") or 8021),
            password=s.get("esl_password") or "ClueCon",
            timeout=float(s.get("esl_timeout") or 5.0),
        )
        if clear_tiers:
            try:
                out["tiers_removed"] = await esl.callcenter_clear_agent_tiers(agent_name)
                logger.info("ESL tiers_removed for %s: %s", agent_name, out["tiers_removed"])
            except Exception as e: out["errors"].append(f"clear_tiers: {e}")
        # Garante que o agente existe em memória do mod_callcenter ANTES de
        # mexer em contact/status/tier. Idempotente — se já existir, FreeSWITCH
        # retorna -ERR e seguimos. Isso resolve o caso onde o agente foi
        # descarregado da memória (ex: após Logged Out + reload) e os
        # comandos seguintes seriam ignorados silenciosamente.
        if contact is not None or status is not None or state is not None or add_tier_queues:
            try:
                r = await esl.callcenter_agent_add(agent_name, "callback")
                logger.info("ESL agent add %s | reply=%s", agent_name, r[:120])
            except Exception as e:
                # silencioso — agente já existir é normal
                logger.debug("ESL agent add (já existe?) %s: %s", agent_name, e)
        if contact is not None:
            try:
                r = await esl.callcenter_agent_set(agent_name, "contact", contact)
                out["contact"] = r
                logger.info("ESL set contact %s -> %s | reply=%s", agent_name, contact, r[:120])
            except Exception as e: out["errors"].append(f"contact: {e}")
        if status is not None:
            try:
                r = await esl.callcenter_agent_set(agent_name, "status", status)
                out["status"] = r
                logger.info("ESL set status %s -> %s | reply=%s", agent_name, status, r[:120])
            except Exception as e: out["errors"].append(f"status: {e}")
        if state is not None:
            try:
                r = await esl.callcenter_agent_set(agent_name, "state", state)
                out["state"] = r
                logger.info("ESL set state %s -> %s | reply=%s", agent_name, state, r[:120])
            except Exception as e: out["errors"].append(f"state: {e}")
        if add_tier_queues:
            for qname in add_tier_queues:
                try:
                    r = await esl.callcenter_tier_add(qname, agent_name, 1, 1)
                    out["tiers_added"].append(qname)
                    logger.info("ESL tier add %s <- %s | reply=%s", qname, agent_name, r[:120])
                except Exception as e: out["errors"].append(f"tier_add {qname}: {e}")
    except Exception as e:
        out["errors"].append(str(e))
        logger.warning("_pbx_apply_agent_live exception: %s", e)
    return out

EMAIL_RE = re.compile(r"^[A-Za-z0-9._%+\-]+@[A-Za-z0-9.\-]+\.[A-Za-z]{2,}$")

def validate_email_str(v: str) -> str:
    if not v or not isinstance(v, str): raise ValueError("Email inválido")
    v = v.strip().lower()
    if not EMAIL_RE.match(v): raise ValueError("Formato de email inválido")
    return v

# ---------- DB ----------
mongo_url = os.environ['MONGO_URL']
client = AsyncIOMotorClient(mongo_url)
db = client[os.environ['DB_NAME']]

app = FastAPI(title="Voxyra CCA - Callcenter Analytical")
api = APIRouter(prefix="/api")

# ---------- Auth helpers ----------
JWT_ALGO = "HS256"

def _secret(): return os.environ["JWT_SECRET"]

def hash_password(pw: str) -> str:
    return bcrypt.hashpw(pw.encode(), bcrypt.gensalt()).decode()

def verify_password(pw: str, hashed: str) -> bool:
    try: return bcrypt.checkpw(pw.encode(), hashed.encode())
    except Exception: return False

def create_access_token(user_id: str, email: str, role: str, tenant_id: Optional[str]) -> str:
    payload = {
        "sub": user_id, "email": email, "role": role, "tenant_id": tenant_id,
        "exp": datetime.now(timezone.utc) + timedelta(hours=8),
        "type": "access",
    }
    return jwt.encode(payload, _secret(), algorithm=JWT_ALGO)

async def get_current_user(request: Request) -> dict:
    token = request.cookies.get("access_token")
    if not token:
        auth = request.headers.get("Authorization", "")
        if auth.startswith("Bearer "): token = auth[7:]
    if not token:
        # Fallback: token via query string (used by <audio src> and <a download> tags
        # since browsers cannot attach Authorization headers on those native requests).
        token = request.query_params.get("token") or request.query_params.get("access_token")
    if not token:
        raise HTTPException(status_code=401, detail="Não autenticado")
    try:
        payload = jwt.decode(token, _secret(), algorithms=[JWT_ALGO])
        if payload.get("type") != "access":
            raise HTTPException(status_code=401, detail="Token inválido")
        user = await db.users.find_one({"id": payload["sub"]}, {"_id": 0, "password_hash": 0})
        if not user:
            raise HTTPException(status_code=401, detail="Usuário não encontrado")
        # Super admin can have an active tenant context via header (for impersonation)
        if user.get("role") == "super_admin":
            ctx = request.headers.get("X-Tenant-Context")
            user["_tenant_context"] = ctx or None
        return user
    except jwt.ExpiredSignatureError:
        raise HTTPException(status_code=401, detail="Token expirado")
    except jwt.InvalidTokenError:
        raise HTTPException(status_code=401, detail="Token inválido")

# ---------- Permissions ----------
ALL_PERMISSIONS = [
    {"key": "dashboard.view",       "label": "Ver Dashboard",                 "group": "Dashboard"},
    {"key": "realtime.view",        "label": "Ver Chamadas em Tempo Real",    "group": "Operação"},
    {"key": "tv.view",              "label": "Acessar Painel TV",             "group": "Operação"},
    {"key": "recordings.view_own",  "label": "Ouvir suas próprias gravações", "group": "Gravações"},
    {"key": "recordings.view_all",  "label": "Ouvir gravações de toda equipe","group": "Gravações"},
    {"key": "recordings.download",  "label": "Baixar gravações",              "group": "Gravações"},
    {"key": "recordings.edit_notes","label": "Adicionar anotações em gravações","group": "Gravações"},
    {"key": "reports.view",         "label": "Ver relatórios",                "group": "Relatórios"},
    {"key": "reports.export",       "label": "Exportar relatórios (Excel/PDF)","group": "Relatórios"},
    {"key": "queues.view",          "label": "Ver filas",                     "group": "Filas"},
    {"key": "queues.edit",          "label": "Editar filas",                  "group": "Filas"},
    {"key": "agents.view",          "label": "Ver agentes",                   "group": "Agentes"},
    {"key": "agents.edit",          "label": "Editar agentes",                "group": "Agentes"},
    {"key": "agent.change_extension","label": "Trocar de ramal sem deslogar",  "group": "Agentes"},
    {"key": "users.manage",         "label": "Gerenciar usuários",            "group": "Administração"},
    {"key": "tenant.settings",      "label": "Configurar empresa (tenant)",   "group": "Administração"},
]

DEFAULT_PERMISSIONS_BY_ROLE = {
    "super_admin": [p["key"] for p in ALL_PERMISSIONS],
    "admin": [p["key"] for p in ALL_PERMISSIONS if p["key"] != "tenants.manage"],
    "supervisor": [
        "dashboard.view", "realtime.view", "tv.view",
        "recordings.view_all", "recordings.download", "recordings.edit_notes",
        "reports.view", "reports.export",
        "queues.view", "queues.edit",
        "agents.view", "agents.edit",
    ],
    "agent": ["dashboard.view", "recordings.view_own", "reports.view",
              "agent.change_extension"],
}

# Built-in roles (não podem ser deletados, mas as permissões podem ser editadas)
BUILTIN_ROLES = ["admin", "supervisor", "agent"]

# Cache simples por tenant (TTL implícito — invalidado em escritas)
_role_template_cache: Dict[str, Dict[str, List[str]]] = {}

async def _load_tenant_role_templates(tid: str) -> Dict[str, List[str]]:
    """Carrega todos os templates de roles do tenant. Retorna {role_key: [perms]}."""
    if not tid:
        return {}
    if tid in _role_template_cache:
        return _role_template_cache[tid]
    docs = await db.role_templates.find({"tenant_id": tid}, {"_id": 0}).to_list(50)
    out = {d["key"]: list(d.get("permissions") or []) for d in docs}
    _role_template_cache[tid] = out
    return out

def _invalidate_role_template_cache(tid: Optional[str] = None):
    if tid is None:
        _role_template_cache.clear()
    else:
        _role_template_cache.pop(tid, None)

async def effective_permissions_async(user: dict) -> List[str]:
    """Permissões efetivas de um usuário considerando:
    1) user.permissions custom (se setado, vence tudo)
    2) template do role no tenant atual (db.role_templates)
    3) DEFAULT_PERMISSIONS_BY_ROLE (fallback)"""
    role = user.get("role", "agent")
    if role == "super_admin":
        return DEFAULT_PERMISSIONS_BY_ROLE["super_admin"]
    perms = user.get("permissions")
    if perms is not None:
        return perms
    tid = user.get("tenant_id")
    if tid:
        tpl = await _load_tenant_role_templates(tid)
        if role in tpl:
            return tpl[role]
    return DEFAULT_PERMISSIONS_BY_ROLE.get(role, [])

def effective_permissions(user: dict) -> List[str]:
    """Versão sync — usado em código sync/legado. Não consulta DB,
    cai no DEFAULT se o user não tiver custom. Para verificações
    com templates use require_permission (que é async)."""
    if user.get("role") in ("super_admin", "admin"):
        return DEFAULT_PERMISSIONS_BY_ROLE[user["role"]]
    perms = user.get("permissions")
    if perms is None:
        # Tenta cache em memória (foi populado por require_permission antes)
        tid = user.get("tenant_id")
        if tid and tid in _role_template_cache:
            tpl = _role_template_cache[tid]
            if user.get("role") in tpl:
                return tpl[user["role"]]
        return DEFAULT_PERMISSIONS_BY_ROLE.get(user.get("role", "agent"), [])
    return perms

def require_permission(perm: str):
    async def checker(user: dict = Depends(get_current_user)):
        perms = await effective_permissions_async(user)
        if perm not in perms:
            raise HTTPException(status_code=403, detail="Sem permissão")
        return user
    return checker

def require_super_admin():
    async def checker(user: dict = Depends(get_current_user)):
        if user.get("role") != "super_admin":
            raise HTTPException(status_code=403, detail="Apenas super-admin")
        return user
    return checker

def tenant_scope(user: dict) -> Optional[str]:
    """Returns the tenant_id to scope queries by. None means no scoping (super admin without context)."""
    if user.get("role") == "super_admin":
        return user.get("_tenant_context")  # may be None => global
    return user.get("tenant_id")

def tenant_filter(user: dict) -> Dict[str, Any]:
    tid = tenant_scope(user)
    return {"tenant_id": tid} if tid else {}

async def require_tenant_or_super(user: dict) -> str:
    """For write operations: returns the tenant_id to use. Super admin must have context."""
    tid = tenant_scope(user)
    if not tid:
        raise HTTPException(status_code=400, detail="Selecione um tenant antes de executar esta operação")
    return tid

# ---------- Models ----------
class LoginReq(BaseModel):
    domain: Optional[str] = None
    email: str
    password: str
    @validator("email")
    def _e(cls, v): return validate_email_str(v)

class TenantCreate(BaseModel):
    domain: str
    name: str
    accent_color: str = "#0EA5E9"
    logo_url: Optional[str] = None
    timezone: str = "America/Sao_Paulo"
    max_users: int = 50
    max_agents: int = 50
    active: bool = True
    plan_id: Optional[str] = None
    contract_value: Optional[float] = None
    contract_start: Optional[str] = None
    contract_end: Optional[str] = None
    payment_status: str = "pending"  # paid | pending | overdue | trial

class TenantUpdate(BaseModel):
    name: Optional[str] = None
    accent_color: Optional[str] = None
    logo_url: Optional[str] = None
    timezone: Optional[str] = None
    max_users: Optional[int] = None
    max_agents: Optional[int] = None
    active: Optional[bool] = None
    plan_id: Optional[str] = None
    contract_value: Optional[float] = None
    contract_start: Optional[str] = None
    contract_end: Optional[str] = None
    payment_status: Optional[str] = None

class PlanCreate(BaseModel):
    name: str
    description: Optional[str] = ""
    monthly_price: float
    max_users: int = 10
    max_agents: int = 10
    features: List[str] = []
    active: bool = True

class PlanUpdate(BaseModel):
    name: Optional[str] = None
    description: Optional[str] = None
    monthly_price: Optional[float] = None
    max_users: Optional[int] = None
    max_agents: Optional[int] = None
    features: Optional[List[str]] = None
    active: Optional[bool] = None

def _serialize_tenant(t: dict) -> dict:
    return {k: t.get(k) for k in [
        "id", "domain", "name", "accent_color", "logo_url",
        "timezone", "max_users", "max_agents", "active", "created_at",
        "plan_id", "contract_value", "contract_start", "contract_end", "payment_status",
    ]}

def _serialize_plan(p: dict) -> dict:
    return {k: p.get(k) for k in ["id", "name", "description", "monthly_price",
                                   "max_users", "max_agents", "features", "active", "created_at"]}

# ---------- Auth Endpoints ----------
def _set_auth_cookie(response: Response, token: str):
    response.set_cookie(
        key="access_token", value=token, httponly=True,
        secure=False, samesite="lax", max_age=8 * 3600, path="/",
    )

@api.post("/auth/login")
async def login(body: LoginReq, response: Response):
    email = body.email.lower().strip()
    domain = (body.domain or "").strip().lower()

    # Auto-extract domain from email if not provided (ex: "user@empresa.com.br")
    if not domain and "@" in email:
        domain_from_email = email.split("@", 1)[1]
        if domain_from_email:
            tenant_check = await db.tenants.find_one({"domain": domain_from_email})
            if tenant_check:
                domain = domain_from_email

    if domain:
        tenant = await db.tenants.find_one({"domain": domain})
        if not tenant:
            raise HTTPException(status_code=401, detail="Domínio não encontrado")
        if not tenant.get("active", True):
            raise HTTPException(status_code=403, detail="Tenant suspenso")
        user = await db.users.find_one({"tenant_id": tenant["id"], "email": email})
    else:
        # Super-admin login (no tenant)
        user = await db.users.find_one({"email": email, "role": "super_admin"})
    if not user or not verify_password(body.password, user["password_hash"]):
        raise HTTPException(status_code=401, detail="Credenciais inválidas")
    if user.get("active") is False:
        raise HTTPException(status_code=403, detail="Usuário desativado")
    token = create_access_token(user["id"], user["email"], user["role"], user.get("tenant_id"))
    _set_auth_cookie(response, token)
    try:
        await write_audit(user, "login", "user", user["id"], f"{user.get('name')} <{user.get('email')}>", {"domain": domain})
    except Exception: pass
    return {
        "id": user["id"], "email": user["email"], "name": user["name"],
        "role": user["role"], "tenant_id": user.get("tenant_id"),
        "permissions": effective_permissions(user), "token": token,
    }

@api.post("/auth/logout")
async def logout(response: Response):
    response.delete_cookie("access_token", path="/")
    return {"ok": True}

@api.get("/auth/me")
async def me(user: dict = Depends(get_current_user)):
    out = {k: user.get(k) for k in ["id", "email", "name", "role", "tenant_id", "active"]}
    out["permissions"] = effective_permissions(user)
    if user.get("role") == "super_admin":
        out["tenant_context"] = user.get("_tenant_context")
    # attach tenant info
    tid = user.get("tenant_id") or user.get("_tenant_context")
    if tid:
        t = await db.tenants.find_one({"id": tid}, {"_id": 0})
        if t: out["tenant"] = _serialize_tenant(t)
    return out

@api.get("/auth/branding")
async def get_branding(domain: str = Query(...)):
    """Public endpoint: return tenant branding by domain (used on login screen)."""
    tenant = await db.tenants.find_one({"domain": domain.strip().lower(), "active": True}, {"_id": 0})
    if not tenant:
        return {"found": False}
    return {
        "found": True,
        "name": tenant.get("name"),
        "accent_color": tenant.get("accent_color"),
        "logo_url": tenant.get("logo_url"),
    }

# ---------- Audit ----------
async def write_audit(actor: dict, action: str, target_type: str, target_id: str, target_label: str, changes: Optional[dict] = None):
    doc = {
        "id": str(uuid.uuid4()),
        "tenant_id": actor.get("tenant_id"),
        "action": action, "target_type": target_type,
        "target_id": target_id, "target_label": target_label,
        "actor_id": actor.get("id"), "actor_email": actor.get("email"), "actor_name": actor.get("name"),
        "changes": changes or {},
        "created_at": datetime.now(timezone.utc).isoformat(),
    }
    await db.audit_logs.insert_one(doc)

# ---------- Tenants (super admin) ----------
@api.get("/tenants")
async def list_tenants(user: dict = Depends(require_super_admin())):
    docs = await db.tenants.find({}, {"_id": 0}).sort("created_at", 1).to_list(500)
    out = []
    for t in docs:
        users = await db.users.count_documents({"tenant_id": t["id"]})
        agents = await db.agents.count_documents({"tenant_id": t["id"]})
        out.append({**_serialize_tenant(t), "user_count": users, "agent_count": agents})
    return {"tenants": out}

@api.post("/tenants")
async def create_tenant(body: TenantCreate, user: dict = Depends(require_super_admin())):
    domain = body.domain.strip().lower()
    if not domain or " " in domain:
        raise HTTPException(status_code=400, detail="Domínio inválido")
    if await db.tenants.find_one({"domain": domain}):
        raise HTTPException(status_code=400, detail="Domínio já cadastrado")
    tid = str(uuid.uuid4())
    doc = {
        "id": tid, "domain": domain, "name": body.name,
        "accent_color": body.accent_color, "logo_url": body.logo_url,
        "timezone": body.timezone, "max_users": body.max_users,
        "max_agents": body.max_agents, "active": body.active,
        "created_at": datetime.now(timezone.utc).isoformat(),
    }
    await db.tenants.insert_one(doc)
    await write_audit(user, "create", "tenant", tid, f"{body.name} ({domain})", {"domain": domain})
    return _serialize_tenant(doc)

@api.patch("/tenants/{tid}")
async def update_tenant(tid: str, body: TenantUpdate, user: dict = Depends(require_super_admin())):
    target = await db.tenants.find_one({"id": tid})
    if not target: raise HTTPException(status_code=404, detail="Tenant não encontrado")
    update = {}
    changes = {}
    for f in ["name", "accent_color", "logo_url", "timezone", "max_users", "max_agents", "active"]:
        v = getattr(body, f)
        if v is not None and v != target.get(f):
            update[f] = v
            changes[f] = {"from": target.get(f), "to": v}
    if update:
        await db.tenants.update_one({"id": tid}, {"$set": update})
        await write_audit(user, "update", "tenant", tid, target.get("name", ""), changes)
    fresh = await db.tenants.find_one({"id": tid}, {"_id": 0})
    return _serialize_tenant(fresh)

@api.delete("/tenants/{tid}")
async def delete_tenant(tid: str, user: dict = Depends(require_super_admin())):
    target = await db.tenants.find_one({"id": tid})
    if not target: raise HTTPException(status_code=404, detail="Tenant não encontrado")
    # Cascade delete tenant data
    for col in ("users", "agents", "queues", "calls", "recordings", "audit_logs"):
        await db[col].delete_many({"tenant_id": tid})
    await db.tenants.delete_one({"id": tid})
    await write_audit(user, "delete", "tenant", tid, target.get("name", ""), {"domain": target.get("domain")})
    return {"ok": True}

@api.post("/tenants/{tid}/impersonate")
async def impersonate_tenant(tid: str, user: dict = Depends(require_super_admin())):
    """Set super admin's active tenant context. Frontend should send X-Tenant-Context header."""
    t = await db.tenants.find_one({"id": tid})
    if not t: raise HTTPException(status_code=404, detail="Tenant não encontrado")
    return {"tenant": _serialize_tenant(t)}

@api.get("/tenants/{tid}/stats")
async def tenant_stats(tid: str, user: dict = Depends(require_super_admin())):
    """Usage statistics for a single tenant - super admin view."""
    t = await db.tenants.find_one({"id": tid})
    if not t: raise HTTPException(status_code=404, detail="Tenant não encontrado")
    f = {"tenant_id": tid}
    cutoff_30d = (datetime.now(timezone.utc) - timedelta(days=30)).isoformat()
    users_count = await db.users.count_documents(f)
    agents_count = await db.agents.count_documents(f)
    queues_count = await db.queues.count_documents(f)
    calls_30d = await db.calls.count_documents({**f, "started_at": {"$gte": cutoff_30d}})
    answered_30d = await db.calls.count_documents({**f, "started_at": {"$gte": cutoff_30d}, "disposition": "answered"})
    recordings = await db.recordings.find({**f, "started_at": {"$gte": cutoff_30d}}, {"_id": 0, "duration_sec": 1, "size_mb": 1}).to_list(5000)
    total_minutes = sum(r.get("duration_sec", 0) for r in recordings) / 60
    total_storage_mb = sum(r.get("size_mb", 0) for r in recordings)
    # Plan info
    plan = None
    if t.get("plan_id"):
        plan = await db.plans.find_one({"id": t["plan_id"]}, {"_id": 0})
    return {
        "tenant": _serialize_tenant(t),
        "plan": _serialize_plan(plan) if plan else None,
        "usage": {
            "users": users_count, "users_limit": t.get("max_users", 0),
            "agents": agents_count, "agents_limit": t.get("max_agents", 0),
            "queues": queues_count,
            "calls_30d": calls_30d,
            "answered_30d": answered_30d,
            "recording_minutes_30d": round(total_minutes, 1),
            "recording_storage_mb": round(total_storage_mb, 1),
            "recordings_count_30d": len(recordings),
        },
    }

# ---------- Plan Features Catalog ----------
PLAN_FEATURES_CATALOG = [
    {"group": "Operação", "key": "dashboard",            "label": "Dashboard analítico",            "description": "KPIs, gráficos e abandonos"},
    {"group": "Operação", "key": "realtime",             "label": "Monitor em Tempo Real",          "description": "Chamadas ao vivo + status agentes"},
    {"group": "Operação", "key": "tv_panel",             "label": "Painel TV (wallboard)",          "description": "Modo TV com auto-refresh"},
    {"group": "Operação", "key": "tv_customization",     "label": "Personalização Painel TV",       "description": "Temas, rotação, alertas sonoros"},
    {"group": "Gravações","key": "recordings",           "label": "Gravações de chamadas",          "description": "Acesso ao acervo de gravações"},
    {"group": "Gravações","key": "recordings_download",  "label": "Download de gravações",          "description": "Baixar arquivos de áudio"},
    {"group": "Gravações","key": "recordings_notes",     "label": "Anotações em gravações",         "description": "Adicionar comentários por chamada"},
    {"group": "Gravações","key": "ai_analysis",          "label": "Análise por IA (sentimento)",    "description": "Transcrição + score de qualidade"},
    {"group": "Filas",    "key": "queues_view",          "label": "Visualizar filas",               "description": "Estatísticas e fila atual"},
    {"group": "Filas",    "key": "queues_edit",          "label": "Editar filas",                   "description": "Criar/configurar filas"},
    {"group": "Filas",    "key": "abandoned_analytics",  "label": "Análise de abandonos",           "description": "Por hora/dia/semana e tipo"},
    {"group": "Agentes",  "key": "agents_view",          "label": "Listar agentes",                 "description": "Visualizar equipe"},
    {"group": "Agentes",  "key": "agents_edit",          "label": "Gerenciar agentes",              "description": "Editar perfis e status"},
    {"group": "Agentes",  "key": "agent_scoring",        "label": "Performance/CSAT por agente",    "description": "Ranking e métricas"},
    {"group": "Relatórios","key":"reports",              "label": "Relatórios padrão",              "description": "6 tipos de relatório"},
    {"group": "Relatórios","key":"reports_export",       "label": "Exportar Excel/PDF",             "description": "Download de relatórios"},
    {"group": "Relatórios","key":"custom_reports",       "label": "Relatórios customizados",        "description": "Criar relatórios sob medida"},
    {"group": "Administração","key":"users_management",  "label": "Gerenciar usuários",             "description": "CRUD de usuários e permissões"},
    {"group": "Administração","key":"audit_logs",        "label": "Logs de auditoria",              "description": "Histórico de ações"},
    {"group": "Administração","key":"multi_supervisor",  "label": "Múltiplos supervisores",         "description": "Mais de 1 supervisor"},
    {"group": "Marca",    "key": "white_label",          "label": "White-label",                    "description": "Logo + cor próprios"},
    {"group": "Marca",    "key": "custom_domain",        "label": "Domínio personalizado",          "description": "Acesso por domínio próprio"},
    {"group": "Integrações","key":"api_access",          "label": "Acesso à API REST",              "description": "Tokens de API para integração"},
    {"group": "Integrações","key":"webhooks",            "label": "Webhooks",                       "description": "Notificações de eventos"},
    {"group": "Suporte",  "key": "support_email",        "label": "Suporte por email",              "description": "Resposta em até 48h"},
    {"group": "Suporte",  "key": "support_priority",     "label": "Suporte prioritário",            "description": "Resposta em até 4h"},
    {"group": "Suporte",  "key": "support_dedicated",    "label": "Gerente dedicado",               "description": "Atendimento exclusivo"},
    {"group": "Suporte",  "key": "sla_99",               "label": "SLA 99,9%",                      "description": "Disponibilidade garantida"},
]

@api.get("/plans/features-catalog")
async def plans_features_catalog(user: dict = Depends(get_current_user)):
    return {"features": PLAN_FEATURES_CATALOG}

# ---------- Plans (super admin) ----------
@api.get("/plans")
async def list_plans(user: dict = Depends(get_current_user)):
    """Public to authenticated users so tenants can see their plan info."""
    docs = await db.plans.find({}, {"_id": 0}).sort("monthly_price", 1).to_list(100)
    return {"plans": [_serialize_plan(p) for p in docs]}

@api.post("/plans")
async def create_plan(body: PlanCreate, user: dict = Depends(require_super_admin())):
    pid = str(uuid.uuid4())
    doc = {"id": pid, **body.dict(), "created_at": datetime.now(timezone.utc).isoformat()}
    await db.plans.insert_one(doc)
    await write_audit(user, "create", "plan", pid, body.name, {"price": body.monthly_price})
    return _serialize_plan(doc)

@api.patch("/plans/{pid}")
async def update_plan(pid: str, body: PlanUpdate, user: dict = Depends(require_super_admin())):
    target = await db.plans.find_one({"id": pid})
    if not target: raise HTTPException(status_code=404, detail="Plano não encontrado")
    update = {k: v for k, v in body.dict(exclude_unset=True).items() if v is not None}
    if update:
        await db.plans.update_one({"id": pid}, {"$set": update})
        await write_audit(user, "update", "plan", pid, target.get("name", ""), update)
    fresh = await db.plans.find_one({"id": pid}, {"_id": 0})
    return _serialize_plan(fresh)

@api.delete("/plans/{pid}")
async def delete_plan(pid: str, user: dict = Depends(require_super_admin())):
    target = await db.plans.find_one({"id": pid})
    if not target: raise HTTPException(status_code=404, detail="Plano não encontrado")
    in_use = await db.tenants.count_documents({"plan_id": pid})
    if in_use:
        raise HTTPException(status_code=400, detail=f"Plano em uso por {in_use} tenant(s). Mova-os antes de excluir.")
    await db.plans.delete_one({"id": pid})
    await write_audit(user, "delete", "plan", pid, target.get("name", ""), {})
    return {"ok": True}

# ---------- Logo upload (local FS) ----------
import shutil
from fastapi import UploadFile, File
from fastapi.staticfiles import StaticFiles

UPLOAD_DIR = ROOT_DIR / "uploads"
UPLOAD_DIR.mkdir(exist_ok=True)

@api.post("/uploads/logo")
async def upload_logo(file: UploadFile = File(...), user: dict = Depends(require_super_admin())):
    ext = (file.filename or "").rsplit(".", 1)[-1].lower()
    if ext not in {"png", "jpg", "jpeg", "webp", "svg", "gif"}:
        raise HTTPException(status_code=400, detail="Formato inválido (use png/jpg/webp/svg)")
    filename = f"{uuid.uuid4()}.{ext}"
    path = UPLOAD_DIR / filename
    with path.open("wb") as f:
        shutil.copyfileobj(file.file, f)
    # URL served via /uploads/<filename>
    return {"url": f"/uploads/{filename}", "filename": filename, "size": path.stat().st_size}


@api.post("/uploads/asset")
async def upload_asset(file: UploadFile = File(...), kind: str = "logo",
                        user: dict = Depends(require_super_admin())):
    """Generic asset upload (logo / wallpaper / favicon).
    `kind` is informational only; storage is the same.
    """
    allowed = {
        "logo":      {"png", "jpg", "jpeg", "webp", "svg", "gif"},
        "wallpaper": {"png", "jpg", "jpeg", "webp", "gif"},
        "favicon":   {"png", "ico", "svg"},
    }.get(kind, {"png", "jpg", "jpeg", "webp", "svg", "gif", "ico"})
    ext = (file.filename or "").rsplit(".", 1)[-1].lower()
    if ext not in allowed:
        raise HTTPException(status_code=400,
            detail=f"Formato inválido para {kind}. Use: {', '.join(sorted(allowed))}")
    # Light size cap: 8MB for wallpapers, 2MB for logo/favicon
    max_bytes = 8 * 1024 * 1024 if kind == "wallpaper" else 2 * 1024 * 1024
    filename = f"{kind}-{uuid.uuid4()}.{ext}"
    path = UPLOAD_DIR / filename
    written = 0
    with path.open("wb") as f:
        while True:
            chunk = await file.read(64 * 1024)
            if not chunk:
                break
            written += len(chunk)
            if written > max_bytes:
                f.close()
                path.unlink(missing_ok=True)
                raise HTTPException(status_code=413,
                    detail=f"Arquivo maior que {max_bytes // (1024*1024)} MB")
            f.write(chunk)
    return {"url": f"/uploads/{filename}", "filename": filename, "size": written, "kind": kind}


# ---------- Site Branding (global, super admin) ----------
class SiteBrandingMode(BaseModel):
    hero_title: Optional[str] = None
    hero_subtitle: Optional[str] = None
    accent_color: Optional[str] = None
    wallpaper_url: Optional[str] = None
    logo_url: Optional[str] = None


class SiteBranding(BaseModel):
    brand_name: Optional[str] = None
    brand_subtitle: Optional[str] = None
    login_title: Optional[str] = None
    login_subtitle: Optional[str] = None
    logo_url: Optional[str] = None
    wallpaper_url: Optional[str] = None
    favicon_url: Optional[str] = None
    footer_text: Optional[str] = None
    release_version: Optional[str] = None
    accent_color: Optional[str] = None
    modes: Optional[Dict[str, SiteBrandingMode]] = None  # {agent, master, admin}


def _site_branding_defaults() -> dict:
    return {
        "id": "global",
        "brand_name": "Voxyra CCA",
        "brand_subtitle": "Callcenter Analytical",
        "login_title": "",
        "login_subtitle": "",
        "logo_url": "",
        "wallpaper_url": "",
        "favicon_url": "",
        "footer_text": "",
        "release_version": "",
        "accent_color": "#09090b",
        "modes": {
            "agent":  {"hero_title": "", "hero_subtitle": "", "accent_color": "", "wallpaper_url": "", "logo_url": ""},
            "master": {"hero_title": "", "hero_subtitle": "", "accent_color": "", "wallpaper_url": "", "logo_url": ""},
            "admin":  {"hero_title": "", "hero_subtitle": "", "accent_color": "", "wallpaper_url": "", "logo_url": ""},
        },
    }


@api.get("/branding/site")
async def get_site_branding():
    """Public: returns global site branding for login pages and document head."""
    doc = await db.site_branding.find_one({"id": "global"}, {"_id": 0})
    if not doc:
        return _site_branding_defaults()
    # Ensure 'modes' is always present and complete
    defaults = _site_branding_defaults()
    modes = doc.get("modes") or {}
    out_modes = {}
    for k in ("agent", "master", "admin"):
        out_modes[k] = {**defaults["modes"][k], **(modes.get(k) or {})}
    doc["modes"] = out_modes
    return doc


@api.put("/branding/site")
async def update_site_branding(body: SiteBranding,
                                user: dict = Depends(require_super_admin())):
    raw = body.dict(exclude_unset=True)
    payload = {k: v for k, v in raw.items() if v is not None and k != "modes"}
    base = await db.site_branding.find_one({"id": "global"}, {"_id": 0}) or _site_branding_defaults()
    base.update(payload)
    if body.modes is not None:
        existing_modes = base.get("modes") or {}
        for mode_key, mode_val in body.modes.items():
            if mode_key not in ("agent", "master", "admin"):
                continue
            current = existing_modes.get(mode_key) or {}
            mv = mode_val.dict(exclude_unset=True) if hasattr(mode_val, "dict") else (mode_val or {})
            current.update({k: v for k, v in mv.items() if v is not None})
            existing_modes[mode_key] = current
        base["modes"] = existing_modes
    base["id"] = "global"
    base["updated_at"] = datetime.now(timezone.utc).isoformat()
    await db.site_branding.update_one({"id": "global"}, {"$set": base}, upsert=True)
    await write_audit(user, "update", "site_branding", "global", "Branding global", payload)
    out = {**base}
    out.pop("_id", None)
    return out

# ---------- Billing settings (placeholder for Asaas/PayPal credentials) ----------
class BillingSettings(BaseModel):
    asaas_api_key: Optional[str] = None
    asaas_environment: str = "production"  # production | sandbox
    asaas_webhook_token: Optional[str] = None
    paypal_client_id: Optional[str] = None
    paypal_client_secret: Optional[str] = None
    paypal_environment: str = "live"  # live | sandbox
    paypal_webhook_id: Optional[str] = None
    enabled_methods: List[str] = []  # ["pix", "boleto", "credit_card", "paypal"]

@api.get("/billing/settings")
async def get_billing_settings(user: dict = Depends(require_super_admin())):
    doc = await db.billing_settings.find_one({"id": "global"}, {"_id": 0}) or {}
    # mask secrets
    out = {**doc}
    for k in ("asaas_api_key", "asaas_webhook_token", "paypal_client_secret"):
        if out.get(k):
            v = out[k]
            out[k] = v[:6] + "•" * (len(v) - 10) + v[-4:] if len(v) > 12 else "••••"
            out[f"{k}_set"] = True
        else:
            out[f"{k}_set"] = False
    return out

@api.put("/billing/settings")
async def update_billing_settings(body: BillingSettings, user: dict = Depends(require_super_admin())):
    payload = body.dict(exclude_unset=True)
    # Don't overwrite secrets if a masked value was sent back
    existing = await db.billing_settings.find_one({"id": "global"}) or {}
    for k in ("asaas_api_key", "asaas_webhook_token", "paypal_client_secret"):
        if k in payload and payload[k] and "•" in str(payload[k]):
            payload[k] = existing.get(k)  # keep existing
    await db.billing_settings.update_one({"id": "global"}, {"$set": {"id": "global", **payload}}, upsert=True)
    await write_audit(user, "update", "billing_settings", "global", "Configurações de cobrança", {"fields": list(payload.keys())})
    return {"ok": True}

# ---------- Permissions metadata ----------
@api.get("/permissions")
async def list_permissions(user: dict = Depends(require_permission("users.manage"))):
    """Retorna metadados de permissões + grupos (templates) do tenant atual.
    Cada item de `roles` traz já a lista de permissões efetivas do grupo
    (custom do tenant ou fallback default)."""
    tid = tenant_scope(user)
    tpl = await _load_tenant_role_templates(tid) if tid else {}
    builtin_meta = {
        "admin": "Administrador",
        "supervisor": "Supervisor",
        "agent": "Agente",
    }
    roles = []
    for k, label in builtin_meta.items():
        roles.append({
            "key": k,
            "label": label,
            "is_builtin": True,
            "permissions": tpl.get(k, DEFAULT_PERMISSIONS_BY_ROLE.get(k, [])),
            "is_custom": k in tpl,
        })
    if tid:
        custom_docs = await db.role_templates.find(
            {"tenant_id": tid, "key": {"$nin": list(builtin_meta.keys())}},
            {"_id": 0}).to_list(50)
        for d in custom_docs:
            roles.append({
                "key": d["key"],
                "label": d.get("label") or d["key"],
                "is_builtin": False,
                "permissions": d.get("permissions") or [],
                "is_custom": True,
            })
    return {
        "permissions": ALL_PERMISSIONS,
        "defaults": DEFAULT_PERMISSIONS_BY_ROLE,
        "roles": roles,
    }


class RoleTemplateUpsert(BaseModel):
    key: str
    label: Optional[str] = None
    permissions: List[str] = []
    @validator("key")
    def _k(cls, v):
        if not v or not v.strip():
            raise ValueError("Chave do grupo obrigatória")
        v = v.strip().lower()
        if not all(c.isalnum() or c in "_-" for c in v):
            raise ValueError("Chave só pode ter letras, números, _ e -")
        if v in ("super_admin",):
            raise ValueError("Chave reservada")
        return v


@api.get("/role-templates")
async def list_role_templates(user: dict = Depends(require_permission("users.manage"))):
    """Lista templates de roles do tenant atual, incluindo built-ins (admin,
    supervisor, agent) com fallback para os defaults quando não há override."""
    tid = await require_tenant_or_super(user)
    tpl = await _load_tenant_role_templates(tid)
    docs = await db.role_templates.find({"tenant_id": tid}, {"_id": 0}).to_list(50)
    docs_by_key = {d["key"]: d for d in docs}
    out = []
    builtin_labels = {"admin": "Administrador", "supervisor": "Supervisor", "agent": "Agente"}
    for k in BUILTIN_ROLES:
        d = docs_by_key.get(k)
        out.append({
            "key": k,
            "label": (d.get("label") if d else None) or builtin_labels[k],
            "is_builtin": True,
            "permissions": tpl.get(k) or DEFAULT_PERMISSIONS_BY_ROLE.get(k, []),
            "user_count": await db.users.count_documents({"tenant_id": tid, "role": k}),
            "has_override": k in tpl,
        })
    for k, d in docs_by_key.items():
        if k in BUILTIN_ROLES:
            continue
        out.append({
            "key": k,
            "label": d.get("label") or k,
            "is_builtin": False,
            "permissions": d.get("permissions") or [],
            "user_count": await db.users.count_documents({"tenant_id": tid, "role": k}),
            "has_override": True,
        })
    return {"roles": out}


@api.put("/role-templates/{role_key}")
async def upsert_role_template(role_key: str, body: RoleTemplateUpsert,
                                  user: dict = Depends(require_permission("users.manage"))):
    """Cria ou atualiza o template de um role no tenant. Para built-ins
    (admin/supervisor/agent), apenas as permissões são sobrescritas. Para
    roles custom, a chave do path deve ser igual ao body.key."""
    tid = await require_tenant_or_super(user)
    role_key = role_key.strip().lower()
    if role_key != body.key:
        raise HTTPException(status_code=400, detail="Chave da URL difere do body")
    valid_perms = {p["key"] for p in ALL_PERMISSIONS}
    invalid = [p for p in body.permissions if p not in valid_perms]
    if invalid:
        raise HTTPException(status_code=400, detail=f"Permissões inválidas: {invalid}")
    is_builtin = role_key in BUILTIN_ROLES
    label = body.label or (
        {"admin": "Administrador", "supervisor": "Supervisor", "agent": "Agente"}.get(role_key, role_key)
    )
    now = datetime.now(timezone.utc).isoformat()
    existing = await db.role_templates.find_one({"tenant_id": tid, "key": role_key}, {"_id": 0})
    doc = {
        "id": existing.get("id") if existing else str(uuid.uuid4()),
        "tenant_id": tid, "key": role_key, "label": label,
        "permissions": list(set(body.permissions)),
        "is_builtin": is_builtin,
        "updated_at": now,
        "created_at": existing.get("created_at") if existing else now,
    }
    await db.role_templates.update_one(
        {"tenant_id": tid, "key": role_key},
        {"$set": doc}, upsert=True)
    _invalidate_role_template_cache(tid)
    await write_audit(user, "upsert", "role_template", role_key,
                       f"Grupo {label} · {len(doc['permissions'])} permissão(ões)",
                       {"permissions": doc["permissions"]})
    return {"ok": True, "role": {**doc, "user_count": await db.users.count_documents({"tenant_id": tid, "role": role_key})}}


@api.delete("/role-templates/{role_key}")
async def delete_role_template(role_key: str,
                                  user: dict = Depends(require_permission("users.manage"))):
    """Remove um role custom. Built-ins não podem ser removidos — apenas
    redefinidos para o default ao deletar o override."""
    tid = await require_tenant_or_super(user)
    role_key = role_key.strip().lower()
    if role_key in BUILTIN_ROLES:
        # Para built-in, apenas remove o override (volta ao default)
        await db.role_templates.delete_one({"tenant_id": tid, "key": role_key})
        _invalidate_role_template_cache(tid)
        await write_audit(user, "reset", "role_template", role_key,
                           f"Grupo {role_key} resetado para padrão", {})
        return {"ok": True, "reset": True}
    in_use = await db.users.count_documents({"tenant_id": tid, "role": role_key})
    if in_use:
        raise HTTPException(status_code=400, detail=f"{in_use} usuário(s) ainda usam este grupo. Mude-os antes de remover.")
    await db.role_templates.delete_one({"tenant_id": tid, "key": role_key})
    _invalidate_role_template_cache(tid)
    await write_audit(user, "delete", "role_template", role_key,
                       f"Grupo {role_key} removido", {})
    return {"ok": True, "deleted": True}


# ---------- Users ----------
class UserCreate(BaseModel):
    email: str; password: str; name: str
    role: str = "agent"
    permissions: Optional[List[str]] = None
    allowed_extensions: Optional[List[str]] = None  # None/[] = vê todos os ramais
    active: bool = True
    agent_id: Optional[str] = None
    # Provisionamento opcional no FusionPBX
    provision_extension: bool = False     # cria ramal SIP
    extension_number: Optional[str] = None  # ex: "1001"
    extension_sip_password: Optional[str] = None  # vazio = gerada
    provision_pbx_user: bool = False      # cria login web no FusionPBX
    pbx_password: Optional[str] = None
    provision_call_center_agent: bool = False  # cria call_center_agent
    cc_agent_id: Optional[str] = None     # login do agente (default = extension)
    queue_uuids: List[str] = []           # filas para vincular agente
    @validator("email")
    def _e(cls, v): return validate_email_str(v)

class UserUpdate(BaseModel):
    name: Optional[str] = None
    role: Optional[str] = None
    password: Optional[str] = None
    permissions: Optional[List[str]] = None
    allowed_extensions: Optional[List[str]] = None  # None = não altera, [] = limpar (ver todos)
    active: Optional[bool] = None
    agent_id: Optional[str] = None

def _serialize_user(u: dict, tenant_template: Optional[Dict[str, List[str]]] = None) -> dict:
    role = u.get("role", "agent")
    if u.get("permissions") is not None:
        eff = u["permissions"]
    elif tenant_template and role in tenant_template:
        eff = tenant_template[role]
    else:
        eff = DEFAULT_PERMISSIONS_BY_ROLE.get(role, [])
    return {
        "id": u["id"], "email": u["email"], "name": u.get("name", ""),
        "role": role, "tenant_id": u.get("tenant_id"),
        "permissions": eff,
        "is_custom_permissions": u.get("permissions") is not None,
        "allowed_extensions": u.get("allowed_extensions") or [],
        "active": u.get("active", True), "agent_id": u.get("agent_id"),
        "created_at": u.get("created_at"),
    }


def allowed_extensions_for(user: dict) -> Optional[set]:
    """Retorna None (sem restrição) ou um set com os ramais permitidos.
    Admins e super_admin sempre veem tudo."""
    if user.get("role") in ("super_admin", "admin"):
        return None
    allowed = user.get("allowed_extensions") or []
    if not allowed:
        return None
    return {str(x) for x in allowed}


async def allowed_agent_ids_for(user: dict) -> Optional[set]:
    """Retorna None (sem restrição) ou um set de agent_id (db.agents.id) cujos
    ramais estão na whitelist do usuário. Útil para filtrar recordings, reports
    e calls que se referem a agent_id."""
    allowed = allowed_extensions_for(user)
    if allowed is None:
        return None
    f = tenant_filter(user)
    docs = await db.agents.find(
        {**f, "extension": {"$in": list(allowed)}},
        {"_id": 0, "id": 1}).to_list(500)
    return {d["id"] for d in docs if d.get("id")}

@api.get("/users")
async def list_users(user: dict = Depends(require_permission("users.manage"))):
    f = tenant_filter(user)
    docs = await db.users.find({**f, "role": {"$ne": "super_admin"}}, {"_id": 0, "password_hash": 0}).sort("created_at", 1).to_list(500)
    tid = tenant_scope(user)
    tpl = await _load_tenant_role_templates(tid) if tid else {}
    return {"users": [_serialize_user(u, tpl) for u in docs]}

@api.post("/users")
async def create_user(body: UserCreate, user: dict = Depends(require_permission("users.manage"))):
    tid = await require_tenant_or_super(user)
    # Validar role: built-in OU custom existente no tenant
    valid_roles = set(BUILTIN_ROLES)
    custom_roles = await db.role_templates.distinct("key", {"tenant_id": tid})
    valid_roles.update(custom_roles)
    if body.role not in valid_roles:
        raise HTTPException(status_code=400, detail=f"Papel inválido. Use um destes: {sorted(valid_roles)}")
    tenant = await db.tenants.find_one({"id": tid})
    cnt = await db.users.count_documents({"tenant_id": tid})
    if tenant and cnt >= tenant.get("max_users", 999):
        raise HTTPException(status_code=400, detail=f"Limite de {tenant['max_users']} usuários atingido para este tenant")
    email = body.email.lower()
    if await db.users.find_one({"tenant_id": tid, "email": email}):
        raise HTTPException(status_code=400, detail="Email já cadastrado neste tenant")
    if body.permissions is not None:
        invalid = [p for p in body.permissions if p not in {x["key"] for x in ALL_PERMISSIONS}]
        if invalid: raise HTTPException(status_code=400, detail=f"Permissões inválidas: {invalid}")

    # Provisionamento FusionPBX (opcional)
    provisioned: Dict[str, Any] = {}
    fpbx_client = None
    fpbx_settings = None
    needs_fpbx = body.provision_extension or body.provision_pbx_user or body.provision_call_center_agent
    if needs_fpbx:
        try:
            fpbx_client, fpbx_settings = await _get_db_client(tid)
        except HTTPException as e:
            raise HTTPException(status_code=400, detail=f"Provisionamento FPBX requer modo PostgreSQL: {e.detail}")
        if not body.extension_number and (body.provision_extension or body.provision_call_center_agent):
            raise HTTPException(status_code=400, detail="Número do ramal é obrigatório para provisionamento")

    sip_password = body.extension_sip_password or _gen_pwd(12)
    pbx_password = body.pbx_password or _gen_pwd(10)
    cc_agent_id = body.cc_agent_id or str(body.extension_number or "")
    voxyra_agent_id_for_link = body.agent_id

    try:
        if body.provision_extension:
            ext_res = await fpbx_client.provision_extension(
                extension=str(body.extension_number), sip_password=sip_password,
                caller_id_name=body.name, caller_id_number=str(body.extension_number),
                description=f"Voxyra · {body.name}",
            )
            provisioned["extension"] = {**ext_res, "sip_password": sip_password}

        if body.provision_pbx_user:
            try:
                user_res = await fpbx_client.provision_user(
                    username=cc_agent_id or email.split("@")[0],
                    password_hash=pbx_password,
                )
                provisioned["pbx_user"] = {**user_res, "password": pbx_password}
                if "extension" in provisioned:
                    await fpbx_client.link_extension_to_user(
                        provisioned["extension"]["extension_uuid"],
                        user_res["user_uuid"],
                    )
            except FusionPBXDBError as e:
                logger.warning("Falha pbx_user: %s", e)
                provisioned["pbx_user_error"] = str(e)

        if body.provision_call_center_agent:
            ag_res = await fpbx_client.provision_call_center_agent(
                agent_name=body.name, agent_id=cc_agent_id,
                extension=str(body.extension_number),
                domain_name=fpbx_settings.get("domain_name") or "",
            )
            provisioned["call_center_agent"] = ag_res
            for qid in body.queue_uuids or []:
                try:
                    await fpbx_client.assign_agent_to_queue(
                        ag_res["call_center_agent_uuid"], qid,
                    )
                except Exception as e:
                    logger.warning("Falha vincular fila %s: %s", qid, e)
            # cria entidade agent local linkada
            voxyra_agent_uuid = str(uuid.uuid4())
            await db.agents.insert_one({
                "id": voxyra_agent_uuid, "tenant_id": tid,
                "external_id": ag_res["call_center_agent_uuid"],
                "name": body.name, "username": cc_agent_id,
                "extension": str(body.extension_number or ""),
                "email": email, "source": "call_center_agent",
                "avatar": AGENT_AVATARS[hash(cc_agent_id) % len(AGENT_AVATARS)],
                "status": "offline", "queues": [], "calls_handled": 0,
                "avg_handle_sec": 0, "csat": 0, "adherence_pct": 0,
                "created_at": datetime.now(timezone.utc).isoformat(),
                "updated_at": datetime.now(timezone.utc).isoformat(),
            })
            provisioned["voxyra_agent_id"] = voxyra_agent_uuid
            voxyra_agent_id_for_link = voxyra_agent_uuid

    except HTTPException:
        raise
    except FusionPBXDBError as e:
        raise HTTPException(status_code=502, detail=f"Falha no FusionPBX: {e}")

    uid = str(uuid.uuid4())
    doc = {
        "id": uid, "tenant_id": tid, "email": email, "name": body.name, "role": body.role,
        "password_hash": hash_password(body.password),
        "permissions": body.permissions, "active": body.active,
        "allowed_extensions": [str(x) for x in (body.allowed_extensions or [])],
        "agent_id": voxyra_agent_id_for_link,
        "created_at": datetime.now(timezone.utc).isoformat(),
    }
    await db.users.insert_one(doc)
    await write_audit(user, "create", "user", uid, f"{body.name} <{email}>", {
        "role": body.role, "tenant_id": tid,
        "permissions_mode": "custom" if body.permissions is not None else "default",
        "provisioned": list(provisioned.keys()),
    })
    out = _serialize_user(doc, await _load_tenant_role_templates(tid))
    if provisioned:
        out["provisioned"] = provisioned
    return out

@api.patch("/users/{user_id}")
async def update_user(user_id: str, body: UserUpdate, user: dict = Depends(require_permission("users.manage"))):
    f = tenant_filter(user)
    target = await db.users.find_one({"id": user_id, **f})
    if not target: raise HTTPException(status_code=404, detail="Usuário não encontrado")
    update = {}; changes = {}
    if body.name is not None and body.name != target.get("name"):
        update["name"] = body.name; changes["name"] = {"from": target.get("name"), "to": body.name}
    if body.role is not None and body.role != target.get("role"):
        valid_roles = set(BUILTIN_ROLES)
        if target.get("tenant_id"):
            valid_roles.update(
                await db.role_templates.distinct("key", {"tenant_id": target["tenant_id"]}))
        if body.role not in valid_roles:
            raise HTTPException(status_code=400, detail=f"Papel inválido. Use um destes: {sorted(valid_roles)}")
        update["role"] = body.role; changes["role"] = {"from": target.get("role"), "to": body.role}
    if body.password:
        update["password_hash"] = hash_password(body.password); changes["password"] = "changed"
    if body.permissions is not None:
        invalid = [p for p in body.permissions if p not in {x["key"] for x in ALL_PERMISSIONS}]
        if invalid: raise HTTPException(status_code=400, detail=f"Permissões inválidas: {invalid}")
        update["permissions"] = body.permissions
        changes["permissions"] = {"count": len(body.permissions), "mode": "custom"}
    if body.allowed_extensions is not None:
        normalized = [str(x) for x in body.allowed_extensions]
        update["allowed_extensions"] = normalized
        changes["allowed_extensions"] = {"count": len(normalized)}
    if body.active is not None and body.active != target.get("active", True):
        update["active"] = body.active; changes["active"] = {"from": target.get("active", True), "to": body.active}
    if body.agent_id is not None and body.agent_id != target.get("agent_id"):
        update["agent_id"] = body.agent_id; changes["agent_id"] = {"from": target.get("agent_id"), "to": body.agent_id}
    if update:
        await db.users.update_one({"id": user_id}, {"$set": update})
        await write_audit(user, "update", "user", user_id, f"{target.get('name')} <{target.get('email')}>", changes)
    fresh = await db.users.find_one({"id": user_id}, {"_id": 0, "password_hash": 0})
    return _serialize_user(fresh)

@api.delete("/users/{user_id}")
async def delete_user(user_id: str, user: dict = Depends(require_permission("users.manage"))):
    if user["id"] == user_id: raise HTTPException(status_code=400, detail="Você não pode excluir a si mesmo")
    f = tenant_filter(user)
    target = await db.users.find_one({"id": user_id, **f})
    if not target: raise HTTPException(status_code=404, detail="Usuário não encontrado")
    if target.get("role") == "admin":
        cnt = await db.users.count_documents({"role": "admin", "tenant_id": target.get("tenant_id")})
        if cnt <= 1: raise HTTPException(status_code=400, detail="Não é possível remover o último administrador do tenant")
    await db.users.delete_one({"id": user_id})
    await write_audit(user, "delete", "user", user_id, f"{target.get('name')} <{target.get('email')}>", {"role": target.get("role")})
    return {"ok": True}


# ---------- Super-admin users (cross-tenant root accounts) ----------
class SuperAdminCreate(BaseModel):
    email: str
    password: str
    name: str
    active: bool = True
    @validator("email")
    def _e(cls, v): return validate_email_str(v)


class SuperAdminUpdate(BaseModel):
    name: Optional[str] = None
    password: Optional[str] = None
    active: Optional[bool] = None


@api.get("/super-admins")
async def list_super_admins(user: dict = Depends(require_super_admin())):
    docs = await db.users.find({"role": "super_admin"}, {"_id": 0, "password_hash": 0}).sort("created_at", 1).to_list(200)
    return {"users": [_serialize_user(u) for u in docs]}


@api.post("/super-admins")
async def create_super_admin(body: SuperAdminCreate, user: dict = Depends(require_super_admin())):
    email = body.email.lower()
    if await db.users.find_one({"email": email, "role": "super_admin"}):
        raise HTTPException(status_code=400, detail="Email já cadastrado como super admin")
    if len(body.password) < 8:
        raise HTTPException(status_code=400, detail="Senha mínima 8 caracteres")
    doc = {
        "id": str(uuid.uuid4()),
        "tenant_id": None,
        "email": email,
        "name": body.name.strip() or email,
        "role": "super_admin",
        "permissions": None,
        "active": bool(body.active),
        "agent_id": None,
        "password_hash": hash_password(body.password),
        "created_at": datetime.now(timezone.utc).isoformat(),
    }
    await db.users.insert_one(doc)
    await write_audit(user, "create", "super_admin", doc["id"], f"{doc['name']} <{email}>", {"active": doc["active"]})
    fresh = await db.users.find_one({"id": doc["id"]}, {"_id": 0, "password_hash": 0})
    return _serialize_user(fresh)


@api.put("/super-admins/{user_id}")
async def update_super_admin(user_id: str, body: SuperAdminUpdate,
                              user: dict = Depends(require_super_admin())):
    target = await db.users.find_one({"id": user_id, "role": "super_admin"})
    if not target:
        raise HTTPException(status_code=404, detail="Super admin não encontrado")
    update: Dict[str, Any] = {}
    changes: Dict[str, Any] = {}
    if body.name is not None and body.name != target.get("name"):
        update["name"] = body.name.strip()
        changes["name"] = {"from": target.get("name"), "to": body.name.strip()}
    if body.password:
        if len(body.password) < 8:
            raise HTTPException(status_code=400, detail="Senha mínima 8 caracteres")
        update["password_hash"] = hash_password(body.password)
        changes["password"] = "changed"
    if body.active is not None and body.active != target.get("active", True):
        # Não permite desativar o último super admin ativo
        if not body.active:
            others_active = await db.users.count_documents(
                {"role": "super_admin", "active": True, "id": {"$ne": user_id}})
            if others_active == 0:
                raise HTTPException(status_code=400,
                    detail="Não é possível desativar o último super admin ativo")
        update["active"] = body.active
        changes["active"] = {"from": target.get("active", True), "to": body.active}
    if update:
        await db.users.update_one({"id": user_id}, {"$set": update})
        await write_audit(user, "update", "super_admin", user_id,
                           f"{target.get('name')} <{target.get('email')}>", changes)
    fresh = await db.users.find_one({"id": user_id}, {"_id": 0, "password_hash": 0})
    return _serialize_user(fresh)


@api.delete("/super-admins/{user_id}")
async def delete_super_admin(user_id: str, user: dict = Depends(require_super_admin())):
    if user["id"] == user_id:
        raise HTTPException(status_code=400, detail="Você não pode excluir a si mesmo")
    target = await db.users.find_one({"id": user_id, "role": "super_admin"})
    if not target:
        raise HTTPException(status_code=404, detail="Super admin não encontrado")
    others_active = await db.users.count_documents(
        {"role": "super_admin", "active": True, "id": {"$ne": user_id}})
    if others_active == 0:
        raise HTTPException(status_code=400,
            detail="Não é possível remover o último super admin ativo")
    await db.users.delete_one({"id": user_id})
    await write_audit(user, "delete", "super_admin", user_id,
                       f"{target.get('name')} <{target.get('email')}>", {})
    return {"ok": True}


@api.get("/audit-logs")
async def list_audit_logs(
    user: dict = Depends(require_permission("users.manage")),
    target_type: Optional[str] = None, action: Optional[str] = None,
    actor_id: Optional[str] = None, limit: int = Query(200, le=1000),
):
    q = {**tenant_filter(user)}
    if target_type: q["target_type"] = target_type
    if action: q["action"] = action
    if actor_id: q["actor_id"] = actor_id
    docs = await db.audit_logs.find(q, {"_id": 0}).sort("created_at", -1).to_list(limit)
    return {"logs": docs}

# ---------- Helpers / Constants ----------
DISPOSITION_LABELS = {"answered": "Atendida", "missed": "Perdida", "abandoned": "Abandonada", "voicemail": "Correio de Voz"}
DIRECTION_LABELS = {"inbound": "Entrada", "outbound": "Saída"}
ABANDON_LABELS = {"agent_loss": "Abandono do Agente", "queue_abandon": "Cliente na Fila"}
SAMPLE_AUDIO = "https://www.soundhelix.com/examples/mp3/SoundHelix-Song-1.mp3"
STATUSES = ["online", "incall", "paused", "offline"]
CALL_DIR = ["inbound", "outbound"]
DISPOSITIONS = ["answered", "missed", "abandoned", "voicemail"]
AGENT_AVATARS = [
    "https://images.unsplash.com/photo-1770058428276-7ca3d2f98568?crop=entropy&cs=srgb&fm=jpg&ixid=M3w4NTYxODF8MHwxfHNlYXJjaHwyfHxjYWxsJTIwY2VudGVyJTIwYWdlbnQlMjBwcm9mZXNzaW9uYWwlMjBwb3J0cmFpdCUyMHdoaXRlJTIwYmFja2dyb3VuZHxlbnwwfHx8fDE3Nzc3NDcwODh8MA&ixlib=rb-4.1.0&q=85",
    "https://images.unsplash.com/photo-1612276529418-52e6ad86ee1d?crop=entropy&cs=srgb&fm=jpg&ixid=M3w4NTYxODF8MHwxfHNlYXJjaHw0fHxjYWxsJTIwY2VudGVyJTIwYWdlbnQlMjBwcm9mZXNzaW9uYWwlMjBwb3J0cmFpdCUyMHdoaXRlJTIwYmFja2dyb3VuZHxlbnwwfHx8fDE3Nzc3NDcwODh8MA&ixlib=rb-4.1.0&q=85",
    "https://images.unsplash.com/photo-1712744626457-3ffa4ba32c8c?crop=entropy&cs=srgb&fm=jpg&ixid=M3w4NTYxODF8MHwxfHNlYXJjaHwxfHxjYWxsJTIwY2VudGVyJTIwYWdlbnQlMjBwcm9mZXNzaW9uYWwlMjBwb3J0cmFpdCUyMHdoaXRlJTIwYmFja2dyb3VuZHxlbnwwfHx8fDE3Nzc3NDcwODh8MA&ixlib=rb-4.1.0&q=85",
    "https://images.unsplash.com/photo-1685688739798-bce206ab6b42?crop=entropy&cs=srgb&fm=jpg&ixid=M3w4NTYxODF8MHwxfHNlYXJjaHwzfHxjYWxsJTIwY2VudGVyJTIwYWdlbnQlMjBwcm9mZXNzaW9uYWwlMjBwb3J0cmFpdCUyMHdoaXRlJTIwYmFja2dyb3VuZHxlbnwwfHx8fDE3Nzc3NDcwODh8MA&ixlib=rb-4.1.0&q=85",
]

def fmt_br_datetime(iso: Optional[str]) -> str:
    if not iso: return "—"
    try: return datetime.fromisoformat(iso).strftime("%d/%m/%Y %H:%M")
    except Exception: return iso

# ---------- Dashboard endpoints ----------
@api.get("/dashboard/stats")
async def dashboard_stats(user: dict = Depends(require_permission("dashboard.view"))):
    f = tenant_filter(user)
    tid = tenant_scope(user)
    total_agents = await db.agents.count_documents(f)
    online_agents = await db.agents.count_documents({**f, "status": {"$in": ["online", "incall", "paused"]}})
    incall = await db.agents.count_documents({**f, "status": "incall"})
    today = datetime.now(timezone.utc).replace(hour=0, minute=0, second=0, microsecond=0).isoformat()
    answered = await db.calls.count_documents({**f, "disposition": "answered", "started_at": {"$gte": today}})
    missed = await db.calls.count_documents({**f, "disposition": {"$in": ["missed", "abandoned"]}, "started_at": {"$gte": today}})
    queues = await db.queues.find(f, {"_id": 0}).to_list(50)
    waiting = sum(q.get("waiting", 0) for q in queues)
    avg_wait = int(sum(q.get("avg_wait_sec", 0) for q in queues) / max(len(queues), 1))
    # Live waiting via ESL (overrides cached if available)
    if tid:
        s = await db.fusionpbx_settings.find_one({"tenant_id": tid})
        if s and s.get("enabled") and s.get("esl_host"):
            try:
                esl = FreeSwitchESL(host=s["esl_host"], port=int(s.get("esl_port") or 8021),
                                    password=s.get("esl_password") or "ClueCon",
                                    timeout=float(s.get("esl_timeout") or 5.0))
                rows = await esl.show_channels()
                queue_exts = {q.get("extension"): q for q in queues if q.get("extension")}
                live_waiting = 0
                live_incall_exts = set()
                rows_by_uuid = {(r.get("uuid") or ""): r for r in (rows or [])}
                child_uuids = set()
                for r in (rows or []):
                    bu = r.get("b_uuid") or ""
                    if bu in rows_by_uuid: child_uuids.add(bu)
                for r in (rows or []):
                    if (r.get("uuid") or "") in child_uuids: continue
                    dest = str(r.get("dest") or "")
                    state = (r.get("callstate") or "").upper()
                    # Em fila = destino bate com queue.extension e ainda não foi bridged (sem b_uuid)
                    if dest in queue_exts and not r.get("b_uuid"):
                        live_waiting += 1
                # Atualiza waiting nas filas (in-memory, não persiste)
                queue_waiting = {qext: 0 for qext in queue_exts}
                for r in (rows or []):
                    if (r.get("uuid") or "") in child_uuids: continue
                    dest = str(r.get("dest") or "")
                    if dest in queue_waiting and not r.get("b_uuid"):
                        queue_waiting[dest] += 1
                # Persiste contagem ao vivo nas filas
                for ext, count in queue_waiting.items():
                    q = queue_exts[ext]
                    await db.queues.update_one({"id": q["id"]}, {"$set": {"waiting": count}})
                waiting = live_waiting
            except Exception as e:
                logger.warning("dashboard: ESL waiting probe falhou: %s", e)
    cutoff = (datetime.now(timezone.utc) - timedelta(hours=23)).isoformat()
    cur = db.calls.find({**f, "started_at": {"$gte": cutoff}}, {"_id": 0, "started_at": 1, "disposition": 1})
    buckets = {h: {"hour": f"{h:02d}h", "answered": 0, "missed": 0} for h in range(24)}
    async for c in cur:
        try:
            h = datetime.fromisoformat(c["started_at"]).hour
            if c["disposition"] == "answered": buckets[h]["answered"] += 1
            elif c["disposition"] in ("missed", "abandoned"): buckets[h]["missed"] += 1
        except Exception: pass
    return {
        "total_agents": total_agents, "online_agents": online_agents, "incall_agents": incall,
        "answered_today": answered, "missed_today": missed,
        "waiting_in_queue": waiting, "avg_wait_sec": avg_wait,
        "hourly": [buckets[h] for h in range(24)],
    }

@api.get("/dashboard/abandoned")
async def dashboard_abandoned(user: dict = Depends(require_permission("dashboard.view"))):
    f = tenant_filter(user)
    now = datetime.now(timezone.utc)
    hour_buckets = {h: {"label": f"{h:02d}h", "agent_loss": 0, "queue_abandon": 0} for h in range(24)}
    day_buckets = []
    for i in range(6, -1, -1):
        d = (now - timedelta(days=i)).replace(hour=0, minute=0, second=0, microsecond=0)
        day_buckets.append({"key": d.isoformat(), "label": d.strftime("%a %d/%m"), "agent_loss": 0, "queue_abandon": 0})
    week_buckets = []
    for i in range(3, -1, -1):
        day = now - timedelta(days=now.weekday() + 7 * i)
        day = day.replace(hour=0, minute=0, second=0, microsecond=0)
        week_buckets.append({"key": day.isoformat(), "label": f"Sem {day.strftime('%d/%m')}", "agent_loss": 0, "queue_abandon": 0})
    cutoff_hour = (now - timedelta(hours=23)).isoformat()
    cutoff_day = (now - timedelta(days=6)).replace(hour=0, minute=0, second=0, microsecond=0)
    cutoff_week = (now - timedelta(days=now.weekday() + 7 * 3)).replace(hour=0, minute=0, second=0, microsecond=0)
    earliest = min(cutoff_week, datetime.fromisoformat(cutoff_hour))
    cursor = db.calls.find({**f, "disposition": {"$in": ["missed", "abandoned"]}, "started_at": {"$gte": earliest.isoformat()}},
                           {"_id": 0, "started_at": 1, "abandonment_type": 1, "queue_name": 1})
    total_hour = {"agent_loss": 0, "queue_abandon": 0}
    total_day = {"agent_loss": 0, "queue_abandon": 0}
    total_week = {"agent_loss": 0, "queue_abandon": 0}
    by_queue: Dict[str, dict] = {}
    async for c in cursor:
        try: dt = datetime.fromisoformat(c["started_at"])
        except Exception: continue
        at = c.get("abandonment_type")
        if at not in ("agent_loss", "queue_abandon"): continue
        if dt.isoformat() >= cutoff_hour:
            hour_buckets[dt.hour][at] += 1; total_hour[at] += 1
        day_key = dt.replace(hour=0, minute=0, second=0, microsecond=0)
        if day_key >= cutoff_day:
            for b in day_buckets:
                if b["key"] == day_key.isoformat(): b[at] += 1; total_day[at] += 1; break
        week_start = (dt - timedelta(days=dt.weekday())).replace(hour=0, minute=0, second=0, microsecond=0)
        if week_start >= cutoff_week:
            for b in week_buckets:
                if b["key"] == week_start.isoformat(): b[at] += 1; total_week[at] += 1; break
        if day_key >= cutoff_day:
            qn = c.get("queue_name", "—")
            if qn not in by_queue: by_queue[qn] = {"queue": qn, "agent_loss": 0, "queue_abandon": 0}
            by_queue[qn][at] += 1
    return {
        "by_hour": [hour_buckets[h] for h in range(24)],
        "by_day": day_buckets, "by_week": week_buckets,
        "totals": {"last_24h": total_hour, "last_7d": total_day, "last_4w": total_week},
        "by_queue": sorted(by_queue.values(), key=lambda x: x["agent_loss"] + x["queue_abandon"], reverse=True),
    }

@api.get("/realtime/calls")
async def realtime_calls(user: dict = Depends(require_permission("realtime.view"))):
    f = tenant_filter(user)
    tid = tenant_scope(user)
    allowed_ext = allowed_extensions_for(user)
    def _filter_calls(calls: List[Dict[str, Any]]) -> List[Dict[str, Any]]:
        if allowed_ext is None:
            return calls
        return [c for c in calls if str(c.get("agent_extension") or "") in allowed_ext]
    if tid:
        s = await db.fusionpbx_settings.find_one({"tenant_id": tid})
        if s and s.get("enabled"):
            # 1) Try ESL first (most reliable for live channels)
            if s.get("esl_host"):
                try:
                    esl = FreeSwitchESL(
                        host=s["esl_host"], port=int(s.get("esl_port") or 8021),
                        password=s.get("esl_password") or "ClueCon",
                        timeout=float(s.get("esl_timeout") or 5.0),
                    )
                    rows = await esl.show_channels()
                    if rows is not None:
                        agents_db = await db.agents.find(f, {"_id": 0}).to_list(500)
                        ext_to_agent = {a.get("extension"): a for a in agents_db if a.get("extension")}
                        queues_db = await db.queues.find(f, {"_id": 0}).to_list(500)
                        ext_to_queue = {q.get("extension"): q for q in queues_db if q.get("extension")}
                        domain_name = (s.get("domain_name") or "").strip().lower()
                        # Group bridged calls (b-leg has a `b_uuid` pointing to a-leg)
                        # Build map: uuid -> row, and a set of "child" uuids to skip
                        rows_by_uuid = {(r.get("uuid") or ""): r for r in rows}
                        child_uuids = set()
                        for r in rows:
                            buid = r.get("b_uuid") or ""
                            if buid and buid in rows_by_uuid:
                                child_uuids.add(buid)
                        calls = []
                        for raw in rows:
                            uid = raw.get("uuid") or ""
                            if uid in child_uuids:
                                continue  # skip b-leg of a bridged call (we'll show the parent)
                            n = normalize_esl_channel(raw)
                            # Domain filter: check presence_id, context, dest_addr, sip_to_user
                            if domain_name:
                                blob = " ".join([
                                    str(raw.get("presence_id") or ""),
                                    str(raw.get("context") or ""),
                                    str(raw.get("dest") or ""),
                                    str(raw.get("sip_to_user") or ""),
                                    str(raw.get("sip_from_user") or ""),
                                ]).lower()
                                # if no domain match AND no extension/queue match, skip
                                ext_dest = str(raw.get("dest") or "")
                                in_known = ext_dest in ext_to_agent or ext_dest in ext_to_queue
                                if domain_name not in blob and not in_known:
                                    continue
                            ext = n["destination_number"] or n["caller_id_number"]
                            ag = ext_to_agent.get(ext) or {}
                            queue = ext_to_queue.get(ext) or {}
                            # se está numa fila, status = "queued"
                            in_queue = bool(queue and not ag)
                            elapsed = 0
                            if n["created_epoch"]:
                                elapsed = int(time.time() - n["created_epoch"])
                            # check b-leg for richer agent info if this is parent
                            buid = raw.get("b_uuid") or ""
                            b_leg = rows_by_uuid.get(buid) if buid else None
                            if b_leg and not ag:
                                b_dest = str(b_leg.get("dest") or b_leg.get("destination_number") or "")
                                if b_dest in ext_to_agent:
                                    ag = ext_to_agent[b_dest]
                                    ext = b_dest
                            status = "incall" if (n["answer_state"] == "answered" and ag) \
                                     else ("queued" if in_queue else "ringing")
                            calls.append({
                                "id": uid or str(uuid.uuid4()),
                                "agent_name": ag.get("name") or (queue.get("name") and f"Fila: {queue['name']}") or n["caller_id_name"] or "—",
                                "agent_extension": ext or "—",
                                "agent_avatar": ag.get("avatar"),
                                "queue_name": queue.get("name") or (b_leg and ext_to_queue.get(str(b_leg.get("dest") or "")) or {}).get("name") or "—",
                                "caller_number": n["caller_id_number"] or "—",
                                "direction": n["direction"],
                                "elapsed_sec": elapsed,
                                "status": status,
                            })
                        return {"calls": _filter_calls(calls), "source": "esl"}
                except FreeSwitchESLError as e:
                    logger.warning("ESL falhou, tentando fallback: %s", e)

            # 2) Fallback DB or REST (v_channels — pode não existir)
            ctype = s.get("connection_type") or "rest"
            client = None
            ClientErr: type = Exception
            try:
                if ctype == "db" and s.get("db_host"):
                    client = FusionPBXDBClient(
                        host=s["db_host"], port=int(s.get("db_port") or 5432),
                        database=s.get("db_name") or "fusionpbx",
                        username=s.get("db_username") or "", password=s.get("db_password") or "",
                        domain_uuid=s.get("domain_uuid"), ssl=bool(s.get("db_ssl")),
                    )
                    ClientErr = FusionPBXDBError
                elif ctype == "rest" and s.get("base_url"):
                    client = FusionPBXClient(
                        base_url=s["base_url"], api_key=s.get("api_key"),
                        username=s.get("username"), password=s.get("password"),
                        domain_uuid=s.get("domain_uuid"), domain_name=s.get("domain_name"),
                        verify_ssl=bool(s.get("verify_ssl", True)),
                    )
                    ClientErr = FusionPBXError
                if client:
                    raw = await client.list_active_calls()
                    if raw:
                        agents_db = await db.agents.find(f, {"_id": 0}).to_list(500)
                        ext_to_agent = {a.get("extension"): a for a in agents_db if a.get("extension")}
                        calls = []
                        for c in raw:
                            ext = str(c.get("destination_number") or c.get("destination") or c.get("extension") or "")
                            ag = ext_to_agent.get(ext) or {}
                            answer_state = (c.get("answer_state") or c.get("channel_state") or "").lower()
                            is_answered = "answered" in answer_state or "exchange_media" in answer_state or answer_state == "cs_execute"
                            elapsed = 0
                            try:
                                if c.get("created_epoch"):
                                    elapsed = int(time.time() - int(c["created_epoch"]))
                                elif c.get("duration"):
                                    elapsed = int(c["duration"])
                            except Exception:
                                pass
                            calls.append({
                                "id": c.get("uuid") or c.get("call_uuid") or str(uuid.uuid4()),
                                "agent_name": ag.get("name") or c.get("caller_id_name") or "—",
                                "agent_extension": ext or "—",
                                "agent_avatar": ag.get("avatar"),
                                "queue_name": c.get("queue_name") or "—",
                                "caller_number": c.get("caller_id_number") or c.get("cid_num") or "—",
                                "direction": c.get("direction") if c.get("direction") in ("inbound", "outbound") else "inbound",
                                "elapsed_sec": elapsed,
                                "status": "incall" if is_answered else "ringing",
                            })
                        return {"calls": _filter_calls(calls), "source": "fusionpbx"}
            except ClientErr as e:
                logger.warning("realtime/calls fallback %s falhou: %s", ctype, e)
            except Exception as e:
                logger.warning("realtime/calls erro inesperado: %s", e)
    # Final fallback
    agents = await db.agents.find({**f, "status": "incall"}, {"_id": 0}).to_list(100)
    queues = {q["id"]: q for q in await db.queues.find(f, {"_id": 0}).to_list(100)}
    active = []
    for a in agents:
        qid = a["queues"][0] if a.get("queues") else None
        q = queues.get(qid, {})
        active.append({
            "id": str(uuid.uuid4()),
            "agent_name": a["name"], "agent_extension": a["extension"], "agent_avatar": a.get("avatar"),
            "queue_name": q.get("name", "—"),
            "caller_number": "—",
            "direction": "inbound", "elapsed_sec": 0, "status": "incall",
        })
    return {"calls": _filter_calls(active), "source": "local"}

@api.post("/agents/cleanup-extensions")
async def cleanup_extension_agents(user: dict = Depends(require_super_admin())):
    """Remove documentos de agente que vieram de v_extensions (source=extension)
    quando existem agentes reais (source=call_center_agent) no tenant. Útil
    para limpar a página de Agentes que lista ramais por engano."""
    res = {"removed_per_tenant": {}}
    tids = await db.agents.distinct("tenant_id")
    for tid in tids:
        if not tid:
            continue
        has_real = await db.agents.count_documents(
            {"tenant_id": tid, "source": "call_center_agent"})
        if has_real == 0:
            continue
        deleted = await db.agents.delete_many(
            {"tenant_id": tid, "source": "extension"})
        if deleted.deleted_count:
            res["removed_per_tenant"][tid] = deleted.deleted_count
    await write_audit(user, "cleanup", "agents", "all",
                       "Removido ramais soltos da lista de agentes", res)
    return {"ok": True, **res}


@api.get("/agents")
async def list_agents(user: dict = Depends(require_permission("agents.view")),
                       include_extensions: bool = False):
    """Lista agentes reais do Call Center (source != "extension").
    Ramais SIP ficam na página dedicada /extensions. Para incluir ramais
    nesta lista, passe ?include_extensions=true."""
    f = tenant_filter(user)
    if not include_extensions:
        f = {**f, "source": {"$ne": "extension"}}
    allowed_ext = allowed_extensions_for(user)
    if allowed_ext is not None:
        f = {**f, "extension": {"$in": list(allowed_ext)}}
    items = await db.agents.find(f, {"_id": 0}).to_list(500)
    # Enrich with missed_count and avg_wait_sec (last 24h)
    cutoff_24h = (datetime.now(timezone.utc) - timedelta(hours=24)).isoformat()
    pipeline = [
        {"$match": {**f, "started_at": {"$gte": cutoff_24h}, "agent_id": {"$ne": None}}},
        {"$group": {
            "_id": "$agent_id",
            "missed": {"$sum": {"$cond": [{"$in": ["$disposition", ["missed", "abandoned"]]}, 1, 0]}},
            "answered": {"$sum": {"$cond": [{"$eq": ["$disposition", "answered"]}, 1, 0]}},
            "avg_wait": {"$avg": "$wait_sec"},
        }},
    ]
    stats: Dict[str, Dict[str, int]] = {}
    async for r in db.calls.aggregate(pipeline):
        stats[r["_id"]] = {
            "missed_count": int(r.get("missed") or 0),
            "answered_count": int(r.get("answered") or 0),
            "avg_wait_sec": int(r.get("avg_wait") or 0),
        }
    for a in items:
        s = stats.get(a["id"], {})
        a["missed_count"] = s.get("missed_count", 0)
        a["answered_count"] = s.get("answered_count", a.get("calls_handled", 0))
        a["avg_wait_sec"] = s.get("avg_wait_sec", 0)
    return {"agents": items}

@api.get("/agents/{agent_id}")
async def get_agent(agent_id: str, user: dict = Depends(require_permission("agents.view"))):
    f = tenant_filter(user)
    a = await db.agents.find_one({"id": agent_id, **f}, {"_id": 0})
    if not a: raise HTTPException(status_code=404, detail="Agente não encontrado")
    recent = await db.calls.find({"agent_id": agent_id, **f}, {"_id": 0}).sort("started_at", -1).to_list(20)
    return {"agent": a, "recent_calls": recent}


VOXYRA_STATUS_TO_PBX = {
    "online": "Available",
    "available": "Available",
    "paused": "On Break",
    "break": "On Break",
    "offline": "Logged Out",
    "logged_out": "Logged Out",
}
VALID_VOXYRA_STATUSES = {"online", "paused", "offline"}


class AgentStatusReq(BaseModel):
    status: str  # "online" | "paused" | "offline"


class AgentEditReq(BaseModel):
    name: Optional[str] = None
    extension: Optional[str] = None
    voxyra_email: Optional[str] = None
    voxyra_password: Optional[str] = None
    sip_password: Optional[str] = None
    queue_uuids: Optional[List[str]] = None  # external_ids

@api.post("/agents/{agent_id}/pbx-force-logout")
async def pbx_force_logout(agent_id: str,
                            user: dict = Depends(require_permission("agents.edit"))):
    """Limpeza defensiva: força a remoção de TODOS os tiers e marca Logged Out
    para qualquer agente em memória do mod_callcenter cujo nome corresponda à
    extensão deste agente, em qualquer formato (`1001`, `1001@dominio`).
    Use quando o agente fica 'preso' como Available/Waiting no FusionPBX
    mesmo após o logout pelo Voxyra."""
    f = tenant_filter(user)
    agent = await db.agents.find_one({"id": agent_id, **f}, {"_id": 0})
    if not agent:
        raise HTTPException(status_code=404, detail="Agente não encontrado")
    tid = agent.get("tenant_id")
    s = await db.fusionpbx_settings.find_one({"tenant_id": tid}) or {}
    if not s.get("esl_host"):
        raise HTTPException(status_code=400, detail="ESL não configurado")
    ext = str(agent.get("extension") or "").strip()
    if not ext:
        raise HTTPException(status_code=400, detail="Agente sem ramal definido")
    domain = s.get("domain_name") or None
    esl = FreeSwitchESL(
        host=s["esl_host"], port=int(s.get("esl_port") or 8021),
        password=s.get("esl_password") or "ClueCon",
        timeout=float(s.get("esl_timeout") or 5.0),
    )
    result = await esl.callcenter_force_clear_by_extension(ext, domain=domain)
    # Reset DB state too
    await db.agents.update_one(
        {"id": agent_id},
        {"$set": {"status": "offline",
                  "pbx_status": "Logged Out",
                  "active_queues": [],
                  "status_changed_at": datetime.now(timezone.utc).isoformat()}})
    # Update Postgres v_call_center_agents.agent_status if DB mode
    if s.get("connection_type") == "db" and s.get("db_host") and agent.get("external_id"):
        try:
            client = FusionPBXDBClient(
                host=s["db_host"], port=int(s.get("db_port") or 5432),
                database=s.get("db_name") or "fusionpbx",
                username=s["db_username"], password=s.get("db_password") or "",
                domain_uuid=s.get("domain_uuid"), ssl=bool(s.get("db_ssl")),
            )
            try:
                await client.update_agent_status(agent["external_id"], "Logged Out")
            except Exception as e:
                logger.warning("force_logout: DB status update failed: %s", e)
        except Exception:
            pass
    await write_audit(user, "force_logout", "agent", agent_id,
                       f"Force logout PBX · {agent.get('name')} ({ext})", result)
    return {"ok": True, **result}


@api.get("/agents/{agent_id}/pbx-state")
async def get_agent_pbx_state(agent_id: str,
                                user: dict = Depends(require_permission("agents.view"))):
    """Diagnóstico: retorna o estado do agente em MEMÓRIA do mod_callcenter
    (status, state, contact, tiers ativos) e em PostgreSQL (v_call_center_agents).
    Útil para descobrir por que comandos de login/logoff de fila não estão
    surtindo efeito (geralmente porque o agent_id em memória difere do que o
    Voxyra está enviando)."""
    f = tenant_filter(user)
    agent = await db.agents.find_one({"id": agent_id, **f}, {"_id": 0})
    if not agent:
        raise HTTPException(status_code=404, detail="Agente não encontrado")
    tid = agent.get("tenant_id")
    s = await db.fusionpbx_settings.find_one({"tenant_id": tid}) or {}
    out: Dict[str, Any] = {
        "voxyra": {
            "id": agent.get("id"),
            "name": agent.get("name"),
            "extension": agent.get("extension"),
            "external_id": agent.get("external_id"),
            "source": agent.get("source"),
            "status": agent.get("status"),
            "pbx_status": agent.get("pbx_status"),
            "active_queues": agent.get("active_queues") or [],
        },
        "pbx_db": None,
        "esl_resolved_name": None,
        "esl_agent_in_memory": None,
        "esl_tiers_in_memory": [],
        "esl_all_agents_count": 0,
        "esl_errors": [],
    }
    # 1) PBX DB lookup
    if s.get("connection_type") == "db" and s.get("db_host") and agent.get("external_id"):
        try:
            client = FusionPBXDBClient(
                host=s["db_host"], port=int(s.get("db_port") or 5432),
                database=s.get("db_name") or "fusionpbx",
                username=s["db_username"], password=s.get("db_password") or "",
                domain_uuid=s.get("domain_uuid"), ssl=bool(s.get("db_ssl")),
            )
            conn = await client._connect()
            try:
                row = await conn.fetchrow(
                    """SELECT call_center_agent_uuid::text, agent_id, agent_name,
                              agent_status, agent_state, agent_contact, agent_type
                       FROM v_call_center_agents
                       WHERE call_center_agent_uuid = $1::uuid""",
                    agent["external_id"],
                )
                if row:
                    out["pbx_db"] = dict(row)
            finally:
                await conn.close()
        except Exception as e:
            out["esl_errors"].append(f"pbx_db: {e}")
    # 2) Resolve ESL agent name
    esl_name = await _pbx_resolve_agent_esl_name(tid, agent.get("external_id") or "")
    out["esl_resolved_name"] = esl_name
    # 3) ESL queries
    if s.get("esl_host"):
        try:
            esl = FreeSwitchESL(
                host=s["esl_host"], port=int(s.get("esl_port") or 8021),
                password=s.get("esl_password") or "ClueCon",
                timeout=float(s.get("esl_timeout") or 5.0),
            )
            try:
                all_agents = await esl.callcenter_agent_list()
                out["esl_all_agents_count"] = len(all_agents)
                if esl_name:
                    out["esl_agent_in_memory"] = next(
                        (a for a in all_agents if a.get("name") == esl_name), None)
            except Exception as e:
                out["esl_errors"].append(f"agent_list: {e}")
            try:
                all_tiers = await esl.callcenter_tier_list()
                if esl_name:
                    out["esl_tiers_in_memory"] = [
                        t for t in all_tiers if t.get("agent") == esl_name]
                else:
                    out["esl_tiers_in_memory"] = all_tiers[:20]
            except Exception as e:
                out["esl_errors"].append(f"tier_list: {e}")
        except Exception as e:
            out["esl_errors"].append(f"esl: {e}")
    return out


@api.post("/agents/{agent_id}/pbx-resync")
async def pbx_resync_agent(agent_id: str,
                            user: dict = Depends(require_permission("agents.edit"))):
    """Força re-sincronização do agente com a memória do mod_callcenter.
    Útil quando alguém edita contato/status direto no GUI do FusionPBX e o
    `reload mod_callcenter` não atualiza agentes já carregados em memória."""
    f = tenant_filter(user)
    agent = await db.agents.find_one({"id": agent_id, **f}, {"_id": 0})
    if not agent:
        raise HTTPException(status_code=404, detail="Agente não encontrado")
    if not agent.get("external_id"):
        raise HTTPException(status_code=400, detail="Agente sem external_id")
    tid = agent.get("tenant_id")
    result = await _pbx_resync_agent_from_db(tid, agent["external_id"])
    try:
        s = await db.fusionpbx_settings.find_one({"tenant_id": tid}) or {}
        if s.get("connection_type") == "db" and s.get("db_host"):
            pg = FusionPBXDBClient(
                host=s["db_host"], port=int(s.get("db_port") or 5432),
                database=s.get("db_name") or "fusionpbx",
                username=s["db_username"], password=s.get("db_password") or "",
                domain_uuid=s.get("domain_uuid"), ssl=bool(s.get("db_ssl")),
            )
            tiers = await pg.list_agent_tiers(agent["external_id"])
            qnames = []
            for t in tiers:
                qn = t.get("queue_name") or ""
                if qn:
                    if s.get("domain_name") and "@" not in qn:
                        qn = f"{qn}@{s['domain_name']}"
                    qnames.append(qn)
            await _pbx_apply_agent_live(
                tid, agent["external_id"],
                clear_tiers=True, add_tier_queues=qnames,
            )
            result["tiers"] = qnames
    except Exception as e:
        result["errors"] = (result.get("errors") or []) + [f"tiers: {e}"]
    await write_audit(user, "resync", "agent", agent_id,
                       f"Resync PBX: {agent.get('name')}", result)
    return {"ok": True, **result}


@api.get("/extensions")
async def list_extensions(user: dict = Depends(get_current_user)):
    """Lista ramais SIP do FusionPBX (v_extensions). Lê em tempo real do
    FusionPBX via PostgreSQL (preferido) ou REST, respeitando a
    configuração do tenant em fusionpbx_settings."""
    tid = tenant_scope(user)
    if not tid:
        return {"extensions": [], "registrations": {}}
    s = await db.fusionpbx_settings.find_one({"tenant_id": tid}) or {}
    if not s.get("enabled"):
        return {"extensions": [], "registrations": {},
                "warning": "Integração FusionPBX desabilitada para este tenant"}
    ctype = s.get("connection_type") or "rest"
    try:
        if ctype == "db":
            if not s.get("db_host") or not s.get("db_username"):
                return {"extensions": [], "registrations": {},
                        "warning": "FusionPBX PostgreSQL não configurado"}
            client = FusionPBXDBClient(
                host=s["db_host"], port=int(s.get("db_port") or 5432),
                database=s.get("db_name") or "fusionpbx",
                username=s["db_username"], password=s.get("db_password") or "",
                domain_uuid=s.get("domain_uuid"), ssl=bool(s.get("db_ssl")),
            )
            exts = await client.list_extensions()
        else:
            if not s.get("base_url"):
                return {"extensions": [], "registrations": {},
                        "warning": "FusionPBX REST não configurado"}
            custom_paths = {k.replace("path_", ""): s[k] for k in ("path_extensions",) if s.get(k)}
            client = FusionPBXClient(
                base_url=s["base_url"], api_key=s.get("api_key"),
                username=s.get("username"), password=s.get("password"),
                domain_uuid=s.get("domain_uuid"), domain_name=s.get("domain_name"),
                verify_ssl=bool(s.get("verify_ssl", True)),
                custom_paths=custom_paths,
            )
            exts = await client.list_extensions()
    except (FusionPBXDBError, FusionPBXError) as e:
        raise HTTPException(status_code=502, detail=str(e))
    # Registrations via ESL (best-effort)
    regs: Dict[str, bool] = {}
    try:
        if s.get("esl_host"):
            esl = FreeSwitchESL(
                host=s["esl_host"], port=int(s.get("esl_port") or 8021),
                password=s.get("esl_password") or "ClueCon",
                timeout=float(s.get("esl_timeout") or 5.0),
            )
            out = await esl.api("sofia status profile internal reg")
            for line in (out or "").splitlines():
                if "User:" in line:
                    user_part = line.split("User:")[1].strip()
                    ext = user_part.split("@")[0]
                    regs[ext] = True
    except Exception as e:
        logger.warning("list_extensions: ESL regs falhou: %s", e)
    # Mark which agents are linked
    agents = await db.agents.find(
        {"tenant_id": tid, "extension": {"$ne": None}},
        {"_id": 0, "id": 1, "extension": 1, "name": 1, "source": 1,
         "status": 1, "pbx_status": 1, "queues": 1}).to_list(500)
    agent_by_ext: Dict[str, Dict[str, Any]] = {}
    for a in agents:
        ext_k = str(a.get("extension") or "")
        if not ext_k:
            continue
        # Prefer real Call Center Agents over extension-sourced records
        existing = agent_by_ext.get(ext_k)
        if existing and existing.get("source") == "call_center_agent":
            continue
        agent_by_ext[ext_k] = a
    # Apply per-user extension whitelist (None = unrestricted)
    allowed_ext = allowed_extensions_for(user)
    out = []
    seen_exts = set()
    for e in exts:
        ext = str(e.get("extension") or "")
        if allowed_ext is not None and ext not in allowed_ext:
            continue
        enabled_raw = e.get("enabled")
        if isinstance(enabled_raw, bool):
            enabled_val = enabled_raw
        else:
            enabled_val = str(enabled_raw).lower() in ("true", "t", "1", "yes")
        linked = agent_by_ext.get(ext) or {}
        is_agent = linked.get("source") == "call_center_agent"
        seen_exts.add(ext)
        out.append({
            "uuid": e.get("uuid") or e.get("extension_uuid"),
            "extension": ext,
            "caller_id_name": e.get("caller_id_name") or e.get("effective_caller_id_name"),
            "caller_id_number": e.get("caller_id_number") or e.get("effective_caller_id_number"),
            "enabled": enabled_val,
            "description": e.get("description"),
            "registered": bool(regs.get(ext)),
            "agent_name": linked.get("name"),
            "agent_id": linked.get("id") if is_agent else None,
            "is_agent": is_agent,
            "agent_status": linked.get("status") if is_agent else None,
            "agent_pbx_status": linked.get("pbx_status") if is_agent else None,
            "queues_count": len(linked.get("queues") or []) if is_agent else 0,
        })
    # Append Call Center Agents that don't have a matching SIP extension in v_extensions
    # (e.g., agent_contact format not parsed, or pure callcenter agents without phone)
    for a in agents:
        if a.get("source") != "call_center_agent":
            continue
        ext = str(a.get("extension") or "")
        if ext and ext in seen_exts:
            continue
        if allowed_ext is not None and ext and ext not in allowed_ext:
            continue
        out.append({
            "uuid": a.get("id"),
            "extension": ext or "—",
            "caller_id_name": a.get("name"),
            "caller_id_number": ext or None,
            "enabled": True,
            "description": "Agente Call Center (sem ramal SIP correspondente)",
            "registered": bool(regs.get(ext)) if ext else False,
            "agent_name": a.get("name"),
            "agent_id": a.get("id"),
            "is_agent": True,
            "agent_status": a.get("status"),
            "agent_pbx_status": a.get("pbx_status"),
            "queues_count": len(a.get("queues") or []),
        })
    return {"extensions": out}


@api.get("/agents/{agent_id}/linked-user")
async def get_agent_linked_user(agent_id: str,
                                  user: dict = Depends(require_permission("agents.edit"))):
    f = tenant_filter(user)
    agent = await db.agents.find_one({"id": agent_id, **f}, {"_id": 0, "id": 1, "tenant_id": 1})
    if not agent:
        raise HTTPException(status_code=404, detail="Agente não encontrado")
    u = await db.users.find_one(
        {"agent_id": agent_id, "tenant_id": agent["tenant_id"]},
        {"_id": 0, "id": 1, "email": 1, "name": 1, "active": 1})
    if not u:
        return {"email": None}
    return u


@api.put("/agents/{agent_id}")
async def update_agent(agent_id: str, body: AgentEditReq,
                       user: dict = Depends(require_permission("agents.edit"))):
    f = tenant_filter(user)
    agent = await db.agents.find_one({"id": agent_id, **f}, {"_id": 0})
    if not agent:
        raise HTTPException(status_code=404, detail="Agente não encontrado")
    tid = agent.get("tenant_id")
    update: Dict[str, Any] = {}
    changes: Dict[str, Any] = {}
    pbx_warnings: List[str] = []

    if body.name is not None and body.name.strip() and body.name != agent.get("name"):
        update["name"] = body.name.strip()
        changes["name"] = {"from": agent.get("name"), "to": body.name.strip()}

    new_ext = (body.extension or "").strip()
    if new_ext and new_ext != (agent.get("extension") or ""):
        if not new_ext.isdigit() or not (2 <= len(new_ext) <= 8):
            raise HTTPException(status_code=400, detail="Ramal inválido")
        # Garante ramal único no tenant
        dup = await db.agents.find_one(
            {"tenant_id": tid, "extension": new_ext, "id": {"$ne": agent_id}})
        if dup:
            raise HTTPException(status_code=400, detail=f"Ramal {new_ext} já está em uso")
        update["extension"] = new_ext
        changes["extension"] = {"from": agent.get("extension"), "to": new_ext}

    s = await db.fusionpbx_settings.find_one({"tenant_id": tid}) or {}
    pbx_db_ok = (s.get("connection_type") == "db" and s.get("db_host") and agent.get("external_id"))
    client = None
    if pbx_db_ok:
        try:
            client = FusionPBXDBClient(
                host=s["db_host"], port=int(s.get("db_port") or 5432),
                database=s.get("db_name") or "fusionpbx",
                username=s["db_username"], password=s.get("db_password") or "",
                domain_uuid=s.get("domain_uuid"), ssl=bool(s.get("db_ssl")),
            )
        except Exception as e:
            pbx_warnings.append(f"PBX init: {e}")

    # Voxyra user (linked to this agent)
    voxyra_user_doc = await db.users.find_one({"agent_id": agent_id, "tenant_id": tid})
    if body.voxyra_email and voxyra_user_doc and body.voxyra_email != voxyra_user_doc.get("email"):
        new_email = validate_email_str(body.voxyra_email)
        clash = await db.users.find_one({"email": new_email, "id": {"$ne": voxyra_user_doc["id"]}})
        if clash:
            raise HTTPException(status_code=400, detail="Email já está em uso por outro usuário")
        await db.users.update_one({"id": voxyra_user_doc["id"]}, {"$set": {"email": new_email}})
        changes["voxyra_email"] = {"from": voxyra_user_doc.get("email"), "to": new_email}

    if body.voxyra_password:
        if voxyra_user_doc:
            if len(body.voxyra_password) < 6:
                raise HTTPException(status_code=400, detail="Senha Voxyra mínima 6 caracteres")
            await db.users.update_one(
                {"id": voxyra_user_doc["id"]},
                {"$set": {"password_hash": hash_password(body.voxyra_password)}})
            changes["voxyra_password"] = "changed"
        else:
            pbx_warnings.append("Não há usuário Voxyra vinculado para alterar senha")

    # SIP password (FusionPBX extension)
    if body.sip_password and pbx_db_ok and client:
        if len(body.sip_password) < 4:
            raise HTTPException(status_code=400, detail="Senha SIP mínima 4 caracteres")
        try:
            await client.update_extension_password(
                domain_uuid=s.get("domain_uuid"),
                extension=agent.get("extension"),
                new_password=body.sip_password,
            )
            changes["sip_password"] = "changed"
        except Exception as e:
            pbx_warnings.append(f"SIP password: {e}")

    # Queue tiers (set of external_ids)
    if body.queue_uuids is not None and pbx_db_ok and client:
        # Map current queues (mongo) -> external_ids
        cur_queue_docs = await db.queues.find(
            {"tenant_id": tid, "id": {"$in": agent.get("queues") or []}}, {"_id": 0}).to_list(50)
        current_external = {q.get("external_id") for q in cur_queue_docs if q.get("external_id")}
        desired_external = set(body.queue_uuids)
        to_add = desired_external - current_external
        to_remove = current_external - desired_external
        for qu in to_add:
            try: await client.assign_agent_to_queue(agent["external_id"], qu)
            except Exception as e: pbx_warnings.append(f"add tier {qu[:8]}: {e}")
        for qu in to_remove:
            try: await client.remove_agent_from_queue(agent["external_id"], qu)
            except Exception as e: pbx_warnings.append(f"remove tier {qu[:8]}: {e}")
        # Update local mapping
        new_queue_ids = []
        if desired_external:
            new_queues = await db.queues.find(
                {"tenant_id": tid, "external_id": {"$in": list(desired_external)}}, {"_id": 0, "id": 1}).to_list(50)
            new_queue_ids = [q["id"] for q in new_queues]
        update["queues"] = new_queue_ids
        changes["queues"] = {"count": len(new_queue_ids)}

    if update:
        await db.agents.update_one({"id": agent_id}, {"$set": update})
    if changes:
        await write_audit(user, "update", "agent", agent_id,
                           f"Edição de {agent.get('name')}", changes)
    fresh = await db.agents.find_one({"id": agent_id}, {"_id": 0})
    return {"ok": True, "agent": fresh, "changes": changes, "warnings": pbx_warnings}


@api.put("/agents/{agent_id}/status")
async def set_agent_status(agent_id: str, body: AgentStatusReq,
                           user: dict = Depends(get_current_user)):
    """Update agent status (Voxyra + FusionPBX). Agents can update their own;
    admins/supervisors can update anyone in tenant."""
    new_status = (body.status or "").lower().strip()
    if new_status not in VALID_VOXYRA_STATUSES:
        raise HTTPException(status_code=400, detail=f"Status inválido. Use: {sorted(VALID_VOXYRA_STATUSES)}")
    f = tenant_filter(user)
    agent = await db.agents.find_one({"id": agent_id, **f}, {"_id": 0})
    if not agent:
        raise HTTPException(status_code=404, detail="Agente não encontrado")
    # Permission: agent só muda o seu próprio
    if user.get("role") == "agent":
        # Agent must be linked to this agent_id via user.agent_id
        if user.get("agent_id") != agent_id:
            raise HTTPException(status_code=403, detail="Você só pode alterar seu próprio status")
    elif user.get("role") not in ("super_admin", "admin", "supervisor"):
        raise HTTPException(status_code=403, detail="Sem permissão")

    # Atualiza local
    await db.agents.update_one(
        {"id": agent_id}, {"$set": {"status": new_status,
                                    "pbx_status": VOXYRA_STATUS_TO_PBX[new_status],
                                    "status_changed_at": datetime.now(timezone.utc).isoformat()}})

    # Tenta refletir no FusionPBX (modo DB)
    pbx_synced = False
    pbx_error = None
    tid = agent.get("tenant_id") or tenant_scope(user)
    if tid and agent.get("external_id"):
        s = await db.fusionpbx_settings.find_one({"tenant_id": tid})
        if s and (s.get("connection_type") == "db") and s.get("db_host"):
            try:
                client = FusionPBXDBClient(
                    host=s["db_host"], port=int(s.get("db_port") or 5432),
                    database=s.get("db_name") or "fusionpbx",
                    username=s["db_username"], password=s.get("db_password") or "",
                    domain_uuid=s.get("domain_uuid"), ssl=bool(s.get("db_ssl")),
                )
                await client.update_agent_status(agent["external_id"], VOXYRA_STATUS_TO_PBX[new_status])
                pbx_synced = True
            except FusionPBXDBError as e:
                pbx_error = str(e)
                logger.warning("set_agent_status: PBX update falhou: %s", e)
    # Reflete em memória do mod_callcenter (live distributor)
    if tid and agent.get("external_id"):
        # Quando voltando para Available, marca state=Waiting para começar a receber.
        # Em pausa/offline marca state=Idle (mod_callcenter não distribui).
        live_state = "Waiting" if new_status in ("online", "available") else "Idle"
        # Quando o agente fica disponível, garante que os tiers das filas
        # ativas dele estejam aplicados em memória (importante quando o
        # mod_callcenter perde tiers após reload ou se o agente nunca passou
        # por select_my_queues nesta sessão).
        add_q = None
        if new_status in ("online", "available"):
            active_qids = agent.get("active_queues") or agent.get("queues") or []
            if active_qids:
                qdocs = await db.queues.find(
                    {"tenant_id": tid, "id": {"$in": active_qids}},
                    {"_id": 0, "extension": 1}).to_list(50)
                domain = (await db.fusionpbx_settings.find_one({"tenant_id": tid}) or {}).get("domain_name")
                names = []
                for q in qdocs:
                    qext = q.get("extension")
                    if not qext:
                        continue
                    qn = str(qext)
                    if domain and "@" not in qn:
                        qn = f"{qn}@{domain}"
                    names.append(qn)
                add_q = names or None
        await _pbx_apply_agent_live(
            tid, agent.get("external_id") or "",
            status=VOXYRA_STATUS_TO_PBX[new_status],
            state=live_state,
            clear_tiers=bool(add_q),
            add_tier_queues=add_q,
        )

    await write_audit(user, "update", "agent_status", agent_id,
                      f"{agent.get('name')} → {new_status}",
                      {"new_status": new_status, "pbx_synced": pbx_synced})
    return {"ok": True, "status": new_status, "pbx_synced": pbx_synced, "pbx_error": pbx_error}


@api.get("/agents/me/info")
async def get_my_agent(user: dict = Depends(get_current_user)):
    """Returns the agent linked to the logged-in user (for agent dashboard)."""
    if user.get("role") != "agent":
        raise HTTPException(status_code=400, detail="Apenas para usuários com perfil de agente")
    aid = user.get("agent_id")
    if not aid:
        raise HTTPException(status_code=404, detail="Usuário não está vinculado a um agente. Peça a um admin para vincular em /usuários.")
    agent = await db.agents.find_one({"id": aid, "tenant_id": user.get("tenant_id")}, {"_id": 0})
    if not agent:
        raise HTTPException(status_code=404, detail="Agente vinculado não encontrado")
    return {"agent": agent}


class ExtensionReq(BaseModel):
    extension: str

@api.put("/agents/me/extension")
async def set_my_extension(body: ExtensionReq, user: dict = Depends(get_current_user)):
    """Agent updates the extension where calls should ring (FusionPBX agent_contact).
    Persists the new extension on the local agent doc and pushes user/<ext>@<domain>
    to v_call_center_agents.agent_contact.
    """
    if user.get("role") != "agent":
        raise HTTPException(status_code=403, detail="Apenas agentes")
    # Permissão: agent.change_extension. No primeiro login (extension ainda não
    # definida) sempre permitimos. Bloqueia somente quando o agente tenta TROCAR
    # depois de já ter ramal e a permissão foi removida.
    aid = user.get("agent_id")
    if not aid:
        raise HTTPException(status_code=404, detail="Usuário não vinculado a agente")
    existing_agent = await db.agents.find_one(
        {"id": aid, "tenant_id": user.get("tenant_id")},
        {"_id": 0, "extension": 1})
    cur_ext = (existing_agent or {}).get("extension")
    is_changing = bool(cur_ext) and cur_ext not in ("", "999")
    if is_changing and "agent.change_extension" not in effective_permissions(user):
        raise HTTPException(status_code=403,
                              detail="Sem permissão para trocar de ramal. Faça logout e logue novamente.")
    ext = (body.extension or "").strip()
    if not ext.isdigit() or not (2 <= len(ext) <= 8):
        raise HTTPException(status_code=400, detail="Ramal inválido (use somente dígitos, 2 a 8 caracteres)")
    tid = user.get("tenant_id")
    agent = await db.agents.find_one({"id": aid, "tenant_id": tid}, {"_id": 0})
    if not agent:
        raise HTTPException(status_code=404, detail="Agente não encontrado")
    s = await db.fusionpbx_settings.find_one({"tenant_id": tid})
    pbx_synced = False
    pbx_error = None
    new_contact = None
    if s and s.get("connection_type") == "db" and s.get("db_host") and agent.get("external_id"):
        try:
            client = FusionPBXDBClient(
                host=s["db_host"], port=int(s.get("db_port") or 5432),
                database=s.get("db_name") or "fusionpbx",
                username=s["db_username"], password=s.get("db_password") or "",
                domain_uuid=s.get("domain_uuid"), ssl=bool(s.get("db_ssl")),
            )
            new_contact = await client.update_agent_contact(
                agent["external_id"], ext, s.get("domain_name") or "")
            # Reset login state: deslogar de TODAS as filas (limpa tiers antigos)
            try:
                await client.remove_all_tiers_for_agent(agent["external_id"])
            except Exception as e:
                logger.warning("remove_all_tiers no login falhou: %s", e)
            # Marcar agente como Available para começar a receber chamadas
            try:
                await client.update_agent_status(agent["external_id"], "Available")
            except Exception as e:
                logger.warning("update_agent_status Available falhou: %s", e)
            pbx_synced = True
        except FusionPBXDBError as e:
            pbx_error = str(e)
            logger.warning("set_my_extension PBX falhou: %s", e)
    await db.agents.update_one(
        {"id": aid}, {"$set": {"extension": ext,
                                "agent_contact": new_contact,
                                "status": "offline",
                                "pbx_status": "Logged Out",
                                "active_queues": [],
                                "extension_changed_at": datetime.now(timezone.utc).isoformat()}})
    if pbx_synced:
        await _pbx_reload_callcenter(tid)
        # Atualiza apenas contact + limpa tiers antigos. Status fica
        # "Logged Out" (deslogado). O agente precisa clicar manualmente em
        # Disponível no painel para começar a receber chamadas.
        await _pbx_apply_agent_live(
            tid, agent.get("external_id") or "",
            status="Logged Out", state="Idle", contact=new_contact,
            clear_tiers=True,
        )
    await write_audit(user, "update", "agent_extension", aid,
                       f"{agent.get('name')} ramal → {ext}",
                       {"extension": ext, "pbx_synced": pbx_synced})
    return {"ok": True, "extension": ext, "agent_contact": new_contact,
            "pbx_synced": pbx_synced, "pbx_error": pbx_error}


@api.post("/agents/me/logout")
async def agent_logout(request: Request):
    """Cleanup ao fechar a aba/logout do agente: remove tiers e marca Logged Out.
    Aceita JWT via Authorization, cookie OU query string `?token=...` para
    funcionar com navigator.sendBeacon (não envia headers customizados)."""
    token = request.cookies.get("access_token")
    if not token:
        auth = request.headers.get("Authorization", "")
        if auth.startswith("Bearer "): token = auth[7:]
    if not token:
        token = request.query_params.get("token")
    if not token:
        return {"ok": False, "reason": "no token"}
    try:
        payload = jwt.decode(token, _secret(), algorithms=[JWT_ALGO])
    except Exception:
        return {"ok": False, "reason": "invalid token"}
    user_id = payload.get("sub")
    if not user_id:
        return {"ok": False}
    user = await db.users.find_one({"id": user_id}, {"_id": 0})
    if not user or user.get("role") != "agent":
        return {"ok": False}
    aid = user.get("agent_id")
    if not aid:
        return {"ok": False}
    tid = user.get("tenant_id")
    agent = await db.agents.find_one({"id": aid, "tenant_id": tid}, {"_id": 0})
    if not agent:
        return {"ok": False}
    s = await db.fusionpbx_settings.find_one({"tenant_id": tid}) or {}
    pbx_logged_out = False
    tiers_removed = 0
    parking_contact = f"user/999@{s.get('domain_name') or ''}"
    if (s.get("connection_type") == "db" and s.get("db_host")
            and agent.get("external_id")):
        try:
            client = FusionPBXDBClient(
                host=s["db_host"], port=int(s.get("db_port") or 5432),
                database=s.get("db_name") or "fusionpbx",
                username=s["db_username"], password=s.get("db_password") or "",
                domain_uuid=s.get("domain_uuid"), ssl=bool(s.get("db_ssl")),
            )
            try:
                tiers_removed = await client.remove_all_tiers_for_agent(agent["external_id"])
            except Exception as e:
                logger.warning("logout remove_all_tiers: %s", e)
            try:
                await client.update_agent_status(agent["external_id"], "Logged Out")
                pbx_logged_out = True
            except Exception as e:
                logger.warning("logout update_agent_status: %s", e)
            # Reseta contato para ramal "parking" (999) — agente não recebe
            # mais nada mesmo se algum reload futuro relistá-lo.
            try:
                if s.get("domain_name"):
                    await client.update_agent_contact(
                        agent["external_id"], "999", s["domain_name"])
            except Exception as e:
                logger.warning("logout reset contact 999: %s", e)
        except Exception as e:
            logger.warning("agent_logout PBX falhou: %s", e)
    await db.agents.update_one(
        {"id": aid},
        {"$set": {"status": "offline", "pbx_status": "Logged Out",
                   "active_queues": [],
                   "agent_contact": parking_contact,
                   "extension": "999",
                   "logout_at": datetime.now(timezone.utc).isoformat()}})
    if pbx_logged_out or tiers_removed:
        await _pbx_reload_callcenter(tid)
        await _pbx_apply_agent_live(
            tid, agent.get("external_id") or "",
            status="Logged Out", state="Idle",
            contact=parking_contact,
            clear_tiers=True,
        )
    # Defensive: força a remoção de qualquer agente "preso" em memória do
    # mod_callcenter cujo nome corresponda ao ramal anterior (qualquer formato).
    # Resolve o caso onde o nome em memória difere do que enviamos
    # (ex: UUID vs `<ext>@<dominio>`), o que é comum após reloads do PBX.
    prev_ext = str(agent.get("extension") or "").strip()
    if prev_ext and prev_ext != "999" and s.get("esl_host"):
        try:
            esl = FreeSwitchESL(
                host=s["esl_host"], port=int(s.get("esl_port") or 8021),
                password=s.get("esl_password") or "ClueCon",
                timeout=float(s.get("esl_timeout") or 5.0),
            )
            await esl.callcenter_force_clear_by_extension(
                prev_ext, domain=s.get("domain_name"))
        except Exception as e:
            logger.warning("logout force_clear_by_extension falhou: %s", e)
    await write_audit(user, "logout", "agent", aid,
                       f"{agent.get('name')} desconectou",
                       {"tiers_removed": tiers_removed, "pbx_logged_out": pbx_logged_out})
    return {"ok": True, "tiers_removed": tiers_removed, "pbx_logged_out": pbx_logged_out}


@api.get("/agents/me/queues")
async def my_available_queues(user: dict = Depends(get_current_user)):
    """Return queues the agent has access to (via tier membership) plus
    which ones are currently 'active' (selected by the agent at login)."""
    if user.get("role") != "agent":
        raise HTTPException(status_code=403, detail="Apenas agentes")
    aid = user.get("agent_id")
    if not aid:
        raise HTTPException(status_code=404, detail="Agente não vinculado")
    tid = user.get("tenant_id")
    me = await db.agents.find_one({"id": aid, "tenant_id": tid}, {"_id": 0})
    if not me:
        raise HTTPException(status_code=404, detail="Agente não encontrado")
    qids = me.get("queues") or []
    if not qids:
        return {"queues": [], "active_queues": []}
    queues = await db.queues.find(
        {"tenant_id": tid, "id": {"$in": qids}}, {"_id": 0}).to_list(50)
    active = me.get("active_queues") or []
    out = [{
        "id": q["id"], "name": q.get("name"), "extension": q.get("extension"),
        "strategy": q.get("strategy"), "waiting": q.get("waiting", 0),
        "active": q["id"] in active,
    } for q in queues]
    return {"queues": out, "active_queues": active}


class QueueSelectionReq(BaseModel):
    queue_ids: List[str]

@api.post("/agents/me/queues/select")
async def select_my_queues(body: QueueSelectionReq, user: dict = Depends(get_current_user)):
    """Agent chooses queues to actively log into.
    This adds/removes tiers (agent↔queue) in FusionPBX PostgreSQL so the
    distributor only rings selected queues. Falls back to local-only when
    PBX update is unavailable (read-only or non-DB integration)."""
    if user.get("role") != "agent":
        raise HTTPException(status_code=403, detail="Apenas agentes")
    aid = user.get("agent_id")
    if not aid:
        raise HTTPException(status_code=404, detail="Agente não vinculado")
    tid = user.get("tenant_id")
    me = await db.agents.find_one({"id": aid, "tenant_id": tid}, {"_id": 0})
    if not me:
        raise HTTPException(status_code=404, detail="Agente não encontrado")
    allowed_ids = set(me.get("queues") or [])
    chosen_ids = [qid for qid in (body.queue_ids or []) if qid in allowed_ids]
    if not chosen_ids:
        raise HTTPException(status_code=400, detail="Selecione ao menos uma fila autorizada")

    # Resolve queue PBX uuids
    queues = await db.queues.find(
        {"tenant_id": tid, "id": {"$in": list(allowed_ids)}}, {"_id": 0}).to_list(50)
    qmap = {q["id"]: q for q in queues}
    chosen_uuids = {qmap[q].get("external_id") for q in chosen_ids if qmap.get(q) and qmap[q].get("external_id")}
    all_uuids   = {qmap[q].get("external_id") for q in allowed_ids if qmap.get(q) and qmap[q].get("external_id")}
    to_remove = list(all_uuids - chosen_uuids)

    pbx_added: List[str] = []
    pbx_removed: List[str] = []
    pbx_errors: List[str] = []
    s = await db.fusionpbx_settings.find_one({"tenant_id": tid}) or {}
    if s.get("connection_type") == "db" and s.get("db_host") and me.get("external_id"):
        try:
            client = FusionPBXDBClient(
                host=s["db_host"], port=int(s.get("db_port") or 5432),
                database=s.get("db_name") or "fusionpbx",
                username=s["db_username"], password=s.get("db_password") or "",
                domain_uuid=s.get("domain_uuid"), ssl=bool(s.get("db_ssl")),
            )
            for quuid in chosen_uuids:
                try:
                    await client.assign_agent_to_queue(me["external_id"], quuid)
                    pbx_added.append(quuid)
                except FusionPBXDBError as e:
                    pbx_errors.append(f"add {quuid[:8]}: {e}")
            for quuid in to_remove:
                try:
                    ok = await client.remove_agent_from_queue(me["external_id"], quuid)
                    if ok:
                        pbx_removed.append(quuid)
                except FusionPBXDBError as e:
                    pbx_errors.append(f"remove {quuid[:8]}: {e}")
        except Exception as e:
            pbx_errors.append(str(e))

    await db.agents.update_one(
        {"id": aid},
        {"$set": {"active_queues": chosen_ids,
                  "active_queues_changed_at": datetime.now(timezone.utc).isoformat()}})
    if pbx_added or pbx_removed:
        await _pbx_reload_callcenter(tid)
        # Sincroniza tiers em MEMÓRIA do mod_callcenter.
        # IMPORTANTE: o nome da fila no mod_callcenter é
        # `<queue_extension>@<domain_name>` (ex: 620@grupoicore...),
        # NÃO o `queue_name` amigável (FILA-SUPORTE-CORP). Usar o nome
        # amigável faz o tier ser ignorado silenciosamente.
        domain = (await db.fusionpbx_settings.find_one({"tenant_id": tid}) or {}).get("domain_name")
        chosen_queue_names = []
        for qid in chosen_ids:
            q = qmap.get(qid) or {}
            qext = q.get("extension")
            if not qext:
                continue
            qname = str(qext)
            if domain and "@" not in qname:
                qname = f"{qname}@{domain}"
            chosen_queue_names.append(qname)
        await _pbx_apply_agent_live(
            tid, me.get("external_id") or "",
            clear_tiers=True,
            add_tier_queues=chosen_queue_names,
        )
    await write_audit(user, "update", "agent_queues", aid,
                       f"{me.get('name')} ativou {len(chosen_ids)} fila(s)",
                       {"active_queues": chosen_ids,
                        "pbx_added": len(pbx_added), "pbx_removed": len(pbx_removed)})
    return {"ok": True, "active_queues": chosen_ids,
            "pbx_added": pbx_added, "pbx_removed": pbx_removed,
            "pbx_errors": pbx_errors}


@api.get("/agents/me/queue-status")
async def my_queue_status(user: dict = Depends(get_current_user)):
    """For the agent dashboard: returns the agent's queues with
    per-queue counts of logged-in agents (online/paused) and
    waiting calls in queue (from FreeSWITCH ESL when available)."""
    if user.get("role") != "agent":
        raise HTTPException(status_code=403, detail="Apenas agentes")
    aid = user.get("agent_id")
    if not aid:
        raise HTTPException(status_code=404, detail="Agente não vinculado")
    tid = user.get("tenant_id")
    me = await db.agents.find_one({"id": aid, "tenant_id": tid}, {"_id": 0})
    if not me:
        raise HTTPException(status_code=404, detail="Agente não encontrado")
    my_queue_ids = me.get("queues") or []
    # Filter by active_queues if agent has selected (login-time pick).
    active_queues = me.get("active_queues") or []
    if active_queues:
        my_queue_ids = [q for q in my_queue_ids if q in active_queues]
    if not my_queue_ids:
        return {"queues": []}
    queues_db = await db.queues.find(
        {"tenant_id": tid, "id": {"$in": my_queue_ids}}, {"_id": 0}).to_list(50)
    # Find peers per queue (agents that share at least one of my_queue_ids)
    peers = await db.agents.find(
        {"tenant_id": tid, "queues": {"$in": my_queue_ids}},
        {"_id": 0, "id": 1, "name": 1, "avatar": 1, "extension": 1, "status": 1, "queues": 1}
    ).to_list(500)
    # Live waiting count per queue extension via ESL (best-effort)
    waiting_by_qext: Dict[str, int] = {}
    s = await db.fusionpbx_settings.find_one({"tenant_id": tid}) or {}
    if s.get("enabled") and s.get("esl_host"):
        try:
            esl = FreeSwitchESL(
                host=s["esl_host"], port=int(s.get("esl_port") or 8021),
                password=s.get("esl_password") or "ClueCon",
                timeout=float(s.get("esl_timeout") or 5.0),
            )
            rows = await esl.show_channels()
            if rows:
                ext_to_qext = {q.get("extension"): q.get("extension")
                               for q in queues_db if q.get("extension")}
                for r in rows:
                    dest = str(r.get("dest") or "")
                    if dest in ext_to_qext and (r.get("callstate") or "").upper() != "ACTIVE":
                        waiting_by_qext[dest] = waiting_by_qext.get(dest, 0) + 1
        except Exception as e:
            logger.warning("my_queue_status ESL falhou: %s", e)
    out = []
    for q in queues_db:
        peers_in_q = [p for p in peers if q["id"] in (p.get("queues") or [])]
        agents_summary = [{
            "id": p["id"], "name": p.get("name"), "avatar": p.get("avatar"),
            "extension": p.get("extension"), "status": p.get("status") or "offline",
            "is_me": p["id"] == aid,
        } for p in peers_in_q]
        # local counts
        online = sum(1 for p in agents_summary if p["status"] == "online")
        paused = sum(1 for p in agents_summary if p["status"] == "paused")
        offline = sum(1 for p in agents_summary if p["status"] == "offline")
        out.append({
            "id": q["id"], "name": q.get("name"), "extension": q.get("extension"),
            "agents": agents_summary,
            "online": online, "paused": paused, "offline": offline,
            "logged_in": online + paused,
            "waiting": waiting_by_qext.get(q.get("extension") or "", 0),
        })
    return {"queues": out, "fetched_at": datetime.now(timezone.utc).isoformat()}


@api.get("/queues")
async def list_queues(user: dict = Depends(require_permission("queues.view"))):
    f = tenant_filter(user)
    items = await db.queues.find(f, {"_id": 0}).to_list(500)
    agents = await db.agents.find(f, {"_id": 0, "queues": 1}).to_list(500)
    for q in items:
        q["agent_count"] = sum(1 for a in agents if q["id"] in a.get("queues", []))
    return {"queues": items}

# ---------- Recordings ----------
@api.get("/recordings")
async def list_recordings(
    user: dict = Depends(get_current_user),
    agent_id: Optional[str] = None, queue_id: Optional[str] = None,
    search: Optional[str] = None, limit: int = Query(100, le=500),
):
    perms = effective_permissions(user)
    has_all = "recordings.view_all" in perms
    has_own = "recordings.view_own" in perms
    if not (has_all or has_own): raise HTTPException(status_code=403, detail="Sem permissão")
    q = {**tenant_filter(user)}
    if not has_all and has_own:
        own_id = user.get("agent_id")
        if not own_id: return {"recordings": []}
        q["agent_id"] = own_id
    elif agent_id: q["agent_id"] = agent_id
    # Per-user extension whitelist (limits which agents' recordings are visible)
    allowed_agent_ids = await allowed_agent_ids_for(user)
    if allowed_agent_ids is not None:
        if not allowed_agent_ids:
            return {"recordings": []}
        existing = q.get("agent_id")
        if isinstance(existing, str):
            if existing not in allowed_agent_ids:
                return {"recordings": []}
        else:
            q["agent_id"] = {"$in": list(allowed_agent_ids)}
    if queue_id: q["queue_id"] = queue_id
    if search:
        q["$or"] = [
            {"caller_number": {"$regex": search, "$options": "i"}},
            {"agent_name": {"$regex": search, "$options": "i"}},
        ]
    items = await db.recordings.find(q, {"_id": 0}).sort("started_at", -1).to_list(limit)
    # If recording came from PBX (storage_key starts with fusionpbx://), point audio_url to stream endpoint
    tid = tenant_scope(user)
    sftp_ok = False
    if tid:
        s = await db.fusionpbx_settings.find_one({"tenant_id": tid})
        sftp_ok = bool(s and s.get("sftp_host") and s.get("sftp_username")
                       and (s.get("sftp_password") or s.get("sftp_private_key")))
    api_base = os.environ.get("PUBLIC_API_BASE", "")
    for rec in items:
        url = rec.get("audio_url") or ""
        is_pbx = (url and url.startswith("fusionpbx://")) \
                 or (rec.get("storage_key", "").startswith("fusionpbx://"))
        if is_pbx and sftp_ok:
            rec["audio_url"] = f"/api/recordings/{rec['id']}/stream"
            rec["streamable"] = True
        elif is_pbx and not sftp_ok:
            rec["audio_url"] = ""
            rec["streamable"] = False
            rec["unavailable_reason"] = "Configure SFTP em Central PBX → Configuração para ouvir gravações reais"
        else:
            rec["streamable"] = bool(url)
    return {"recordings": items}

@api.get("/recordings/{rec_id}")
async def get_recording(rec_id: str, user: dict = Depends(get_current_user)):
    perms = effective_permissions(user)
    if not ("recordings.view_all" in perms or "recordings.view_own" in perms):
        raise HTTPException(status_code=403, detail="Sem permissão")
    r = await db.recordings.find_one({"id": rec_id, **tenant_filter(user)}, {"_id": 0})
    if not r: raise HTTPException(status_code=404, detail="Gravação não encontrada")
    if "recordings.view_all" not in perms and r.get("agent_id") != user.get("agent_id"):
        raise HTTPException(status_code=403, detail="Sem permissão para esta gravação")
    allowed_agent_ids = await allowed_agent_ids_for(user)
    if allowed_agent_ids is not None and r.get("agent_id") not in allowed_agent_ids:
        raise HTTPException(status_code=403, detail="Sem permissão para esta gravação")
    return r

class NoteUpdate(BaseModel):
    notes: str

@api.patch("/recordings/{rec_id}")
async def update_recording(rec_id: str, body: NoteUpdate, user: dict = Depends(require_permission("recordings.edit_notes"))):
    res = await db.recordings.update_one({"id": rec_id, **tenant_filter(user)}, {"$set": {"notes": body.notes}})
    if res.matched_count == 0: raise HTTPException(status_code=404, detail="Não encontrada")
    return {"ok": True}

# ---------- Reports ----------
def _period_cutoff(period: str) -> datetime:
    now = datetime.now(timezone.utc)
    if period == "today": return now.replace(hour=0, minute=0, second=0, microsecond=0)
    if period == "30d": return now - timedelta(days=30)
    return now - timedelta(days=7)

async def _build_report(user: dict, report_type: str, period: str, agent_id: Optional[str], queue_id: Optional[str]):
    cutoff = _period_cutoff(period).isoformat()
    f = tenant_filter(user)
    # Per-user extension whitelist limits which agents/calls are visible
    allowed_agent_ids = await allowed_agent_ids_for(user)
    allowed_ext = allowed_extensions_for(user)
    agent_filter = {**f}
    if allowed_ext is not None:
        agent_filter["extension"] = {"$in": list(allowed_ext)}
    call_filter = {**f}
    if allowed_agent_ids is not None:
        if not allowed_agent_ids:
            # No allowed agents → empty result
            return {"title": "Relatório", "columns": [], "rows": []}
        call_filter["agent_id"] = {"$in": list(allowed_agent_ids)}
    if report_type == "agents":
        agents = await db.agents.find(agent_filter, {"_id": 0}).to_list(500)
        if agent_id: agents = [a for a in agents if a["id"] == agent_id]
        rows = []
        for a in agents:
            base = {"agent_id": a["id"], "started_at": {"$gte": cutoff}, **f}
            answered = await db.calls.count_documents({**base, "disposition": "answered"})
            missed = await db.calls.count_documents({**base, "disposition": {"$in": ["missed", "abandoned"]}})
            # Percentiles for handle time
            durations = await db.calls.find(
                {**base, "disposition": "answered", "duration_sec": {"$gt": 0}},
                {"_id": 0, "duration_sec": 1}
            ).to_list(5000)
            ds = sorted([d["duration_sec"] for d in durations]) if durations else []
            avg_h = int(sum(ds) / len(ds)) if ds else 0
            median = ds[len(ds)//2] if ds else 0
            p95 = ds[min(len(ds)-1, int(len(ds)*0.95))] if ds else 0
            mx = ds[-1] if ds else 0
            rows.append({"agent_name": a["name"], "extension": a["extension"], "status": a.get("status"),
                         "answered": answered, "missed": missed,
                         "avg_handle_sec": avg_h,
                         "median_handle_sec": median,
                         "p95_handle_sec": p95,
                         "max_handle_sec": mx,
                         "csat": a.get("csat", 0), "adherence_pct": a.get("adherence_pct", 0)})
        rows.sort(key=lambda r: r["answered"], reverse=True)
        return {"title": "Performance de Agentes", "columns": [
            {"key": "agent_name", "label": "Agente"}, {"key": "extension", "label": "Ramal"},
            {"key": "status", "label": "Status"}, {"key": "answered", "label": "Atendidas"},
            {"key": "missed", "label": "Perdidas"},
            {"key": "avg_handle_sec", "label": "TMA (s)"},
            {"key": "median_handle_sec", "label": "Mediana (s)"},
            {"key": "p95_handle_sec", "label": "P95 (s)"},
            {"key": "max_handle_sec", "label": "Máx (s)"},
            {"key": "csat", "label": "CSAT"}, {"key": "adherence_pct", "label": "Aderência %"},
        ], "rows": rows}
    if report_type == "queues":
        queues = await db.queues.find(f, {"_id": 0}).to_list(500)
        if queue_id: queues = [q for q in queues if q["id"] == queue_id]
        rows = []
        for q in queues:
            base = {"queue_id": q["id"], "started_at": {"$gte": cutoff}, **f}
            answered = await db.calls.count_documents({**base, "disposition": "answered"})
            missed = await db.calls.count_documents({**base, "disposition": {"$in": ["missed", "abandoned"]}})
            total = answered + missed
            rows.append({"queue_name": q["name"], "extension": q["extension"], "strategy": q["strategy"],
                         "answered": answered, "missed": missed, "total": total,
                         "sla_pct": round((answered / total) * 100, 1) if total else 0.0,
                         "avg_wait_sec": q.get("avg_wait_sec", 0)})
        rows.sort(key=lambda r: r["total"], reverse=True)
        return {"title": "Chamadas por Fila", "columns": [
            {"key": "queue_name", "label": "Fila"}, {"key": "extension", "label": "Ext."},
            {"key": "strategy", "label": "Estratégia"}, {"key": "answered", "label": "Atendidas"},
            {"key": "missed", "label": "Perdidas"}, {"key": "total", "label": "Total"},
            {"key": "sla_pct", "label": "SLA %"}, {"key": "avg_wait_sec", "label": "TME (s)"},
        ], "rows": rows}
    if report_type == "calls":
        q = {**call_filter, "started_at": {"$gte": cutoff}}
        if agent_id: q["agent_id"] = agent_id
        if queue_id: q["queue_id"] = queue_id
        docs = await db.calls.find(q, {"_id": 0}).sort("started_at", -1).to_list(2000)
        rows = [{"started_at": fmt_br_datetime(d.get("started_at")),
                 "agent_name": d.get("agent_name", "—"), "queue_name": d.get("queue_name", "—"),
                 "direction": DIRECTION_LABELS.get(d.get("direction"), "—"),
                 "caller_number": d.get("caller_number", "—"),
                 "disposition": DISPOSITION_LABELS.get(d.get("disposition"), "—"),
                 "duration_sec": d.get("duration_sec", 0), "wait_sec": d.get("wait_sec", 0)} for d in docs]
        return {"title": "Histórico de Chamadas (CDR)", "columns": [
            {"key": "started_at", "label": "Data/Hora"}, {"key": "agent_name", "label": "Agente"},
            {"key": "queue_name", "label": "Fila"}, {"key": "direction", "label": "Direção"},
            {"key": "caller_number", "label": "Número"}, {"key": "disposition", "label": "Status"},
            {"key": "duration_sec", "label": "Duração (s)"}, {"key": "wait_sec", "label": "Espera (s)"},
        ], "rows": rows}
    if report_type == "abandoned":
        q = {**call_filter, "disposition": {"$in": ["missed", "abandoned"]}, "started_at": {"$gte": cutoff}}
        if agent_id: q["agent_id"] = agent_id
        if queue_id: q["queue_id"] = queue_id
        docs = await db.calls.find(q, {"_id": 0}).sort("started_at", -1).to_list(2000)
        rows = [{"started_at": fmt_br_datetime(d.get("started_at")),
                 "abandonment_type": ABANDON_LABELS.get(d.get("abandonment_type"), "—"),
                 "queue_name": d.get("queue_name", "—"), "agent_name": d.get("agent_name", "—"),
                 "caller_number": d.get("caller_number", "—"), "wait_sec": d.get("wait_sec", 0)} for d in docs]
        return {"title": "Chamadas Abandonadas", "columns": [
            {"key": "started_at", "label": "Data/Hora"}, {"key": "abandonment_type", "label": "Tipo"},
            {"key": "queue_name", "label": "Fila"}, {"key": "agent_name", "label": "Agente"},
            {"key": "caller_number", "label": "Número"}, {"key": "wait_sec", "label": "Espera (s)"},
        ], "rows": rows}
    if report_type == "recordings":
        q = {**call_filter, "started_at": {"$gte": cutoff}}
        if agent_id: q["agent_id"] = agent_id
        if queue_id: q["queue_id"] = queue_id
        docs = await db.recordings.find(q, {"_id": 0}).sort("started_at", -1).to_list(2000)
        rows = [{"started_at": fmt_br_datetime(d.get("started_at")),
                 "agent_name": d.get("agent_name", "—"), "queue_name": d.get("queue_name", "—"),
                 "caller_number": d.get("caller_number", "—"),
                 "duration_sec": d.get("duration_sec", 0), "size_mb": d.get("size_mb", 0)} for d in docs]
        return {"title": "Gravações", "columns": [
            {"key": "started_at", "label": "Data/Hora"}, {"key": "agent_name", "label": "Agente"},
            {"key": "queue_name", "label": "Fila"}, {"key": "caller_number", "label": "Número"},
            {"key": "duration_sec", "label": "Duração (s)"}, {"key": "size_mb", "label": "Tamanho (MB)"},
        ], "rows": rows}
    if report_type == "hourly":
        q = {**call_filter, "started_at": {"$gte": cutoff}}
        if agent_id: q["agent_id"] = agent_id
        if queue_id: q["queue_id"] = queue_id
        docs = await db.calls.find(q, {"_id": 0, "started_at": 1, "disposition": 1}).to_list(5000)
        buckets = {h: {"hour": f"{h:02d}h", "answered": 0, "missed": 0, "total": 0} for h in range(24)}
        for d in docs:
            try:
                h = datetime.fromisoformat(d["started_at"]).hour
                buckets[h]["total"] += 1
                if d["disposition"] == "answered": buckets[h]["answered"] += 1
                elif d["disposition"] in ("missed", "abandoned"): buckets[h]["missed"] += 1
            except Exception: pass
        return {"title": "Produtividade Horária", "columns": [
            {"key": "hour", "label": "Hora"}, {"key": "answered", "label": "Atendidas"},
            {"key": "missed", "label": "Perdidas"}, {"key": "total", "label": "Total"},
        ], "rows": [buckets[h] for h in range(24)]}
    if report_type == "sla":
        queues = await db.queues.find(f, {"_id": 0}).to_list(500)
        if queue_id: queues = [q for q in queues if q["id"] == queue_id]
        rows = []
        for q in queues:
            target = int(q.get("sla_target_sec") or 20)
            base = {"queue_id": q["id"], "started_at": {"$gte": cutoff}, **f}
            answered_total = await db.calls.count_documents({**base, "disposition": "answered"})
            answered_within = await db.calls.count_documents({
                **base, "disposition": "answered", "wait_sec": {"$lte": target}
            })
            missed = await db.calls.count_documents({**base, "disposition": {"$in": ["missed", "abandoned"]}})
            offered = answered_total + missed
            sla = round((answered_within / answered_total) * 100, 1) if answered_total else 0.0
            asa_docs = await db.calls.find({**base, "disposition": "answered"},
                                            {"_id": 0, "wait_sec": 1}).to_list(5000)
            asa = int(sum(d.get("wait_sec", 0) for d in asa_docs) / max(len(asa_docs), 1)) if asa_docs else 0
            color = "green" if sla >= 80 else ("amber" if sla >= 60 else "red")
            rows.append({
                "queue_name": q["name"], "extension": q["extension"],
                "target_sec": target, "offered": offered,
                "answered": answered_total, "answered_within": answered_within,
                "missed": missed, "sla_pct": sla, "asa_sec": asa, "color": color,
            })
        rows.sort(key=lambda r: r["sla_pct"])
        return {"title": "SLA por Fila", "columns": [
            {"key": "queue_name", "label": "Fila"}, {"key": "extension", "label": "Ext."},
            {"key": "target_sec", "label": "Meta (s)"},
            {"key": "offered", "label": "Ofertadas"},
            {"key": "answered", "label": "Atendidas"},
            {"key": "answered_within", "label": "Dentro Meta"},
            {"key": "missed", "label": "Perdidas"},
            {"key": "sla_pct", "label": "SLA %"},
            {"key": "asa_sec", "label": "ASA (s)"},
        ], "rows": rows}

    if report_type == "agent_states":
        # Reads v_call_center_agent_status_log from FusionPBX (PostgreSQL mode)
        tid = tenant_scope(user)
        rows = []
        if tid:
            s = await db.fusionpbx_settings.find_one({"tenant_id": tid})
            if s and (s.get("connection_type") == "db") and s.get("db_host"):
                client = FusionPBXDBClient(
                    host=s["db_host"], port=int(s.get("db_port") or 5432),
                    database=s.get("db_name") or "fusionpbx",
                    username=s["db_username"], password=s.get("db_password") or "",
                    domain_uuid=s.get("domain_uuid"), ssl=bool(s.get("db_ssl")),
                )
                try:
                    conn = await client._connect()
                    try:
                        tbl = await conn.fetchval(
                            "SELECT to_regclass('public.v_call_center_agent_status_log')::text")
                        if tbl:
                            sql = """
                                SELECT a.agent_name, a.agent_id,
                                       l.value AS state, l.start_epoch, l.stop_epoch
                                FROM v_call_center_agent_status_log l
                                LEFT JOIN v_call_center_agents a
                                  ON a.call_center_agent_uuid = l.call_center_agent_uuid
                                WHERE a.domain_uuid = $1::uuid
                                  AND l.start_epoch >= EXTRACT(EPOCH FROM $2::timestamptz)
                                ORDER BY l.start_epoch DESC LIMIT 5000
                            """
                            recs = await conn.fetch(sql, client.domain_uuid, cutoff)
                            # Aggregate per agent + state
                            agg: Dict[str, Dict[str, int]] = {}
                            for r in recs:
                                ag = r["agent_name"] or r["agent_id"] or "—"
                                state = (r["state"] or "Available").strip()
                                start = int(r["start_epoch"] or 0)
                                stop = int(r["stop_epoch"] or 0) or int(time.time())
                                dur = max(0, stop - start)
                                if ag not in agg: agg[ag] = {}
                                agg[ag][state] = agg[ag].get(state, 0) + dur
                            for ag, states in agg.items():
                                total = sum(states.values()) or 1
                                rows.append({
                                    "agent_name": ag,
                                    "logged_in_sec": total,
                                    "available_sec": states.get("Available", 0),
                                    "on_break_sec": states.get("On Break", 0),
                                    "logged_out_sec": states.get("Logged Out", 0),
                                    "available_pct": round((states.get("Available", 0) / total) * 100, 1),
                                })
                    finally:
                        await conn.close()
                except FusionPBXDBError as e:
                    logger.warning("agent_states: %s", e)
        rows.sort(key=lambda r: r["available_pct"], reverse=True)
        return {"title": "Estados de Agente", "columns": [
            {"key": "agent_name", "label": "Agente"},
            {"key": "logged_in_sec", "label": "Tempo logado (s)"},
            {"key": "available_sec", "label": "Disponível (s)"},
            {"key": "on_break_sec", "label": "Em pausa (s)"},
            {"key": "logged_out_sec", "label": "Deslogado (s)"},
            {"key": "available_pct", "label": "% Disponível"},
        ], "rows": rows}

    if report_type == "heatmap":
        q = {**f, "started_at": {"$gte": cutoff}}
        if agent_id: q["agent_id"] = agent_id
        if queue_id: q["queue_id"] = queue_id
        docs = await db.calls.find(q, {"_id": 0, "started_at": 1, "disposition": 1}).to_list(20000)
        # 7 (dia da semana) × 24 (hora)
        grid = [[0 for _ in range(24)] for _ in range(7)]
        days = ["Seg", "Ter", "Qua", "Qui", "Sex", "Sáb", "Dom"]
        max_v = 0
        for d in docs:
            try:
                dt = datetime.fromisoformat(d["started_at"])
                wd = dt.weekday()
                h = dt.hour
                grid[wd][h] += 1
                if grid[wd][h] > max_v: max_v = grid[wd][h]
            except Exception: pass
        rows = []
        for wd in range(7):
            row = {"day": days[wd]}
            for h in range(24):
                row[f"h{h:02d}"] = grid[wd][h]
            rows.append(row)
        cols = [{"key": "day", "label": "Dia"}]
        for h in range(24):
            cols.append({"key": f"h{h:02d}", "label": f"{h:02d}h"})
        return {"title": "Heatmap (Dia × Hora)", "columns": cols, "rows": rows,
                "max_value": max_v}

    if report_type == "compare":
        # Comparativo período atual vs período anterior
        now = datetime.now(timezone.utc)
        if period == "today":
            cur_start = now.replace(hour=0, minute=0, second=0, microsecond=0)
            prev_start = cur_start - timedelta(days=1)
            prev_end = cur_start
        elif period == "30d":
            cur_start = now - timedelta(days=30)
            prev_start = cur_start - timedelta(days=30)
            prev_end = cur_start
        else:  # 7d default
            cur_start = now - timedelta(days=7)
            prev_start = cur_start - timedelta(days=7)
            prev_end = cur_start

        async def _stats(start: datetime, end: Optional[datetime] = None):
            q = {**f, "started_at": {"$gte": start.isoformat()}}
            if end is not None:
                q["started_at"]["$lt"] = end.isoformat()
            answered = await db.calls.count_documents({**q, "disposition": "answered"})
            missed = await db.calls.count_documents({**q, "disposition": {"$in": ["missed", "abandoned"]}})
            recs = await db.calls.find({**q, "disposition": "answered"},
                                        {"_id": 0, "duration_sec": 1, "wait_sec": 1}).to_list(20000)
            durs = [r.get("duration_sec", 0) for r in recs]
            waits = [r.get("wait_sec", 0) for r in recs]
            avg_h = int(sum(durs) / len(durs)) if durs else 0
            avg_w = int(sum(waits) / len(waits)) if waits else 0
            return {"answered": answered, "missed": missed,
                    "total": answered + missed,
                    "avg_handle_sec": avg_h, "avg_wait_sec": avg_w}

        cur = await _stats(cur_start)
        prev = await _stats(prev_start, prev_end)

        def _delta(a, b):
            if not b: return 100.0 if a else 0.0
            return round(((a - b) / b) * 100, 1)

        rows = [
            {"metric": "Atendidas", "current": cur["answered"], "previous": prev["answered"],
             "delta_pct": _delta(cur["answered"], prev["answered"]),
             "trend": "up" if cur["answered"] >= prev["answered"] else "down"},
            {"metric": "Perdidas", "current": cur["missed"], "previous": prev["missed"],
             "delta_pct": _delta(cur["missed"], prev["missed"]),
             "trend": "down" if cur["missed"] >= prev["missed"] else "up"},  # less is better
            {"metric": "Total", "current": cur["total"], "previous": prev["total"],
             "delta_pct": _delta(cur["total"], prev["total"]),
             "trend": "up" if cur["total"] >= prev["total"] else "down"},
            {"metric": "TMA (s)", "current": cur["avg_handle_sec"], "previous": prev["avg_handle_sec"],
             "delta_pct": _delta(cur["avg_handle_sec"], prev["avg_handle_sec"]),
             "trend": "down" if cur["avg_handle_sec"] >= prev["avg_handle_sec"] else "up"},
            {"metric": "TME (s)", "current": cur["avg_wait_sec"], "previous": prev["avg_wait_sec"],
             "delta_pct": _delta(cur["avg_wait_sec"], prev["avg_wait_sec"]),
             "trend": "down" if cur["avg_wait_sec"] >= prev["avg_wait_sec"] else "up"},
        ]
        return {"title": "Comparativo de Períodos", "columns": [
            {"key": "metric", "label": "Métrica"},
            {"key": "current", "label": "Atual"},
            {"key": "previous", "label": "Anterior"},
            {"key": "delta_pct", "label": "Variação %"},
            {"key": "trend", "label": "Tendência"},
        ], "rows": rows}

    raise HTTPException(status_code=400, detail="Tipo de relatório inválido")

@api.get("/reports/agents")
async def reports_agents(user: dict = Depends(require_permission("reports.view"))):
    f = tenant_filter(user)
    allowed_ext = allowed_extensions_for(user)
    if allowed_ext is not None:
        f = {**f, "extension": {"$in": list(allowed_ext)}}
    agents = await db.agents.find(f, {"_id": 0}).to_list(500)
    cutoff = (datetime.now(timezone.utc) - timedelta(days=7)).isoformat()
    rows = []
    for a in agents:
        base = {"agent_id": a["id"], "started_at": {"$gte": cutoff}, **f}
        answered = await db.calls.count_documents({**base, "disposition": "answered"})
        missed = await db.calls.count_documents({**base, "disposition": {"$in": ["missed", "abandoned"]}})
        rows.append({"agent_id": a["id"], "agent_name": a["name"], "avatar": a.get("avatar"),
                     "status": a.get("status"), "answered_7d": answered, "missed_7d": missed,
                     "avg_handle_sec": a.get("avg_handle_sec", 0), "csat": a.get("csat", 0),
                     "adherence_pct": a.get("adherence_pct", 0), "calls_handled": a.get("calls_handled", 0)})
    rows.sort(key=lambda r: r["answered_7d"], reverse=True)
    return {"rows": rows}

@api.get("/reports/types")
async def reports_types(user: dict = Depends(require_permission("reports.view"))):
    return {"types": [
        {"key": "agents", "label": "Performance de Agentes"},
        {"key": "queues", "label": "Chamadas por Fila"},
        {"key": "calls", "label": "Histórico de Chamadas (CDR)"},
        {"key": "abandoned", "label": "Chamadas Abandonadas"},
        {"key": "recordings", "label": "Gravações"},
        {"key": "hourly", "label": "Produtividade Horária"},
        {"key": "sla", "label": "SLA por Fila"},
        {"key": "agent_states", "label": "Estados de Agente"},
        {"key": "heatmap", "label": "Heatmap (Dia × Hora)"},
        {"key": "compare", "label": "Comparativo de Períodos"},
    ]}


class QueueSlaTarget(BaseModel):
    sla_target_sec: int = 20


@api.put("/queues/{queue_id}/sla")
async def update_queue_sla(queue_id: str, body: QueueSlaTarget,
                           user: dict = Depends(require_permission("queues.manage"))):
    f = tenant_filter(user)
    res = await db.queues.update_one({"id": queue_id, **f},
                                     {"$set": {"sla_target_sec": int(body.sla_target_sec)}})
    if res.matched_count == 0:
        raise HTTPException(status_code=404, detail="Fila não encontrada")
    await write_audit(user, "update", "queue", queue_id, "SLA target",
                      {"sla_target_sec": body.sla_target_sec})
    return {"ok": True, "sla_target_sec": body.sla_target_sec}

@api.get("/reports/data")
async def reports_data(type: str, period: str = "7d", agent_id: Optional[str] = None, queue_id: Optional[str] = None,
                       user: dict = Depends(require_permission("reports.view"))):
    return await _build_report(user, type, period, agent_id, queue_id)

@api.get("/reports/export")
async def reports_export(type: str, format: str = "xlsx", period: str = "7d",
                         agent_id: Optional[str] = None, queue_id: Optional[str] = None,
                         user: dict = Depends(require_permission("reports.export"))):
    from fastapi.responses import StreamingResponse
    import io
    data = await _build_report(user, type, period, agent_id, queue_id)
    filename_base = f"{type}_{period}_{datetime.now().strftime('%Y%m%d_%H%M')}"
    if format == "xlsx":
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment
        wb = Workbook(); ws = wb.active; ws.title = data["title"][:30]
        ws.append([data["title"]]); ws["A1"].font = Font(bold=True, size=14)
        ws.append([f"Gerado em: {datetime.now().strftime('%d/%m/%Y %H:%M')}  ·  Período: {period}"])
        ws.append([])
        ws.append([c["label"] for c in data["columns"]])
        for cell in ws[ws.max_row]:
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill("solid", fgColor="09090B")
            cell.alignment = Alignment(horizontal="left")
        for row in data["rows"]:
            ws.append([row.get(c["key"], "") for c in data["columns"]])
        for col in ws.columns:
            max_len = max((len(str(c.value)) if c.value is not None else 0) for c in col)
            ws.column_dimensions[col[0].column_letter].width = min(max_len + 2, 40)
        buf = io.BytesIO(); wb.save(buf); buf.seek(0)
        return StreamingResponse(buf, media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                                 headers={"Content-Disposition": f'attachment; filename="{filename_base}.xlsx"'})
    if format == "pdf":
        from reportlab.lib.pagesizes import A4, landscape
        from reportlab.lib import colors
        from reportlab.lib.styles import getSampleStyleSheet
        from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
        buf = io.BytesIO()
        doc = SimpleDocTemplate(buf, pagesize=landscape(A4), leftMargin=20, rightMargin=20, topMargin=25, bottomMargin=20)
        styles = getSampleStyleSheet(); story = []
        story.append(Paragraph(f"<b>{data['title']}</b>", styles["Title"]))
        story.append(Paragraph(f"Gerado em: {datetime.now().strftime('%d/%m/%Y %H:%M')}  ·  Período: {period}", styles["Normal"]))
        story.append(Spacer(1, 12))
        headers = [c["label"] for c in data["columns"]]
        table_data = [headers]
        for row in data["rows"]:
            table_data.append([str(row.get(c["key"], "")) for c in data["columns"]])
        if len(table_data) == 1:
            table_data.append(["Sem dados para o período selecionado."] + [""] * (len(headers) - 1))
        tbl = Table(table_data, repeatRows=1)
        tbl.setStyle(TableStyle([
            ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#09090B")),
            ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
            ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
            ("FONTSIZE", (0, 0), (-1, -1), 8),
            ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#FAFAFA")]),
            ("GRID", (0, 0), (-1, -1), 0.25, colors.HexColor("#E4E4E7")),
            ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
            ("LEFTPADDING", (0, 0), (-1, -1), 5), ("RIGHTPADDING", (0, 0), (-1, -1), 5),
            ("TOPPADDING", (0, 0), (-1, -1), 4), ("BOTTOMPADDING", (0, 0), (-1, -1), 4),
        ]))
        story.append(tbl); doc.build(story); buf.seek(0)
        return StreamingResponse(buf, media_type="application/pdf",
                                 headers={"Content-Disposition": f'attachment; filename="{filename_base}.pdf"'})
    raise HTTPException(status_code=400, detail="Formato inválido. Use xlsx ou pdf.")

# ---------- Billing: Charges (Asaas + PayPal) ----------
class ChargeCreate(BaseModel):
    tenant_id: str
    gateway: str  # asaas | paypal
    method: str   # pix | boleto | credit_card | paypal
    amount: float
    description: Optional[str] = None
    customer_name: Optional[str] = None
    customer_email: Optional[str] = None
    customer_cpf_cnpj: Optional[str] = None  # required for Asaas
    customer_phone: Optional[str] = None
    due_date: Optional[str] = None  # YYYY-MM-DD; defaults to +7d for Asaas
    currency: str = "BRL"  # PayPal


def _serialize_charge(c: dict) -> dict:
    return {k: c.get(k) for k in [
        "id", "tenant_id", "gateway", "method", "amount", "currency", "status",
        "description", "external_id", "invoice_url", "checkout_url",
        "pix_qrcode", "pix_payload", "boleto_url", "barcode",
        "customer_name", "customer_email", "due_date", "paid_at",
        "created_at", "updated_at",
    ] if c.get(k) is not None}


async def _load_billing_settings() -> dict:
    return await db.billing_settings.find_one({"id": "global"}, {"_id": 0}) or {}


async def _get_asaas_client() -> AsaasClient:
    s = await _load_billing_settings()
    api_key = s.get("asaas_api_key")
    if not api_key:
        raise HTTPException(status_code=400, detail="Asaas não configurado. Vá em Cobrança → Asaas.")
    env = s.get("asaas_environment", "sandbox")
    return AsaasClient(api_key=api_key, environment=env)


async def _get_paypal_client() -> PayPalClient:
    s = await _load_billing_settings()
    cid = s.get("paypal_client_id")
    secret = s.get("paypal_client_secret")
    if not cid or not secret:
        raise HTTPException(status_code=400, detail="PayPal não configurado. Vá em Cobrança → PayPal.")
    env = s.get("paypal_environment", "sandbox")
    return PayPalClient(client_id=cid, client_secret=secret, environment=env)


@api.post("/billing/charges")
async def create_charge(body: ChargeCreate, user: dict = Depends(require_super_admin())):
    tenant = await db.tenants.find_one({"id": body.tenant_id})
    if not tenant:
        raise HTTPException(status_code=404, detail="Tenant não encontrado")

    cid = str(uuid.uuid4())
    now = datetime.now(timezone.utc).isoformat()
    base_doc = {
        "id": cid, "tenant_id": body.tenant_id,
        "gateway": body.gateway, "method": body.method,
        "amount": float(body.amount), "currency": body.currency,
        "description": body.description or f"Mensalidade {tenant.get('name','')}",
        "customer_name": body.customer_name or tenant.get("name"),
        "customer_email": body.customer_email,
        "customer_cpf_cnpj": body.customer_cpf_cnpj,
        "due_date": body.due_date,
        "status": "pending",
        "created_at": now, "updated_at": now,
    }

    try:
        if body.gateway == "asaas":
            method_map = {"pix": "PIX", "boleto": "BOLETO", "credit_card": "CREDIT_CARD"}
            billing_type = method_map.get(body.method)
            if not billing_type:
                raise HTTPException(status_code=400, detail="Método inválido para Asaas. Use pix/boleto/credit_card.")
            if not body.customer_cpf_cnpj:
                raise HTTPException(status_code=400, detail="CPF/CNPJ do cliente é obrigatório para Asaas")
            client = await _get_asaas_client()
            cpf = "".join(ch for ch in body.customer_cpf_cnpj if ch.isdigit())
            customer = await client.find_customer_by_cpf(cpf)
            if not customer:
                customer = await client.create_customer(
                    name=base_doc["customer_name"] or "Cliente", cpf_cnpj=cpf,
                    email=body.customer_email, mobile_phone=body.customer_phone,
                )
            due = body.due_date or (datetime.now(timezone.utc) + timedelta(days=7)).strftime("%Y-%m-%d")
            payment = await client.create_payment(
                customer_id=customer["id"], billing_type=billing_type,
                value=float(body.amount), due_date=due,
                description=base_doc["description"], external_reference=cid,
            )
            base_doc["external_id"] = payment.get("id")
            base_doc["external_customer_id"] = customer["id"]
            base_doc["invoice_url"] = payment.get("invoiceUrl")
            base_doc["status"] = map_asaas_status(payment.get("status", "PENDING"))
            base_doc["due_date"] = due
            # Get payment-method-specific data
            if billing_type == "PIX":
                try:
                    qr = await client.get_pix_qrcode(payment["id"])
                    base_doc["pix_qrcode"] = qr.get("encodedImage")  # base64 PNG
                    base_doc["pix_payload"] = qr.get("payload")
                except Exception as e:
                    logger.warning("PIX QR fetch failed: %s", e)
            elif billing_type == "BOLETO":
                base_doc["boleto_url"] = payment.get("bankSlipUrl")
                try:
                    bd = await client.get_boleto_url(payment["id"])
                    base_doc["barcode"] = bd.get("identificationField")
                except Exception: pass
            base_doc["raw"] = payment

        elif body.gateway == "paypal":
            client_pp = await _get_paypal_client()
            order = await client_pp.create_order(
                amount=float(body.amount), currency=body.currency,
                reference_id=cid, description=base_doc["description"],
            )
            base_doc["external_id"] = order.get("id")
            base_doc["status"] = map_paypal_status(order.get("status", "CREATED"))
            for link in order.get("links", []):
                if link.get("rel") == "approve":
                    base_doc["checkout_url"] = link.get("href")
                    break
            base_doc["raw"] = order
        else:
            raise HTTPException(status_code=400, detail="Gateway inválido. Use asaas ou paypal.")
    except (AsaasError, PayPalError) as e:
        raise HTTPException(status_code=502, detail=str(e))

    await db.charges.insert_one(base_doc)
    await write_audit(user, "create", "charge", cid,
                      f"{body.gateway}/{body.method} R$ {body.amount:.2f}",
                      {"tenant_id": body.tenant_id, "gateway": body.gateway})
    return _serialize_charge(base_doc)

@api.get("/billing/charges")
async def list_charges(
    user: dict = Depends(require_super_admin()),
    tenant_id: Optional[str] = None, status: Optional[str] = None,
    limit: int = Query(100, le=500),
):
    q: Dict[str, Any] = {}
    if tenant_id: q["tenant_id"] = tenant_id
    if status: q["status"] = status
    docs = await db.charges.find(q, {"_id": 0, "raw": 0}).sort("created_at", -1).to_list(limit)
    return {"charges": [_serialize_charge(c) for c in docs]}


@api.get("/billing/charges/{cid}")
async def get_charge(cid: str, user: dict = Depends(require_super_admin())):
    c = await db.charges.find_one({"id": cid}, {"_id": 0, "raw": 0})
    if not c: raise HTTPException(status_code=404, detail="Cobrança não encontrada")
    return _serialize_charge(c)


@api.post("/billing/charges/{cid}/sync")
async def sync_charge(cid: str, user: dict = Depends(require_super_admin())):
    """Pulls fresh status from gateway."""
    c = await db.charges.find_one({"id": cid})
    if not c: raise HTTPException(status_code=404, detail="Cobrança não encontrada")
    try:
        if c["gateway"] == "asaas":
            client = await _get_asaas_client()
            payment = await client.get_payment(c["external_id"])
            new_status = map_asaas_status(payment.get("status", "PENDING"))
        elif c["gateway"] == "paypal":
            client_pp = await _get_paypal_client()
            order = await client_pp.get_order(c["external_id"])
            new_status = map_paypal_status(order.get("status", "CREATED"))
        else:
            raise HTTPException(status_code=400, detail="Gateway inválido")
    except (AsaasError, PayPalError) as e:
        raise HTTPException(status_code=502, detail=str(e))

    update = {"status": new_status, "updated_at": datetime.now(timezone.utc).isoformat()}
    if new_status == "paid" and not c.get("paid_at"):
        update["paid_at"] = update["updated_at"]
        # Also mark tenant as paid
        await db.tenants.update_one({"id": c["tenant_id"]}, {"$set": {"payment_status": "paid"}})
    await db.charges.update_one({"id": cid}, {"$set": update})
    fresh = await db.charges.find_one({"id": cid}, {"_id": 0, "raw": 0})
    return _serialize_charge(fresh)


@api.post("/billing/charges/{cid}/capture")
async def capture_paypal_charge(cid: str, user: dict = Depends(require_super_admin())):
    """For PayPal orders that have been approved by buyer; capture funds."""
    c = await db.charges.find_one({"id": cid})
    if not c: raise HTTPException(status_code=404, detail="Cobrança não encontrada")
    if c["gateway"] != "paypal":
        raise HTTPException(status_code=400, detail="Capture é apenas para PayPal")
    try:
        client_pp = await _get_paypal_client()
        result = await client_pp.capture_order(c["external_id"])
    except PayPalError as e:
        raise HTTPException(status_code=502, detail=str(e))
    new_status = map_paypal_status(result.get("status", "COMPLETED"))
    update = {"status": new_status, "updated_at": datetime.now(timezone.utc).isoformat()}
    if new_status == "paid":
        update["paid_at"] = update["updated_at"]
        await db.tenants.update_one({"id": c["tenant_id"]}, {"$set": {"payment_status": "paid"}})
    await db.charges.update_one({"id": cid}, {"$set": update})
    return {"ok": True, "status": new_status}


# ---------- Webhooks ----------
@api.post("/webhooks/asaas")
async def webhook_asaas(request: Request):
    """Asaas posts JSON. Authenticated via header `asaas-access-token`."""
    settings = await _load_billing_settings()
    expected_token = settings.get("asaas_webhook_token")
    sent_token = request.headers.get("asaas-access-token") or request.headers.get("Asaas-Access-Token")
    if expected_token and sent_token != expected_token:
        raise HTTPException(status_code=401, detail="Webhook token inválido")
    try:
        body = await request.json()
    except Exception:
        raise HTTPException(status_code=400, detail="Corpo JSON inválido")
    if not isinstance(body, dict):
        raise HTTPException(status_code=400, detail="Esperado um objeto JSON")
    event = body.get("event", "")
    payment = body.get("payment", {}) or {}
    delivery_id = body.get("id") or payment.get("id")
    if delivery_id and await db.webhook_events.find_one({"delivery_id": delivery_id, "source": "asaas"}):
        return {"ok": True, "duplicate": True}
    asaas_pid = payment.get("id")
    if asaas_pid:
        c = await db.charges.find_one({"external_id": asaas_pid, "gateway": "asaas"})
        if c:
            new_status = map_asaas_status(payment.get("status", "PENDING"))
            update = {"status": new_status, "updated_at": datetime.now(timezone.utc).isoformat()}
            if new_status == "paid":
                update["paid_at"] = update["updated_at"]
                await db.tenants.update_one({"id": c["tenant_id"]}, {"$set": {"payment_status": "paid"}})
            await db.charges.update_one({"id": c["id"]}, {"$set": update})
    await db.webhook_events.insert_one({
        "id": str(uuid.uuid4()), "source": "asaas", "delivery_id": delivery_id,
        "event": event, "payload": body,
        "created_at": datetime.now(timezone.utc).isoformat(),
    })
    return {"ok": True}


@api.post("/webhooks/paypal")
async def webhook_paypal(request: Request):
    """PayPal sends signed events. We store + try to update charge status."""
    try:
        body = await request.json()
    except Exception:
        raise HTTPException(status_code=400, detail="Corpo JSON inválido")
    if not isinstance(body, dict):
        raise HTTPException(status_code=400, detail="Esperado um objeto JSON")
    event_type = body.get("event_type", "")
    resource = body.get("resource", {}) or {}
    delivery_id = body.get("id")
    if delivery_id and await db.webhook_events.find_one({"delivery_id": delivery_id, "source": "paypal"}):
        return {"ok": True, "duplicate": True}
    # Try to extract order ID; events: PAYMENT.CAPTURE.COMPLETED, CHECKOUT.ORDER.APPROVED, etc.
    order_id = None
    if "supplementary_data" in resource:
        rel = resource.get("supplementary_data", {}).get("related_ids", {})
        order_id = rel.get("order_id")
    order_id = order_id or resource.get("id")
    if order_id:
        c = await db.charges.find_one({"external_id": order_id, "gateway": "paypal"})
        if c:
            status_in_event = resource.get("status") or ("COMPLETED" if "CAPTURE.COMPLETED" in event_type else None)
            new_status = map_paypal_status(status_in_event or "")
            if new_status:
                update = {"status": new_status, "updated_at": datetime.now(timezone.utc).isoformat()}
                if new_status == "paid":
                    update["paid_at"] = update["updated_at"]
                    await db.tenants.update_one({"id": c["tenant_id"]}, {"$set": {"payment_status": "paid"}})
                await db.charges.update_one({"id": c["id"]}, {"$set": update})
    await db.webhook_events.insert_one({
        "id": str(uuid.uuid4()), "source": "paypal", "delivery_id": delivery_id,
        "event": event_type, "payload": body,
        "created_at": datetime.now(timezone.utc).isoformat(),
    })
    return {"ok": True}


# ---------- FusionPBX Integration ----------
class FusionPBXSettings(BaseModel):
    enabled: bool = False
    # Connection type: "rest" (default) or "db" (direct PostgreSQL)
    connection_type: str = "rest"
    # REST mode fields
    base_url: str = ""
    api_key: Optional[str] = None
    username: Optional[str] = None
    password: Optional[str] = None
    verify_ssl: bool = True
    path_extensions: Optional[str] = None
    path_queues: Optional[str] = None
    path_agents: Optional[str] = None
    path_cdr: Optional[str] = None
    # DB mode fields
    db_host: Optional[str] = None
    db_port: int = 5432
    db_name: str = "fusionpbx"
    db_username: Optional[str] = None
    db_password: Optional[str] = None
    db_ssl: bool = False
    # ESL (Event Socket) — for live calls
    esl_host: Optional[str] = None
    esl_port: int = 8021
    esl_password: Optional[str] = None
    esl_timeout: float = 5.0
    # SFTP — para baixar gravações
    sftp_host: Optional[str] = None
    sftp_port: int = 22
    sftp_username: Optional[str] = None
    sftp_password: Optional[str] = None
    sftp_private_key: Optional[str] = None
    sftp_recordings_path: Optional[str] = None  # ex: /var/lib/freeswitch/recordings
    # Common
    domain_uuid: Optional[str] = None
    domain_name: Optional[str] = None
    sync_interval_minutes: int = 1


def _serialize_fusion_settings(s: Optional[dict], mask: bool = True) -> dict:
    if not s: return {"enabled": False, "configured": False, "connection_type": "rest"}
    out = {k: s.get(k) for k in [
        "enabled", "connection_type",
        "base_url", "username", "domain_uuid", "domain_name",
        "verify_ssl", "sync_interval_minutes", "last_sync_at", "last_sync_status",
        "path_extensions", "path_queues", "path_agents", "path_cdr",
        "db_host", "db_port", "db_name", "db_username", "db_ssl",
        "esl_host", "esl_port", "esl_timeout",
        "sftp_host", "sftp_port", "sftp_username", "sftp_recordings_path",
    ]}
    out["connection_type"] = s.get("connection_type") or "rest"
    if out["connection_type"] == "db":
        out["configured"] = bool(s.get("db_host") and s.get("db_username"))
    else:
        out["configured"] = bool(s.get("base_url"))
    out["api_key_set"] = bool(s.get("api_key"))
    out["password_set"] = bool(s.get("password"))
    out["db_password_set"] = bool(s.get("db_password"))
    out["esl_password_set"] = bool(s.get("esl_password"))
    out["esl_configured"] = bool(s.get("esl_host"))
    out["sftp_password_set"] = bool(s.get("sftp_password"))
    out["sftp_key_set"] = bool(s.get("sftp_private_key"))
    out["sftp_configured"] = bool(s.get("sftp_host") and s.get("sftp_username") and (s.get("sftp_password") or s.get("sftp_private_key")))
    if not mask:
        out["api_key"] = s.get("api_key")
        out["password"] = s.get("password")
        out["db_password"] = s.get("db_password")
        out["esl_password"] = s.get("esl_password")
        out["sftp_password"] = s.get("sftp_password")
        out["sftp_private_key"] = s.get("sftp_private_key")
    return out


async def _resolve_tenant_for_fusion(user: dict, tenant_id: Optional[str]) -> str:
    if user.get("role") == "super_admin":
        # Accept explicit ?tenant_id=... OR the X-Tenant-Context header (impersonation)
        tid = tenant_id or user.get("_tenant_context")
        if not tid:
            raise HTTPException(status_code=400, detail="Selecione um tenant antes de configurar a central PBX (entre no tenant via página Tenants)")
        return tid
    return user["tenant_id"]


# ---------- System Updates (Super Admin Self-Host Tool) ----------
import subprocess
import asyncio

_UPDATE_STATE: Dict[str, Any] = {"running": False, "log": [], "started_at": None, "finished_at": None, "success": None}
APP_ROOT = Path("/opt/CallCenter")
FRONTEND_BUILD = APP_ROOT / "frontend" / "build"

# Application version - manually incremented on each release
APP_VERSION = "V3.0 R126"


def _get_build_version() -> str:
    """Returns a unique identifier of the current frontend build.
    Uses the main.js hash from asset-manifest.json (changes every rebuild).
    Falls back to mtime of index.html."""
    try:
        manifest = FRONTEND_BUILD / "asset-manifest.json"
        if manifest.exists():
            with open(manifest) as f:
                data = json.load(f)
            main_js = data.get("files", {}).get("main.js", "")
            if main_js:
                # main.<hash>.js -> hash
                parts = main_js.split("/")[-1].split(".")
                if len(parts) >= 3:
                    return parts[1]
        idx = FRONTEND_BUILD / "index.html"
        if idx.exists():
            return str(int(idx.stat().st_mtime))
    except Exception:
        pass
    return "dev"


@api.get("/system/version")
async def system_version():
    """Public endpoint so any authenticated client can detect updates."""
    return {"build": _get_build_version(), "version": APP_VERSION}


def _get_git_info() -> Dict[str, Any]:
    """Returns current git commit, branch, and remote status."""
    info: Dict[str, Any] = {"installed": False}
    try:
        if not (APP_ROOT / ".git").exists():
            return info
        info["installed"] = True
        info["branch"] = subprocess.check_output(
            ["git", "-C", str(APP_ROOT), "rev-parse", "--abbrev-ref", "HEAD"], text=True, timeout=10
        ).strip()
        info["commit"] = subprocess.check_output(
            ["git", "-C", str(APP_ROOT), "rev-parse", "--short", "HEAD"], text=True, timeout=10
        ).strip()
        info["commit_full"] = subprocess.check_output(
            ["git", "-C", str(APP_ROOT), "rev-parse", "HEAD"], text=True, timeout=10
        ).strip()
        info["last_commit_message"] = subprocess.check_output(
            ["git", "-C", str(APP_ROOT), "log", "-1", "--pretty=%B"], text=True, timeout=10
        ).strip()
        info["last_commit_date"] = subprocess.check_output(
            ["git", "-C", str(APP_ROOT), "log", "-1", "--pretty=%cI"], text=True, timeout=10
        ).strip()
        try:
            subprocess.check_output(
                ["git", "-C", str(APP_ROOT), "fetch", "--quiet"], text=True, timeout=30,
                stderr=subprocess.STDOUT,
            )
            info["fetch_ok"] = True
        except Exception as e:
            info["fetch_ok"] = False
            info["fetch_error"] = str(e)
        try:
            behind = subprocess.check_output(
                ["git", "-C", str(APP_ROOT), "rev-list", "--count", "HEAD..@{u}"],
                text=True, timeout=10, stderr=subprocess.DEVNULL,
            ).strip()
            info["commits_behind"] = int(behind) if behind else 0
            info["has_updates"] = info["commits_behind"] > 0
        except Exception:
            info["commits_behind"] = 0
            info["has_updates"] = False
    except Exception as e:
        info["error"] = str(e)
    return info


async def _run_update_task():
    """Background task that runs git pull + pip install + yarn build + supervisor restart."""
    def append(line: str):
        _UPDATE_STATE["log"].append({"t": datetime.now(timezone.utc).isoformat(), "line": line})
        if len(_UPDATE_STATE["log"]) > 1000:
            _UPDATE_STATE["log"] = _UPDATE_STATE["log"][-500:]

    async def run_cmd(cmd: str, cwd: Optional[str] = None, timeout: int = 600) -> int:
        append(f"$ {cmd}")
        proc = await asyncio.create_subprocess_shell(
            cmd, cwd=cwd, stdout=asyncio.subprocess.PIPE, stderr=asyncio.subprocess.STDOUT,
        )
        try:
            while True:
                line = await asyncio.wait_for(proc.stdout.readline(), timeout=timeout)
                if not line: break
                txt = line.decode("utf-8", errors="replace").rstrip()
                if txt: append(txt)
        except asyncio.TimeoutError:
            proc.kill(); append(f"❌ Timeout após {timeout}s"); return 124
        rc = await proc.wait()
        append(f"[exit {rc}]")
        return rc

    _UPDATE_STATE.update({"running": True, "log": [], "started_at": datetime.now(timezone.utc).isoformat(),
                          "finished_at": None, "success": None})
    try:
        append("🚀 Iniciando atualização…")
        if not (APP_ROOT / ".git").exists():
            append(f"❌ {APP_ROOT} não é um repositório Git. Clone via git para habilitar updates.")
            _UPDATE_STATE["success"] = False
            return
        # Pre-flight: check write permissions
        import os as _os
        if not _os.access(str(APP_ROOT), _os.W_OK):
            append(f"❌ Sem permissão de escrita em {APP_ROOT}")
            append(f"   Rode na VPS: sudo chown -R voxyra:voxyra {APP_ROOT}")
            _UPDATE_STATE["success"] = False
            return
        # Pre-flight: check sudo rights for supervisorctl (try common paths)
        sudo_ok = False
        sudoctl_path = None
        for path in ("supervisorctl", "/usr/bin/supervisorctl", "/usr/local/bin/supervisorctl"):
            check = await asyncio.create_subprocess_shell(
                f"sudo -n {path} status CallCenter-backend",
                stdout=asyncio.subprocess.PIPE, stderr=asyncio.subprocess.PIPE,
            )
            sout, serr = await check.communicate()
            out_combined = (sout or b"") + (serr or b"")
            if check.returncode == 0 or b"RUNNING" in out_combined or b"STOPPED" in out_combined:
                sudo_ok = True
                sudoctl_path = path
                break
        if not sudo_ok:
            # Discover actual path to show in the hint
            which = await asyncio.create_subprocess_shell(
                "which supervisorctl", stdout=asyncio.subprocess.PIPE
            )
            wout, _ = await which.communicate()
            actual = wout.decode().strip() if wout else "/usr/bin/supervisorctl"
            append("❌ Sem permissão sudo para reiniciar o supervisor.")
            append(f"   Supervisorctl detectado em: {actual}")
            append("   Rode na VPS (uma vez):")
            append("   sudo tee /etc/sudoers.d/CallCenter-webupdate > /dev/null <<'EOF'")
            append(f"   voxyra ALL=(ALL) NOPASSWD: {actual}")
            append(f"   voxyra ALL=(ALL) NOPASSWD: /usr/bin/supervisorctl")
            append(f"   voxyra ALL=(ALL) NOPASSWD: /usr/local/bin/supervisorctl")
            append("   EOF")
            append("   sudo chmod 0440 /etc/sudoers.d/CallCenter-webupdate")
            _UPDATE_STATE["success"] = False
            return
        # Configure git safe.directory (idempotent, no permission needed for user config)
        await run_cmd(f"git config --global --add safe.directory {APP_ROOT}", timeout=10)
        # git pull
        rc = await run_cmd(f"git -C {APP_ROOT} pull", timeout=120)
        if rc != 0: _UPDATE_STATE["success"] = False; return
        # pip install
        rc = await run_cmd(
            f"bash -c 'cd {APP_ROOT}/backend && source venv/bin/activate && "
            f"grep -vE \"^(emergentintegrations)\" requirements.txt > /tmp/req.txt && "
            f"pip install -r /tmp/req.txt --quiet --disable-pip-version-check'",
            timeout=600,
        )
        if rc != 0: _UPDATE_STATE["success"] = False; return
        # yarn install in place (doesn't affect the running frontend)
        rc = await run_cmd(f"cd {APP_ROOT}/frontend && yarn install --silent", timeout=600)
        if rc != 0: _UPDATE_STATE["success"] = False; return
        # ATOMIC BUILD: build to temporary BUILD_DIR then swap
        # This prevents the user's browser from seeing a partial/deleted build/ folder
        build_old = APP_ROOT / "frontend" / "build"
        build_new = APP_ROOT / "frontend" / "build.new"
        build_backup = APP_ROOT / "frontend" / "build.old"
        append("🏗️  Fazendo build em pasta temporária (build.new)...")
        await run_cmd(f"rm -rf {build_new} {build_backup}", timeout=30)
        rc = await run_cmd(
            f"cd {APP_ROOT}/frontend && BUILD_PATH=./build.new yarn build",
            timeout=900,
        )
        if rc != 0:
            append("❌ Build falhou. O frontend atual permanece intocado.")
            await run_cmd(f"rm -rf {build_new}", timeout=30)
            _UPDATE_STATE["success"] = False; return
        # Validate build.new has index.html and static/js
        if not (build_new / "index.html").exists():
            append("❌ Build gerado está incompleto (sem index.html). Abortando swap.")
            await run_cmd(f"rm -rf {build_new}", timeout=30)
            _UPDATE_STATE["success"] = False; return
        # Atomic swap: mv build -> build.old, mv build.new -> build
        append("🔀 Swap atômico: build.new → build")
        if build_old.exists():
            await run_cmd(f"mv {build_old} {build_backup}", timeout=10)
        await run_cmd(f"mv {build_new} {build_old}", timeout=10)
        # Clean old backup after successful swap
        await run_cmd(f"rm -rf {build_backup}", timeout=30)
        append("")
        append("✅ Código atualizado com sucesso!")
        append("🔄 Reiniciando backend via supervisor…")
        append("   (a página ficará indisponível por ~5 segundos)")
        _UPDATE_STATE["success"] = True
        _UPDATE_STATE["finished_at"] = datetime.now(timezone.utc).isoformat()
        _UPDATE_STATE["running"] = False
        # Schedule restart after 1s so the HTTP response is delivered
        await asyncio.sleep(1)
        restart_cmd = f"sudo -n {sudoctl_path or 'supervisorctl'} restart CallCenter-backend"
        await asyncio.create_subprocess_shell(f"{restart_cmd} &")
    except Exception as e:
        append(f"❌ Erro: {e}")
        _UPDATE_STATE["success"] = False
    finally:
        _UPDATE_STATE["finished_at"] = datetime.now(timezone.utc).isoformat()
        _UPDATE_STATE["running"] = False


@api.get("/system/info")
async def system_info(user: dict = Depends(require_super_admin())):
    """Returns git/version info for the update panel."""
    git = _get_git_info()
    return {
        "app_version": APP_VERSION,
        "app_name": "Voxyra CCA",
        "app_dir": str(APP_ROOT),
        "git": git,
        "update_state": {k: _UPDATE_STATE.get(k) for k in ("running", "started_at", "finished_at", "success")},
    }


@api.get("/system/update/preflight")
async def system_update_preflight(user: dict = Depends(require_super_admin())):
    """Check all permissions/prerequisites needed for web-based update. Returns fix commands."""
    import os as _os, pwd
    checks: List[Dict[str, Any]] = []
    # 1. Git installed?
    r = await asyncio.create_subprocess_shell("git --version", stdout=asyncio.subprocess.PIPE, stderr=asyncio.subprocess.PIPE)
    await r.communicate()
    checks.append({"name": "Git instalado", "ok": r.returncode == 0, "fix": "sudo apt install -y git"})
    # 2. APP_ROOT é repo git?
    is_git = (APP_ROOT / ".git").exists()
    checks.append({"name": f"{APP_ROOT} é repositório Git", "ok": is_git,
                   "fix": f"Clone via: sudo rm -rf {APP_ROOT} && cd /opt && sudo git clone <URL> {APP_ROOT.name}"})
    # 3. Owner
    try:
        stat = APP_ROOT.stat()
        owner = pwd.getpwuid(stat.st_uid).pw_name
        current_user = pwd.getpwuid(_os.getuid()).pw_name
        checks.append({"name": f"Dono do diretório ({owner}) = usuário do backend ({current_user})",
                       "ok": owner == current_user,
                       "fix": f"sudo chown -R {current_user}:{current_user} {APP_ROOT}"})
    except Exception as e:
        checks.append({"name": "Verificação de dono", "ok": False, "fix": str(e)})
    # 4. Escrita em APP_ROOT
    checks.append({"name": f"Permissão de escrita em {APP_ROOT}",
                   "ok": _os.access(str(APP_ROOT), _os.W_OK),
                   "fix": f"sudo chown -R voxyra:voxyra {APP_ROOT}"})
    # 5. sudoers supervisorctl
    r = await asyncio.create_subprocess_shell(
        "sudo -n supervisorctl status CallCenter-backend 2>&1",
        stdout=asyncio.subprocess.PIPE, stderr=asyncio.subprocess.STDOUT,
    )
    out, _ = await r.communicate()
    out_txt = out.decode("utf-8", errors="replace") if out else ""
    sudo_ok = r.returncode == 0 or "RUNNING" in out_txt or "STOPPED" in out_txt
    checks.append({
        "name": "Sudo sem senha para supervisorctl",
        "ok": sudo_ok,
        "fix": (
            "sudo tee /etc/sudoers.d/CallCenter-webupdate > /dev/null <<'EOF'\n"
            "voxyra ALL=(ALL) NOPASSWD: /usr/bin/supervisorctl restart CallCenter-backend\n"
            "voxyra ALL=(ALL) NOPASSWD: /usr/bin/supervisorctl reload\n"
            "EOF\n"
            "sudo chmod 0440 /etc/sudoers.d/CallCenter-webupdate"
        ),
    })
    # 6. Git safe.directory
    r = await asyncio.create_subprocess_shell(
        f"git config --global --get-all safe.directory",
        stdout=asyncio.subprocess.PIPE, stderr=asyncio.subprocess.PIPE,
    )
    sdo, _ = await r.communicate()
    safe_ok = str(APP_ROOT) in (sdo.decode() if sdo else "")
    checks.append({"name": "git safe.directory configurado", "ok": safe_ok,
                   "fix": f"git config --global --add safe.directory {APP_ROOT}"})
    # 7. yarn
    r = await asyncio.create_subprocess_shell("yarn --version", stdout=asyncio.subprocess.PIPE, stderr=asyncio.subprocess.PIPE)
    await r.communicate()
    checks.append({"name": "Yarn instalado", "ok": r.returncode == 0, "fix": "sudo npm install -g yarn"})

    all_ok = all(c["ok"] for c in checks)
    return {"all_ok": all_ok, "checks": checks}


@api.post("/system/update/check")
async def system_update_check(user: dict = Depends(require_super_admin())):
    info = _get_git_info()
    if not info.get("installed"):
        raise HTTPException(status_code=400, detail=f"{APP_ROOT} não é um repositório Git.")
    return {"branch": info.get("branch"), "current": info.get("commit"),
            "has_updates": info.get("has_updates", False),
            "commits_behind": info.get("commits_behind", 0),
            "fetch_ok": info.get("fetch_ok", False),
            "fetch_error": info.get("fetch_error")}


@api.post("/system/update/run")
async def system_update_run(user: dict = Depends(require_super_admin())):
    if _UPDATE_STATE.get("running"):
        raise HTTPException(status_code=409, detail="Uma atualização já está em andamento")
    if not (APP_ROOT / ".git").exists():
        raise HTTPException(status_code=400, detail=f"{APP_ROOT} não é um repositório Git. Clone via git para habilitar updates pela web.")
    await write_audit(user, "update", "system", "app", "Atualização via web", {"started_at": datetime.now(timezone.utc).isoformat()})
    asyncio.create_task(_run_update_task())
    return {"ok": True, "message": "Atualização iniciada"}


@api.get("/system/update/status")
async def system_update_status(user: dict = Depends(require_super_admin()), since: int = 0):
    log = _UPDATE_STATE.get("log", [])
    return {
        "running": _UPDATE_STATE.get("running", False),
        "success": _UPDATE_STATE.get("success"),
        "started_at": _UPDATE_STATE.get("started_at"),
        "finished_at": _UPDATE_STATE.get("finished_at"),
        "log": log[since:],
        "total_lines": len(log),
    }


# ---------- FusionPBX Integration ----------


@api.get("/fusionpbx/settings")
async def get_fusion_settings(user: dict = Depends(get_current_user), tenant_id: Optional[str] = None):
    tid = await _resolve_tenant_for_fusion(user, tenant_id)
    if user.get("role") not in ("super_admin", "admin"):
        raise HTTPException(status_code=403, detail="Sem permissão")
    s = await db.fusionpbx_settings.find_one({"tenant_id": tid}, {"_id": 0})
    return _serialize_fusion_settings(s)


@api.put("/fusionpbx/settings")
async def put_fusion_settings(body: FusionPBXSettings, user: dict = Depends(get_current_user),
                              tenant_id: Optional[str] = None):
    tid = await _resolve_tenant_for_fusion(user, tenant_id)
    if user.get("role") not in ("super_admin", "admin"):
        raise HTTPException(status_code=403, detail="Sem permissão")
    payload = body.dict(exclude_unset=True)
    existing = await db.fusionpbx_settings.find_one({"tenant_id": tid}) or {}
    # Don't overwrite secrets if a masked/empty value was sent and one already exists
    for k in ("api_key", "password", "db_password", "esl_password",
              "sftp_password", "sftp_private_key"):
        if k in payload and not payload[k]:
            payload[k] = existing.get(k)
    payload["tenant_id"] = tid
    payload["updated_at"] = datetime.now(timezone.utc).isoformat()
    await db.fusionpbx_settings.update_one({"tenant_id": tid}, {"$set": payload}, upsert=True)
    await write_audit(user, "update", "fusionpbx_settings", tid, "FusionPBX config",
                      {"base_url": payload.get("base_url"), "enabled": payload.get("enabled")})
    s = await db.fusionpbx_settings.find_one({"tenant_id": tid}, {"_id": 0})
    return _serialize_fusion_settings(s)


async def _build_fusion_client(tid: str):
    """Returns either FusionPBXClient (REST) or FusionPBXDBClient based on connection_type."""
    s = await db.fusionpbx_settings.find_one({"tenant_id": tid})
    if not s:
        raise HTTPException(status_code=400, detail="FusionPBX não configurado para este tenant")
    if not s.get("enabled", False):
        raise HTTPException(status_code=400, detail="Integração FusionPBX desativada")
    ctype = s.get("connection_type") or "rest"
    if ctype == "db":
        if not s.get("db_host") or not s.get("db_username"):
            raise HTTPException(status_code=400, detail="Configuração DB incompleta (host/usuário)")
        return FusionPBXDBClient(
            host=s["db_host"], port=int(s.get("db_port") or 5432),
            database=s.get("db_name") or "fusionpbx",
            username=s["db_username"], password=s.get("db_password") or "",
            domain_uuid=s.get("domain_uuid"), ssl=bool(s.get("db_ssl")),
        )
    if not s.get("base_url"):
        raise HTTPException(status_code=400, detail="FusionPBX não configurado para este tenant")
    custom_paths = {k.replace("path_", ""): s[k] for k in
                    ("path_extensions", "path_queues", "path_agents", "path_cdr") if s.get(k)}
    return FusionPBXClient(
        base_url=s["base_url"], api_key=s.get("api_key"),
        username=s.get("username"), password=s.get("password"),
        domain_uuid=s.get("domain_uuid"), domain_name=s.get("domain_name"),
        verify_ssl=bool(s.get("verify_ssl", True)),
        custom_paths=custom_paths,
    )


@api.post("/fusionpbx/sftp/test")
async def fusion_sftp_test(user: dict = Depends(get_current_user), tenant_id: Optional[str] = None):
    tid = await _resolve_tenant_for_fusion(user, tenant_id)
    if user.get("role") not in ("super_admin", "admin"):
        raise HTTPException(status_code=403, detail="Sem permissão")
    s = await db.fusionpbx_settings.find_one({"tenant_id": tid})
    if not s or not s.get("sftp_host"):
        raise HTTPException(status_code=400, detail="Configure SFTP host primeiro")
    # TCP pre-check
    import socket as _socket
    try:
        with _socket.create_connection((s["sftp_host"], int(s.get("sftp_port") or 22)), timeout=5):
            pass
    except Exception as e:
        raise HTTPException(status_code=502,
            detail=f"Não consegui abrir TCP em {s['sftp_host']}:{s.get('sftp_port', 22)} → "
                   f"[{type(e).__name__}] {e}. Verifique firewall e SSH server.")
    # Lista os 5 últimos arquivos da pasta de recordings (smoke test)
    import asyncssh
    base = s.get("sftp_recordings_path") or "/var/lib/freeswitch/recordings"
    domain = s.get("domain_name") or ""
    try:
        kwargs = {"host": s["sftp_host"], "port": int(s.get("sftp_port") or 22),
                  "username": s["sftp_username"], "known_hosts": None, "client_keys": None}
        if s.get("sftp_private_key"):
            kwargs["client_keys"] = [asyncssh.import_private_key(s["sftp_private_key"])]
        elif s.get("sftp_password"):
            kwargs["password"] = s["sftp_password"]
        async with asyncssh.connect(**kwargs) as conn:
            search_dir = f"{base}/{domain}" if domain else base
            result = await conn.run(
                f"find {search_dir} -name '*.wav' -o -name '*.mp3' 2>/dev/null | head -5",
                check=False,
            )
            files = (result.stdout or "").strip().splitlines()
            return {"ok": True, "host": s["sftp_host"], "found": len(files), "samples": files}
    except Exception as e:
        raise HTTPException(status_code=502, detail=f"Falha SFTP [{type(e).__name__}]: {e}")


@api.get("/recordings/{recording_id}/stream")
async def stream_recording_endpoint(recording_id: str, request: Request,
                                    user: dict = Depends(get_current_user)):
    """Streams the recording file (.wav/.mp3) from FusionPBX via SFTP.
    Supports HTTP Range for audio seeking."""
    f = tenant_filter(user)
    rec = await db.recordings.find_one({"id": recording_id, **f}, {"_id": 0})
    if not rec:
        raise HTTPException(status_code=404, detail="Gravação não encontrada")
    # Permission: agent with view_own only sees their own
    if user.get("role") == "agent" and not user.get("permissions"):
        if rec.get("agent_id") != user.get("agent_id"):
            raise HTTPException(status_code=403, detail="Sem permissão para esta gravação")
    allowed_agent_ids = await allowed_agent_ids_for(user)
    if allowed_agent_ids is not None and rec.get("agent_id") not in allowed_agent_ids:
        raise HTTPException(status_code=403, detail="Sem permissão para esta gravação")

    tid = rec.get("tenant_id")
    s = await db.fusionpbx_settings.find_one({"tenant_id": tid})
    if not s or not s.get("sftp_host"):
        raise HTTPException(status_code=400, detail="SFTP não configurado para este tenant")

    # Resolve filename: tenta campos diversos. Sync usa `audio_url` (`fusionpbx://...`)
    # e `external_id` (record_name do CDR). Mantém compat com docs antigos.
    name = (rec.get("url") or rec.get("storage_key") or rec.get("file_name")
            or rec.get("recording_uuid") or rec.get("audio_url")
            or rec.get("external_id") or "")
    if not name:
        raise HTTPException(status_code=400, detail="Gravação sem referência de arquivo")

    try:
        remote_path, total = await find_recording(
            host=s["sftp_host"], port=int(s.get("sftp_port") or 22),
            username=s["sftp_username"], password=s.get("sftp_password"),
            key=s.get("sftp_private_key"),
            relative_or_name=name,
            base_path=s.get("sftp_recordings_path"),
            domain_name=s.get("domain_name"),
        )
    except RecordingFetchError as e:
        raise HTTPException(status_code=502, detail=str(e))

    # Detect content type
    lower = remote_path.lower()
    if lower.endswith(".mp3"): ctype = "audio/mpeg"
    elif lower.endswith(".wav"): ctype = "audio/wav"
    elif lower.endswith(".ogg"): ctype = "audio/ogg"
    else: ctype = "application/octet-stream"
    ext = lower.rsplit(".", 1)[-1] if "." in lower else "mp3"

    # Friendly download filename: <agente>_<YYYY-MM-DD_HHMM>_<numero>.<ext>
    def _slug(v: str, maxlen: int = 40) -> str:
        v = (v or "").strip()
        # keep only safe chars; replace spaces and accents
        import re as _re
        import unicodedata as _ud
        v = _ud.normalize("NFKD", v).encode("ascii", "ignore").decode("ascii")
        v = _re.sub(r"[^A-Za-z0-9._-]+", "-", v).strip("-_.")
        return (v[:maxlen] or "rec")
    started = rec.get("started_at") or ""
    ts_part = ""
    if started:
        ts_part = started.replace("T", "_").replace(":", "")[:15]  # YYYY-MM-DD_HHMMSS
    parts = [
        _slug(rec.get("agent_name") or "agente", 30),
        ts_part or _slug(recording_id[:8], 10),
        _slug(rec.get("caller_number") or "", 20),
    ]
    download_name = "_".join(p for p in parts if p) + f".{ext}"
    # ?download=1 force attachment so browser saves with the friendly filename
    is_download = (request.query_params.get("download") in ("1", "true", "yes"))
    if is_download:
        disposition = f'attachment; filename="{download_name}"'
    else:
        # Inline sem filename: evita que alguns navegadores tratem como download
        disposition = "inline"

    # Range header
    range_header = request.headers.get("range") or request.headers.get("Range")
    start, end = 0, total - 1 if total else None
    status_code = 200
    headers: Dict[str, str] = {
        "Accept-Ranges": "bytes",
        "Content-Type": ctype,
        "Cache-Control": "private, max-age=3600",
        "Content-Disposition": disposition,
    }
    if range_header and total:
        try:
            r = range_header.replace("bytes=", "").split("-")
            start = int(r[0]) if r[0] else 0
            end = int(r[1]) if len(r) > 1 and r[1] else total - 1
            length = end - start + 1
            headers["Content-Range"] = f"bytes {start}-{end}/{total}"
            headers["Content-Length"] = str(length)
            status_code = 206
        except Exception:
            length = total
            headers["Content-Length"] = str(total)
    else:
        length = total if total else None
        if total: headers["Content-Length"] = str(total)

    async def gen():
        try:
            async for chunk in stream_recording(
                host=s["sftp_host"], port=int(s.get("sftp_port") or 22),
                username=s["sftp_username"], password=s.get("sftp_password"),
                key=s.get("sftp_private_key"),
                remote_path=remote_path,
                offset=start,
                length=length if range_header else None,
            ):
                yield chunk
        except RecordingFetchError as e:
            logger.error("stream_recording falhou: %s", e)
            return

    await write_audit(user, "play", "recording", recording_id,
                      f"Gravação {recording_id}", {"size": total, "range": range_header})
    return StreamingResponse(gen(), status_code=status_code, headers=headers, media_type=ctype)


@api.post("/fusionpbx/esl/test")
async def fusion_esl_test(user: dict = Depends(get_current_user), tenant_id: Optional[str] = None):
    tid = await _resolve_tenant_for_fusion(user, tenant_id)
    if user.get("role") not in ("super_admin", "admin"):
        raise HTTPException(status_code=403, detail="Sem permissão")
    s = await db.fusionpbx_settings.find_one({"tenant_id": tid})
    if not s or not s.get("esl_host"):
        raise HTTPException(status_code=400, detail="Configure ESL host primeiro")
    # TCP pré-check
    import socket
    try:
        with socket.create_connection((s["esl_host"], int(s.get("esl_port") or 8021)), timeout=4):
            pass
    except Exception as e:
        raise HTTPException(status_code=502,
            detail=f"Não consegui abrir TCP em {s['esl_host']}:{s.get('esl_port', 8021)} → "
                   f"[{type(e).__name__}] {e}. Libere a porta no firewall do FusionPBX e configure event_socket.conf.xml.")
    esl = FreeSwitchESL(
        host=s["esl_host"], port=int(s.get("esl_port") or 8021),
        password=s.get("esl_password") or "ClueCon",
        timeout=float(s.get("esl_timeout") or 5.0),
    )
    try:
        result = await esl.ping()
        # também conta canais ativos
        try:
            rows = await esl.show_channels()
            result["active_channels"] = len(rows)
        except Exception:
            result["active_channels"] = None
        return {"ok": True, **result}
    except FreeSwitchESLError as e:
        raise HTTPException(status_code=502, detail=str(e))


@api.post("/fusionpbx/test")
async def fusion_test(user: dict = Depends(get_current_user), tenant_id: Optional[str] = None):
    tid = await _resolve_tenant_for_fusion(user, tenant_id)
    if user.get("role") not in ("super_admin", "admin"):
        raise HTTPException(status_code=403, detail="Sem permissão")
    s = await db.fusionpbx_settings.find_one({"tenant_id": tid})
    if not s:
        raise HTTPException(status_code=400, detail="Configure a integração primeiro")
    ctype = s.get("connection_type") or "rest"

    if ctype == "db":
        if not s.get("db_host") or not s.get("db_username"):
            raise HTTPException(status_code=400, detail="Configure host e usuário do PostgreSQL")
        # Pré-checagem: TCP reach (firewall/route)
        import socket
        try:
            with socket.create_connection((s["db_host"], int(s.get("db_port") or 5432)), timeout=5):
                tcp_ok = True
        except Exception as e:
            raise HTTPException(
                status_code=502,
                detail=f"Não consegui abrir TCP em {s['db_host']}:{s.get('db_port', 5432)} → "
                       f"[{type(e).__name__}] {e}. Verifique firewall (iptables/ufw), pg_hba.conf e listen_addresses."
            )
        client = FusionPBXDBClient(
            host=s["db_host"], port=int(s.get("db_port") or 5432),
            database=s.get("db_name") or "fusionpbx",
            username=s["db_username"], password=s.get("db_password") or "",
            domain_uuid=s.get("domain_uuid"), ssl=bool(s.get("db_ssl")),
        )
        try:
            result = await client.ping()
            return {"ok": True, "tcp_reachable": tcp_ok, "mode": "db", **result}
        except FusionPBXDBError as e:
            raise HTTPException(status_code=502, detail=str(e))

    # REST mode
    if not s.get("base_url"):
        raise HTTPException(status_code=400, detail="Configure base_url primeiro")
    client = FusionPBXClient(
        base_url=s["base_url"], api_key=s.get("api_key"),
        username=s.get("username"), password=s.get("password"),
        domain_uuid=s.get("domain_uuid"), domain_name=s.get("domain_name"),
        verify_ssl=bool(s.get("verify_ssl", True)),
    )
    try:
        result = await client.ping()
        return {"ok": True, "mode": "rest", **result}
    except FusionPBXError as e:
        raise HTTPException(status_code=502, detail=str(e))


@api.post("/fusionpbx/fix-call-dates")
async def fix_call_dates(user: dict = Depends(get_current_user), tenant_id: Optional[str] = None):
    """One-shot: normalize legacy started_at/ended_at strings stored in Mongo
    (e.g. 'YYYY-MM-DD HH:MM:SS' → ISO 'YYYY-MM-DDTHH:MM:SS+00:00').
    Necessary because Mongo string-compare breaks date filters when format mixes."""
    tid = await _resolve_tenant_for_fusion(user, tenant_id)
    if user.get("role") not in ("super_admin", "admin"):
        raise HTTPException(status_code=403, detail="Sem permissão")
    fixed = 0
    cursor = db.calls.find({"tenant_id": tid, "started_at": {"$regex": " "}},
                           {"_id": 0, "id": 1, "started_at": 1, "ended_at": 1})
    async for c in cursor:
        upd = {}
        for key in ("started_at", "ended_at"):
            v = c.get(key) or ""
            if v and " " in v and "T" not in v:
                iso = v.replace(" ", "T", 1)
                if "+" not in iso and "Z" not in iso:
                    iso = iso + "+00:00"
                upd[key] = iso
        if upd:
            await db.calls.update_one({"id": c["id"]}, {"$set": upd})
            fixed += 1
    return {"ok": True, "fixed": fixed, "tenant_id": tid}


@api.post("/fusionpbx/clear-demo-data")
async def fusion_clear_demo_data(user: dict = Depends(get_current_user), tenant_id: Optional[str] = None):
    """Deletes all mocked/seeded data (entities without external_id) from the tenant.
    Real data synced from FusionPBX has external_id set, so it stays intact."""
    tid = await _resolve_tenant_for_fusion(user, tenant_id)
    if user.get("role") not in ("super_admin", "admin"):
        raise HTTPException(status_code=403, detail="Sem permissão")
    # Delete entities without external_id (= seeded demo data)
    demo_filter = {"tenant_id": tid, "$or": [
        {"external_id": {"$exists": False}}, {"external_id": None}
    ]}
    deleted = {}
    for col in ("agents", "queues", "calls", "recordings"):
        res = await db[col].delete_many(demo_filter)
        deleted[col] = res.deleted_count
    await write_audit(user, "clear", "demo_data", tid, "Limpeza de dados simulados", deleted)
    return {"ok": True, "deleted": deleted, "tenant_id": tid}


@api.post("/fusionpbx/resync-agents")
async def fusion_resync_agents(user: dict = Depends(get_current_user), tenant_id: Optional[str] = None):
    """Wipe all synced agents (real ones from PBX) and re-fetch from FusionPBX.
    Use when the source changed (e.g., switched from extensions to call_center_agents)
    or when agents got renamed in the PBX."""
    tid = await _resolve_tenant_for_fusion(user, tenant_id)
    if user.get("role") not in ("super_admin", "admin"):
        raise HTTPException(status_code=403, detail="Sem permissão")
    # Remove apenas agents sincronizados (com external_id)
    res = await db.agents.delete_many({"tenant_id": tid, "external_id": {"$ne": None}})
    deleted_agents = res.deleted_count
    # Roda sync para repopular
    summary = await _run_sync_for_tenant(tid)
    await write_audit(user, "resync", "agents", tid,
                      f"Resincronização de agentes ({deleted_agents} removidos)",
                      {"deleted": deleted_agents, "summary": summary})
    return {"ok": True, "deleted": deleted_agents, "summary": summary}


# ──────────────────────────────────────────────────────────────────────────
# PROVISIONING — cria entidades no FusionPBX a partir do Voxyra
# ──────────────────────────────────────────────────────────────────────────

class ProvisionQueueReq(BaseModel):
    name: str
    extension: str
    strategy: str = "ring-all"
    max_wait_time: int = 120


class ProvisionAgentReq(BaseModel):
    name: str
    extension: str
    agent_id: Optional[str] = None
    sip_password: Optional[str] = None
    pbx_password: Optional[str] = None
    voxyra_email: Optional[str] = None
    voxyra_password: Optional[str] = None
    queue_uuids: List[str] = []
    create_pbx_user: bool = True
    create_voxyra_user: bool = True


def _gen_pwd(n: int = 10) -> str:
    import secrets, string
    alphabet = string.ascii_letters + string.digits
    return "".join(secrets.choice(alphabet) for _ in range(n))


async def _get_db_client(tid: str):
    """Returns (FusionPBXDBClient, settings). Only PostgreSQL mode supports provisioning."""
    s = await db.fusionpbx_settings.find_one({"tenant_id": tid})
    if not s:
        raise HTTPException(status_code=400, detail="FusionPBX não configurado")
    if (s.get("connection_type") or "rest") != "db":
        raise HTTPException(status_code=400,
            detail="Provisionamento requer modo PostgreSQL Direto. Mude em Central PBX → Configuração.")
    if not s.get("db_host") or not s.get("db_username") or not s.get("domain_uuid"):
        raise HTTPException(status_code=400,
            detail="Configuração incompleta: host, usuário e domain_uuid são obrigatórios.")
    return FusionPBXDBClient(
        host=s["db_host"], port=int(s.get("db_port") or 5432),
        database=s.get("db_name") or "fusionpbx",
        username=s["db_username"], password=s.get("db_password") or "",
        domain_uuid=s["domain_uuid"], ssl=bool(s.get("db_ssl")),
    ), s


@api.post("/fusionpbx/provision/queue")
async def provision_queue(body: ProvisionQueueReq,
                          user: dict = Depends(get_current_user),
                          tenant_id: Optional[str] = None):
    tid = await _resolve_tenant_for_fusion(user, tenant_id)
    if user.get("role") not in ("super_admin", "admin"):
        raise HTTPException(status_code=403, detail="Sem permissão")
    client, _ = await _get_db_client(tid)
    try:
        result = await client.provision_queue(
            name=body.name, extension=str(body.extension),
            strategy=body.strategy, max_wait_time=body.max_wait_time,
        )
    except FusionPBXDBError as e:
        raise HTTPException(status_code=502, detail=str(e))
    doc = {
        "id": str(uuid.uuid4()), "tenant_id": tid,
        "external_id": result["call_center_queue_uuid"],
        "name": body.name, "extension": str(body.extension),
        "strategy": body.strategy, "max_wait": body.max_wait_time,
        "waiting": 0, "answered_today": 0, "missed_today": 0, "avg_wait_sec": 0,
        "created_at": datetime.now(timezone.utc).isoformat(),
        "updated_at": datetime.now(timezone.utc).isoformat(),
    }
    await db.queues.insert_one(doc)
    await write_audit(user, "create", "queue", result["call_center_queue_uuid"],
                      f"{body.name} (ext {body.extension})", body.dict())
    return {"ok": True, **result, "voxyra_queue_id": doc["id"]}


@api.post("/fusionpbx/provision/agent")
async def provision_agent(body: ProvisionAgentReq,
                          user: dict = Depends(get_current_user),
                          tenant_id: Optional[str] = None):
    tid = await _resolve_tenant_for_fusion(user, tenant_id)
    if user.get("role") not in ("super_admin", "admin"):
        raise HTTPException(status_code=403, detail="Sem permissão")
    client, settings = await _get_db_client(tid)
    domain_name = settings.get("domain_name") or ""
    agent_id = body.agent_id or str(body.extension)
    sip_password = body.sip_password or _gen_pwd(12)
    pbx_password = body.pbx_password or _gen_pwd(10)

    created: Dict[str, Any] = {"sip_password": sip_password}

    try:
        ext_res = await client.provision_extension(
            extension=str(body.extension), sip_password=sip_password,
            caller_id_name=body.name, caller_id_number=str(body.extension),
            description=f"Voxyra · {body.name}",
        )
        created["extension"] = ext_res
    except FusionPBXDBError as e:
        raise HTTPException(status_code=502, detail=f"extension: {e}")

    pbx_user_uuid: Optional[str] = None
    if body.create_pbx_user:
        try:
            user_res = await client.provision_user(
                username=agent_id, password_hash=pbx_password,
            )
            pbx_user_uuid = user_res["user_uuid"]
            created["pbx_user"] = {**user_res, "password": pbx_password}
            await client.link_extension_to_user(ext_res["extension_uuid"], pbx_user_uuid)
        except FusionPBXDBError as e:
            await client.delete_extension(ext_res["extension_uuid"])
            raise HTTPException(status_code=502, detail=f"pbx_user: {e}")

    try:
        ag_res = await client.provision_call_center_agent(
            agent_name=body.name, agent_id=agent_id,
            extension=str(body.extension), domain_name=domain_name,
        )
        created["agent"] = ag_res
    except FusionPBXDBError as e:
        await client.delete_extension(ext_res["extension_uuid"])
        if pbx_user_uuid:
            try:
                conn_del = await client._connect()
                await conn_del.execute(
                    "DELETE FROM v_users WHERE user_uuid = $1::uuid", pbx_user_uuid,
                )
                await conn_del.close()
            except Exception:
                pass
        raise HTTPException(status_code=502, detail=f"call_center_agent: {e}")

    queues_linked = 0
    for qid in body.queue_uuids or []:
        try:
            await client.assign_agent_to_queue(ag_res["call_center_agent_uuid"], qid)
            queues_linked += 1
        except Exception as e:
            logger.warning("Falha ao vincular agente à fila %s: %s", qid, e)
    created["queues_linked"] = queues_linked

    voxyra_agent_id = str(uuid.uuid4())
    avatar = AGENT_AVATARS[hash(agent_id) % len(AGENT_AVATARS)]
    await db.agents.insert_one({
        "id": voxyra_agent_id, "tenant_id": tid,
        "external_id": ag_res["call_center_agent_uuid"],
        "name": body.name, "username": agent_id,
        "extension": str(body.extension), "email": body.voxyra_email or "",
        "source": "call_center_agent", "avatar": avatar,
        "status": "offline", "queues": [], "calls_handled": 0,
        "avg_handle_sec": 0, "csat": 0, "adherence_pct": 0,
        "created_at": datetime.now(timezone.utc).isoformat(),
        "updated_at": datetime.now(timezone.utc).isoformat(),
    })
    created["voxyra_agent_id"] = voxyra_agent_id

    if body.create_voxyra_user:
        tenant = await db.tenants.find_one({"id": tid}, {"_id": 0})
        email = (body.voxyra_email or "").lower().strip()
        if not email and tenant:
            email = f"{agent_id}@{tenant.get('domain', 'voxyra.local')}"
        voxyra_password = body.voxyra_password or _gen_pwd(10)
        existing_u = await db.users.find_one({"tenant_id": tid, "email": email})
        if existing_u:
            created["voxyra_user"] = {"email": email, "id": existing_u["id"],
                                      "warning": "usuário já existia, não foi recriado"}
        else:
            voxyra_user_id = str(uuid.uuid4())
            await db.users.insert_one({
                "id": voxyra_user_id, "tenant_id": tid, "email": email,
                "name": body.name, "role": "agent",
                "password_hash": hash_password(voxyra_password),
                "permissions": None, "active": True,
                "agent_id": voxyra_agent_id,
                "created_at": datetime.now(timezone.utc).isoformat(),
            })
            created["voxyra_user"] = {"id": voxyra_user_id, "email": email,
                                      "password": voxyra_password}

    await write_audit(user, "create", "agent", ag_res["call_center_agent_uuid"],
                      f"{body.name} ({agent_id} · ext {body.extension})",
                      {"queues_linked": queues_linked,
                       "create_pbx_user": body.create_pbx_user,
                       "create_voxyra_user": body.create_voxyra_user})
    return {"ok": True, **created}


@api.delete("/fusionpbx/provision/queue/{queue_id}")
async def deprovision_queue(queue_id: str,
                            user: dict = Depends(get_current_user),
                            tenant_id: Optional[str] = None):
    tid = await _resolve_tenant_for_fusion(user, tenant_id)
    if user.get("role") not in ("super_admin", "admin"):
        raise HTTPException(status_code=403, detail="Sem permissão")
    q = await db.queues.find_one({"id": queue_id, "tenant_id": tid})
    if not q:
        raise HTTPException(status_code=404, detail="Fila não encontrada")
    if q.get("external_id"):
        try:
            client, _ = await _get_db_client(tid)
            await client.delete_queue(q["external_id"])
        except HTTPException:
            raise
        except Exception as e:
            logger.warning("Falha ao deletar fila no PBX: %s", e)
    await db.queues.delete_one({"id": queue_id, "tenant_id": tid})
    await write_audit(user, "delete", "queue", queue_id, q.get("name", ""), {})
    return {"ok": True}


@api.delete("/fusionpbx/provision/agent/{agent_id}")
async def deprovision_agent(agent_id: str,
                            user: dict = Depends(get_current_user),
                            tenant_id: Optional[str] = None):
    tid = await _resolve_tenant_for_fusion(user, tenant_id)
    if user.get("role") not in ("super_admin", "admin"):
        raise HTTPException(status_code=403, detail="Sem permissão")
    a = await db.agents.find_one({"id": agent_id, "tenant_id": tid})
    if not a:
        raise HTTPException(status_code=404, detail="Agente não encontrado")
    if a.get("external_id"):
        try:
            client, _ = await _get_db_client(tid)
            await client.delete_call_center_agent(a["external_id"])
            try:
                conn = await client._connect()
                await conn.execute(
                    """DELETE FROM v_extensions
                       WHERE domain_uuid = $1::uuid AND extension = $2""",
                    client.domain_uuid, a.get("extension", ""),
                )
                await conn.close()
            except Exception:
                pass
        except HTTPException:
            raise
        except Exception as e:
            logger.warning("Falha ao deletar agente no PBX: %s", e)
    await db.agents.delete_one({"id": agent_id, "tenant_id": tid})
    await db.users.delete_many({"tenant_id": tid, "agent_id": agent_id})
    await write_audit(user, "delete", "agent", agent_id, a.get("name", ""), {})
    return {"ok": True}




async def _run_sync_for_tenant(tid: str, cdr_limit: int = 5000) -> Dict[str, Any]:
    """Core sync logic reusable by manual endpoint + scheduler."""
    s = await db.fusionpbx_settings.find_one({"tenant_id": tid})
    if not s or not s.get("enabled"):
        return {"skipped": True, "reason": "not_enabled"}
    ctype = s.get("connection_type") or "rest"
    if ctype == "db":
        if not s.get("db_host") or not s.get("db_username"):
            return {"skipped": True, "reason": "db_not_configured"}
        client = FusionPBXDBClient(
            host=s["db_host"], port=int(s.get("db_port") or 5432),
            database=s.get("db_name") or "fusionpbx",
            username=s["db_username"], password=s.get("db_password") or "",
            domain_uuid=s.get("domain_uuid"), ssl=bool(s.get("db_ssl")),
        )
        ClientErr = FusionPBXDBError
    else:
        if not s.get("base_url"):
            return {"skipped": True, "reason": "rest_not_configured"}
        custom_paths = {k.replace("path_", ""): s[k] for k in ("path_extensions", "path_queues", "path_agents", "path_cdr") if s.get(k)}
        client = FusionPBXClient(
            base_url=s["base_url"], api_key=s.get("api_key"),
            username=s.get("username"), password=s.get("password"),
            domain_uuid=s.get("domain_uuid"), domain_name=s.get("domain_name"),
            verify_ssl=bool(s.get("verify_ssl", True)),
            custom_paths=custom_paths,
        )
        ClientErr = FusionPBXError
    summary = {"agents_synced": 0, "queues_synced": 0, "calls_synced": 0,
               "recordings_synced": 0,
               "errors": [], "started_at": datetime.now(timezone.utc).isoformat(),
               "agent_source": None}
    # Agents — APENAS Call Center Agents reais (v_call_center_agents).
    # Ramais SIP (v_extensions) NÃO são sincronizados como agentes — eles são
    # apenas o "telefone" onde o agente atende. O usuário cadastra agentes
    # diretamente em Apps → Call Center → Agents do FusionPBX.
    agent_records: list = []
    agent_source = "call_center_agent"
    try:
        cc_agents = await client.list_call_center_agents()
        if cc_agents:
            agent_records = [normalize_agent(a) for a in cc_agents]
    except ClientErr as e:
        summary["errors"].append(f"agents: {e}")
    summary["agent_source"] = agent_source
    # Sempre limpa ramais legados de db.agents (caso de instalações antigas)
    del_res = await db.agents.delete_many({
        "tenant_id": tid, "source": "extension",
    })
    if del_res.deleted_count > 0:
        summary["legacy_extensions_removed"] = del_res.deleted_count
        logger.info("[sync] Removidos %d ramais antigos de db.agents",
                    del_res.deleted_count)
    # Map FusionPBX agent_status → Voxyra status
    def _map_status(pbx_status: Optional[str]) -> Optional[str]:
        if not pbx_status: return None
        s = str(pbx_status).strip().lower()
        if "available" in s and "demand" not in s: return "online"
        if "available (on demand)" in s: return "online"
        if "break" in s or "pause" in s: return "paused"
        if "logged out" in s or "logged_out" in s: return "offline"
        return None
    for ag in agent_records:
        if not ag["external_id"]: continue
        existing = await db.agents.find_one({"tenant_id": tid, "external_id": ag["external_id"]})
        mapped = _map_status(ag.get("agent_status"))
        new_status = mapped if mapped else (existing.get("status") if existing else "offline")
        doc = {
            "tenant_id": tid, "external_id": ag["external_id"],
            "name": ag["name"], "username": ag.get("username", ""),
            "extension": ag.get("extension", ""), "email": ag.get("email", ""),
            "source": ag.get("source", "extension"),
            "agent_type": ag.get("agent_type"),
            "pbx_status": ag.get("agent_status"),
            "avatar": existing.get("avatar") if existing else AGENT_AVATARS[summary["agents_synced"] % len(AGENT_AVATARS)],
            "status": new_status,
            "queues": existing.get("queues", []) if existing else [],
            "calls_handled": existing.get("calls_handled", 0) if existing else 0,
            "avg_handle_sec": existing.get("avg_handle_sec", 0) if existing else 0,
            "csat": existing.get("csat", 0) if existing else 0,
            "adherence_pct": existing.get("adherence_pct", 0) if existing else 0,
            "updated_at": datetime.now(timezone.utc).isoformat(),
        }
        if existing:
            await db.agents.update_one({"id": existing["id"]}, {"$set": doc})
        else:
            doc["id"] = str(uuid.uuid4())
            doc["created_at"] = doc["updated_at"]
            await db.agents.insert_one(doc)
        summary["agents_synced"] += 1
    # Active calls → mark agents as 'incall'
    try:
        active = await client.list_active_calls()
        in_call_extensions = set()
        for ac in active or []:
            ext = ac.get("destination_number") or ac.get("caller_id_number")
            cs = (ac.get("channel_state") or "").lower()
            if cs in ("cs_execute", "cs_exchange_media", "cs_consume_media") and ext:
                in_call_extensions.add(str(ext))
        if in_call_extensions:
            await db.agents.update_many(
                {"tenant_id": tid, "extension": {"$in": list(in_call_extensions)}},
                {"$set": {"status": "incall"}},
            )
    except Exception as e:
        logger.warning("active-calls status sync falhou: %s", e)
    # Queues
    try:
        queues = await client.list_call_center_queues()
        for raw in queues:
            q = normalize_queue(raw)
            if not q["external_id"]: continue
            existing = await db.queues.find_one({"tenant_id": tid, "external_id": q["external_id"]})
            doc = {
                "tenant_id": tid, "external_id": q["external_id"],
                "name": q["name"], "extension": q["extension"],
                "strategy": q["strategy"], "max_wait": q["max_wait"],
                "waiting": existing.get("waiting", 0) if existing else 0,
                "answered_today": existing.get("answered_today", 0) if existing else 0,
                "missed_today": existing.get("missed_today", 0) if existing else 0,
                "avg_wait_sec": existing.get("avg_wait_sec", 0) if existing else 0,
                "updated_at": datetime.now(timezone.utc).isoformat(),
            }
            if existing:
                await db.queues.update_one({"id": existing["id"]}, {"$set": doc})
            else:
                doc["id"] = str(uuid.uuid4())
                doc["created_at"] = doc["updated_at"]
                await db.queues.insert_one(doc)
            summary["queues_synced"] += 1
    except ClientErr as e:
        summary["errors"].append(f"queues: {e}")
    # CDR
    try:
        cdrs = await client.list_cdr(limit=cdr_limit)
        # Build multiple lookup keys to match agents from CDR (cc_agent can be: uuid, agent_id, extension)
        agents_db = await db.agents.find(
            {"tenant_id": tid, "external_id": {"$ne": None}}, {"_id": 0}).to_list(2000)
        agent_map = {}
        for a in agents_db:
            if a.get("external_id"): agent_map[a["external_id"]] = a
            if a.get("username"):    agent_map[a["username"]] = a
            if a.get("extension"):   agent_map[a["extension"]] = a
            # FusionPBX uses "agent_id@domain" — also index without domain
            if a.get("username") and "@" not in a["username"] and s.get("domain_name"):
                agent_map[f"{a['username']}@{s['domain_name']}"] = a
        queue_map = {q["external_id"]: q for q in await db.queues.find(
            {"tenant_id": tid, "external_id": {"$ne": None}}, {"_id": 0}).to_list(2000)}
        for raw in cdrs:
            n = normalize_cdr(raw)
            if not n["external_id"]: continue
            existing = await db.calls.find_one({"tenant_id": tid, "external_id": n["external_id"]})
            # try multiple candidates for agent matching
            cc_agent = (n.get("agent_external_id") or "").strip()
            agent = None
            if cc_agent:
                agent = agent_map.get(cc_agent)
                if not agent and "@" in cc_agent:
                    agent = agent_map.get(cc_agent.split("@", 1)[0])
            queue = queue_map.get(n["queue_external_id"]) if n["queue_external_id"] else None
            doc = {
                "tenant_id": tid, "external_id": n["external_id"],
                "agent_id": agent["id"] if agent else None,
                "agent_name": agent["name"] if agent else "—",
                "queue_id": queue["id"] if queue else None,
                "queue_name": queue["name"] if queue else (n["queue_name"] or "—"),
                "direction": n["direction"],
                "caller_number": n["caller_number"], "callee_number": n["callee_number"],
                "disposition": n["disposition"], "abandonment_type": n["abandonment_type"],
                "wait_sec": n["wait_sec"], "duration_sec": n["duration_sec"],
                "started_at": n["started_at"], "ended_at": n["ended_at"],
                "recording_uuid": n.get("recording_uuid"),
            }
            if existing:
                await db.calls.update_one({"id": existing["id"]}, {"$set": doc})
                call_id_for_rec = existing["id"]
            else:
                doc["id"] = str(uuid.uuid4())
                await db.calls.insert_one(doc)
                call_id_for_rec = doc["id"]
            # Idempotent recording upsert: insert if recording_uuid present, answered,
            # and no existing recording doc has this external_id for this tenant.
            if n.get("recording_uuid") and n["disposition"] == "answered":
                rec_exists = await db.recordings.find_one(
                    {"tenant_id": tid, "external_id": n["recording_uuid"]},
                    {"_id": 0, "id": 1},
                )
                if not rec_exists:
                    rec_url = await client.get_recording_url(n["recording_uuid"])
                    await db.recordings.insert_one({
                        "id": str(uuid.uuid4()), "tenant_id": tid,
                        "external_id": n["recording_uuid"], "call_id": call_id_for_rec,
                        "agent_id": doc["agent_id"], "agent_name": doc["agent_name"],
                        "queue_id": doc["queue_id"], "queue_name": doc["queue_name"],
                        "caller_number": n["caller_number"],
                        "duration_sec": n["duration_sec"],
                        "audio_url": rec_url, "size_mb": round(n["duration_sec"] * 0.012, 2),
                        "started_at": n["started_at"], "notes": "",
                    })
                    summary["recordings_synced"] = summary.get("recordings_synced", 0) + 1
            summary["calls_synced"] += 1
    except ClientErr as e:
        summary["errors"].append(f"cdr: {e}")
    summary["finished_at"] = datetime.now(timezone.utc).isoformat()
    # Recompute calls_handled & avg_handle_sec per agent (last 24h answered calls)
    try:
        cutoff_24h = (datetime.now(timezone.utc) - timedelta(hours=24)).isoformat()
        pipeline = [
            {"$match": {"tenant_id": tid, "disposition": "answered",
                        "started_at": {"$gte": cutoff_24h},
                        "agent_id": {"$ne": None}}},
            {"$group": {"_id": "$agent_id",
                        "calls_handled": {"$sum": 1},
                        "avg_handle_sec": {"$avg": "$duration_sec"}}},
        ]
        async for row in db.calls.aggregate(pipeline):
            await db.agents.update_one(
                {"id": row["_id"], "tenant_id": tid},
                {"$set": {"calls_handled": int(row["calls_handled"]),
                          "avg_handle_sec": int(row.get("avg_handle_sec") or 0)}},
            )
    except Exception as e:
        logger.warning("[stats] Falha ao recomputar métricas por agente: %s", e)
    # Recompute per-queue stats today (UTC)
    try:
        today_iso = datetime.now(timezone.utc).replace(hour=0, minute=0, second=0, microsecond=0).isoformat()
        q_pipeline = [
            {"$match": {"tenant_id": tid, "started_at": {"$gte": today_iso},
                        "queue_id": {"$ne": None}}},
            {"$group": {"_id": "$queue_id",
                        "answered": {"$sum": {"$cond": [{"$eq": ["$disposition", "answered"]}, 1, 0]}},
                        "missed": {"$sum": {"$cond": [{"$in": ["$disposition", ["missed", "abandoned"]]}, 1, 0]}},
                        "avg_wait": {"$avg": "$wait_sec"}}},
        ]
        async for row in db.calls.aggregate(q_pipeline):
            await db.queues.update_one(
                {"id": row["_id"], "tenant_id": tid},
                {"$set": {"answered_today": int(row["answered"]),
                          "missed_today": int(row["missed"]),
                          "avg_wait_sec": int(row.get("avg_wait") or 0)}},
            )
    except Exception as e:
        logger.warning("[stats] Falha ao recomputar métricas por fila: %s", e)
    # Re-aplica contatos e status em memória do mod_callcenter para TODOS os
    # agentes sincronizados. reload mod_callcenter não atualiza agentes já
    # carregados — só o agent set contact/status via ESL resolve.
    try:
        s_cfg = await db.fusionpbx_settings.find_one({"tenant_id": tid}) or {}
        if s_cfg.get("enabled") and s_cfg.get("esl_host"):
            synced_agents = await db.agents.find(
                {"tenant_id": tid, "external_id": {"$ne": None}},
                {"_id": 0, "external_id": 1}).to_list(500)
            for a in synced_agents:
                try:
                    await _pbx_resync_agent_from_db(tid, a["external_id"])
                except Exception as e:
                    logger.warning("resync live de agente falhou: %s", e)
    except Exception as e:
        logger.warning("[sync] Falha no resync live de agentes: %s", e)
    summary["status"] = "error" if summary["errors"] and summary["agents_synced"] == 0 else "ok"
    await db.fusionpbx_settings.update_one(
        {"tenant_id": tid},
        {"$set": {"last_sync_at": summary["finished_at"], "last_sync_status": summary["status"],
                  "last_sync_summary": summary}},
    )
    return summary


_SYNC_SCHEDULER_STATE = {"running": False, "last_run": None, "enabled": True}


async def _fusionpbx_scheduler():
    """Background task that syncs all enabled tenants every minute."""
    await asyncio.sleep(30)  # small delay on startup
    while _SYNC_SCHEDULER_STATE["enabled"]:
        try:
            tenants_to_sync = await db.fusionpbx_settings.find(
                {"enabled": True, "$or": [
                    {"connection_type": "db", "db_host": {"$nin": [None, ""]}},
                    {"connection_type": {"$in": [None, "rest"]}, "base_url": {"$nin": [None, ""]}},
                ]},
                {"_id": 0, "tenant_id": 1, "sync_interval_minutes": 1, "last_sync_at": 1},
            ).to_list(1000)
            now = datetime.now(timezone.utc)
            for t in tenants_to_sync:
                tid = t["tenant_id"]
                interval_min = max(1, int(t.get("sync_interval_minutes") or 1))
                last = t.get("last_sync_at")
                if last:
                    try:
                        last_dt = datetime.fromisoformat(last.replace("Z", "+00:00"))
                        if (now - last_dt).total_seconds() < interval_min * 60 - 5:
                            continue
                    except Exception:
                        pass
                logger.info("[scheduler] Syncing tenant %s (interval=%dmin)", tid, interval_min)
                try:
                    result = await _run_sync_for_tenant(tid)
                    if not result.get("skipped"):
                        logger.info("[scheduler] Synced %s: %d agents, %d queues, %d calls",
                                    tid, result.get("agents_synced", 0),
                                    result.get("queues_synced", 0), result.get("calls_synced", 0))
                except Exception as e:
                    logger.exception("[scheduler] Erro no tenant %s: %s", tid, e)
            _SYNC_SCHEDULER_STATE["last_run"] = now.isoformat()
        except Exception as e:
            logger.exception("[scheduler] Falha: %s", e)
        await asyncio.sleep(30)  # check every 30s; per-tenant interval gate handles frequency


@api.get("/fusionpbx/scheduler/status")
async def fusion_scheduler_status(user: dict = Depends(get_current_user)):
    if user.get("role") not in ("super_admin", "admin"):
        raise HTTPException(status_code=403, detail="Sem permissão")
    enabled_tenants = await db.fusionpbx_settings.count_documents(
        {"enabled": True, "$or": [
            {"connection_type": "db", "db_host": {"$nin": [None, ""]}},
            {"connection_type": {"$in": [None, "rest"]}, "base_url": {"$nin": [None, ""]}},
        ]}
    )
    return {
        "running": _SYNC_SCHEDULER_STATE["running"] or True,  # if backend is up, scheduler is up
        "last_check": _SYNC_SCHEDULER_STATE.get("last_run"),
        "enabled_tenants": enabled_tenants,
    }


@api.get("/fusionpbx/diagnostics")
async def fusion_diagnostics(user: dict = Depends(get_current_user), tenant_id: Optional[str] = None):
    """Returns comprehensive status: settings, last sync, counts of real vs demo data,
    and the latest synced records so the admin can verify data is coming in."""
    tid = await _resolve_tenant_for_fusion(user, tenant_id)
    if user.get("role") not in ("super_admin", "admin"):
        raise HTTPException(status_code=403, detail="Sem permissão")

    settings = await db.fusionpbx_settings.find_one({"tenant_id": tid}, {"_id": 0}) or {}

    async def _count_real_vs_demo(col: str) -> Dict[str, int]:
        real = await db[col].count_documents({"tenant_id": tid, "external_id": {"$exists": True, "$ne": None}})
        demo = await db[col].count_documents({"tenant_id": tid, "$or": [{"external_id": {"$exists": False}}, {"external_id": None}]})
        return {"real": real, "demo": demo, "total": real + demo}

    counts = {
        "agents": await _count_real_vs_demo("agents"),
        "queues": await _count_real_vs_demo("queues"),
        "calls": await _count_real_vs_demo("calls"),
        "recordings": await _count_real_vs_demo("recordings"),
    }

    # Last 10 synced calls, 10 agents, 50 queues
    recent_real_calls = await db.calls.find(
        {"tenant_id": tid, "external_id": {"$exists": True, "$ne": None}}, {"_id": 0}
    ).sort("started_at", -1).to_list(10)
    recent_real_agents = await db.agents.find(
        {"tenant_id": tid, "external_id": {"$exists": True, "$ne": None}}, {"_id": 0}
    ).sort("updated_at", -1).to_list(10)
    recent_real_queues = await db.queues.find(
        {"tenant_id": tid, "external_id": {"$exists": True, "$ne": None}}, {"_id": 0}
    ).to_list(50)

    # Audit log of recent syncs
    sync_history = await db.audit_logs.find(
        {"tenant_id": tid, "resource": "fusionpbx", "action": "sync"}, {"_id": 0}
    ).sort("created_at", -1).to_list(5)

    return {
        "tenant_id": tid,
        "settings": {
            "enabled": settings.get("enabled", False),
            "configured": bool(settings.get("base_url")),
            "base_url": settings.get("base_url"),
            "domain_uuid": settings.get("domain_uuid"),
            "domain_name": settings.get("domain_name"),
            "last_sync_at": settings.get("last_sync_at"),
            "last_sync_status": settings.get("last_sync_status"),
            "last_sync_summary": settings.get("last_sync_summary", {}),
        },
        "counts": counts,
        "recent_calls": recent_real_calls,
        "recent_agents": recent_real_agents,
        "recent_queues": recent_real_queues,
        "sync_history": sync_history,
    }



@api.post("/fusionpbx/sync")
async def fusion_sync(user: dict = Depends(get_current_user), tenant_id: Optional[str] = None,
                      cdr_limit: int = 5000):
    """Sync manual (chamado pelo botão 'Sincronizar Agora')."""
    tid = await _resolve_tenant_for_fusion(user, tenant_id)
    if user.get("role") not in ("super_admin", "admin"):
        raise HTTPException(status_code=403, detail="Sem permissão")
    s = await db.fusionpbx_settings.find_one({"tenant_id": tid})
    if not s or not s.get("base_url"):
        raise HTTPException(status_code=400, detail="FusionPBX não configurado para este tenant")
    if not s.get("enabled", False):
        raise HTTPException(status_code=400, detail="Integração FusionPBX desativada")
    try:
        summary = await _run_sync_for_tenant(tid, cdr_limit=cdr_limit)
    except Exception as e:
        raise HTTPException(status_code=502, detail=f"Erro na sincronização: {e}")
    await write_audit(user, "sync", "fusionpbx", tid, "Sync FusionPBX",
                      {"agents": summary.get("agents_synced", 0),
                       "queues": summary.get("queues_synced", 0),
                       "calls": summary.get("calls_synced", 0),
                       "recordings": summary.get("recordings_synced", 0)})
    return summary


# ---------- Seeding ----------
DEMO_AGENTS_A = [("Ana Silva", "ana.silva"), ("Bruno Lima", "bruno.lima"),
                 ("Carla Santos", "carla.santos"), ("Diego Costa", "diego.costa")]
DEMO_AGENTS_B = [("Eliana Rocha", "eliana.rocha"), ("Felipe Souza", "felipe.souza"),
                 ("Gabriela Alves", "gabriela.alves"), ("Henrique Dias", "henrique.dias")]
DEMO_QUEUES_A = [("Vendas", "1001"), ("Suporte Técnico", "1002")]
DEMO_QUEUES_B = [("Financeiro", "1003"), ("Retenção", "1004")]

async def _seed_tenant(tid: str, domain: str, agents_def: list, queues_def: list, n_calls: int = 200):
    queues = []
    for name, ext in queues_def:
        queues.append({
            "id": str(uuid.uuid4()), "tenant_id": tid, "name": name, "extension": ext,
            "strategy": random.choice(["ring-all", "longest-idle", "round-robin"]),
            "max_wait": random.choice([60, 120, 180]),
            "waiting": random.randint(0, 8),
            "answered_today": random.randint(30, 200),
            "missed_today": random.randint(0, 20),
            "avg_wait_sec": random.randint(10, 90),
            "created_at": datetime.now(timezone.utc).isoformat(),
        })
    await db.queues.insert_many(queues)
    agents = []
    for i, (name, username) in enumerate(agents_def):
        agents.append({
            "id": str(uuid.uuid4()), "tenant_id": tid, "name": name, "username": username,
            "extension": str(1100 + i),
            "email": f"{username}@{domain}",
            "avatar": AGENT_AVATARS[i % len(AGENT_AVATARS)],
            "status": random.choice(STATUSES),
            "queues": random.sample([q["id"] for q in queues], k=min(2, len(queues))),
            "calls_handled": random.randint(20, 180),
            "avg_handle_sec": random.randint(120, 480),
            "csat": round(random.uniform(3.8, 4.9), 2),
            "adherence_pct": round(random.uniform(78, 99), 1),
            "created_at": datetime.now(timezone.utc).isoformat(),
        })
    await db.agents.insert_many(agents)
    calls = []; recs = []
    for _ in range(n_calls):
        ag = random.choice(agents); q = random.choice(queues)
        started = datetime.now(timezone.utc) - timedelta(
            days=random.randint(0, 27), hours=random.randint(0, 23), minutes=random.randint(0, 59))
        duration = random.randint(30, 900)
        disp = random.choices(DISPOSITIONS, weights=[70, 15, 10, 5])[0]
        ab_type = "agent_loss" if disp == "missed" else "queue_abandon" if disp == "abandoned" else None
        cid = str(uuid.uuid4())
        calls.append({
            "id": cid, "tenant_id": tid,
            "agent_id": ag["id"], "agent_name": ag["name"],
            "queue_id": q["id"], "queue_name": q["name"],
            "direction": random.choice(CALL_DIR),
            "caller_number": f"+55 11 9{random.randint(1000,9999)}-{random.randint(1000,9999)}",
            "callee_number": ag["extension"], "disposition": disp,
            "abandonment_type": ab_type,
            "wait_sec": random.randint(5, 120) if disp in ("missed", "abandoned") else random.randint(0, 30),
            "duration_sec": duration if disp == "answered" else random.randint(5, 40),
            "started_at": started.isoformat(),
            "ended_at": (started + timedelta(seconds=duration)).isoformat(),
        })
        if disp == "answered":
            recs.append({
                "id": str(uuid.uuid4()), "tenant_id": tid, "call_id": cid,
                "agent_id": ag["id"], "agent_name": ag["name"],
                "queue_id": q["id"], "queue_name": q["name"],
                "caller_number": calls[-1]["caller_number"],
                "duration_sec": duration, "audio_url": SAMPLE_AUDIO,
                "size_mb": round(duration * 0.012, 2),
                "started_at": started.isoformat(), "notes": "",
            })
    if calls: await db.calls.insert_many(calls)
    if recs: await db.recordings.insert_many(recs)
    # Seed default users for tenant
    users_def = [
        (f"admin@{domain}", "admin123", "Administrador", "admin"),
        (f"supervisor@{domain}", "super123", "Supervisor", "supervisor"),
        (f"agent@{domain}", "agent123", "Agente Demo", "agent"),
    ]
    user_docs = []
    for em, pw, nm, role in users_def:
        user_docs.append({
            "id": str(uuid.uuid4()), "tenant_id": tid, "email": em.lower(),
            "name": nm, "role": role, "password_hash": hash_password(pw),
            "permissions": None, "active": True,
            "agent_id": agents[0]["id"] if role == "agent" else None,
            "created_at": datetime.now(timezone.utc).isoformat(),
        })
    await db.users.insert_many(user_docs)

async def seed_data():
    # Super admin
    sa_email = os.environ.get("SUPER_ADMIN_EMAIL", "").lower()
    sa_pw = os.environ.get("SUPER_ADMIN_PASSWORD", "")
    if sa_email and sa_pw:
        existing = await db.users.find_one({"email": sa_email, "role": "super_admin"})
        if not existing:
            await db.users.insert_one({
                "id": str(uuid.uuid4()), "tenant_id": None, "email": sa_email,
                "name": "Super Admin", "role": "super_admin",
                "password_hash": hash_password(sa_pw),
                "permissions": None, "active": True,
                "created_at": datetime.now(timezone.utc).isoformat(),
            })
        elif not verify_password(sa_pw, existing["password_hash"]):
            await db.users.update_one({"id": existing["id"]}, {"$set": {"password_hash": hash_password(sa_pw)}})

    # Default plans (only seed if none exist)
    if await db.plans.count_documents({}) == 0:
        plans = [
            {"id": str(uuid.uuid4()), "name": "Basic", "description": "Ideal para times pequenos",
             "monthly_price": 99.0, "max_users": 5, "max_agents": 5,
             "features": ["dashboard", "recordings", "reports", "queues_view", "agents_view", "support_email"],
             "active": True, "created_at": datetime.now(timezone.utc).isoformat()},
            {"id": str(uuid.uuid4()), "name": "Pro", "description": "Para operações em crescimento",
             "monthly_price": 299.0, "max_users": 25, "max_agents": 25,
             "features": ["dashboard", "realtime", "recordings", "recordings_download", "recordings_notes",
                          "reports", "reports_export", "queues_view", "queues_edit", "abandoned_analytics",
                          "agents_view", "agents_edit", "agent_scoring", "tv_panel", "audit_logs",
                          "users_management", "support_email"],
             "active": True, "created_at": datetime.now(timezone.utc).isoformat()},
            {"id": str(uuid.uuid4()), "name": "Enterprise", "description": "Sem limites para grandes operações",
             "monthly_price": 799.0, "max_users": 999, "max_agents": 999,
             "features": [f["key"] for f in PLAN_FEATURES_CATALOG],
             "active": True, "created_at": datetime.now(timezone.utc).isoformat()},
        ]
        await db.plans.insert_many(plans)

    # Demo tenants
    if os.environ.get("SEED_TENANT_DEMO", "").lower() == "true":
        if await db.tenants.count_documents({}) == 0:
            tA = {
                "id": str(uuid.uuid4()), "domain": "empresa-a.local", "name": "Empresa A",
                "accent_color": "#0EA5E9", "logo_url": None, "timezone": "America/Sao_Paulo",
                "max_users": 50, "max_agents": 50, "active": True,
                "created_at": datetime.now(timezone.utc).isoformat(),
            }
            tB = {
                "id": str(uuid.uuid4()), "domain": "empresa-b.local", "name": "Empresa B",
                "accent_color": "#10B981", "logo_url": None, "timezone": "America/Sao_Paulo",
                "max_users": 25, "max_agents": 25, "active": True,
                "created_at": datetime.now(timezone.utc).isoformat(),
            }
            await db.tenants.insert_many([tA, tB])
            await _seed_tenant(tA["id"], tA["domain"], DEMO_AGENTS_A, DEMO_QUEUES_A, n_calls=250)
            await _seed_tenant(tB["id"], tB["domain"], DEMO_AGENTS_B, DEMO_QUEUES_B, n_calls=180)

# ---------- Startup ----------
@app.on_event("startup")
async def on_startup():
    await db.tenants.create_index("domain", unique=True)
    await db.users.create_index([("tenant_id", 1), ("email", 1)], unique=True)
    await db.agents.create_index([("tenant_id", 1), ("id", 1)])
    await db.queues.create_index([("tenant_id", 1), ("id", 1)])
    await db.calls.create_index([("tenant_id", 1), ("started_at", -1)])
    await db.recordings.create_index([("tenant_id", 1), ("started_at", -1)])
    await db.audit_logs.create_index([("tenant_id", 1), ("created_at", -1)])
    await db.charges.create_index([("tenant_id", 1), ("created_at", -1)])
    await db.charges.create_index("external_id")
    await db.fusionpbx_settings.create_index("tenant_id", unique=True)
    await db.webhook_events.create_index([("source", 1), ("delivery_id", 1)])
    # Start FusionPBX auto-sync scheduler (checks every 30s, syncs per-tenant based on interval)
    asyncio.create_task(_fusionpbx_scheduler())
    logger.info("FusionPBX auto-sync scheduler iniciado")
    await seed_data()

@app.on_event("shutdown")
async def on_shutdown():
    client.close()

app.include_router(api)
# Serve uploaded assets (logos, wallpapers, favicons) at /uploads/<filename>
app.mount("/uploads", StaticFiles(directory=str(UPLOAD_DIR)), name="uploads")
app.add_middleware(
    CORSMiddleware,
    allow_credentials=True, allow_origin_regex=".*",
    allow_methods=["*"], allow_headers=["*"],
)
logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(name)s - %(levelname)s - %(message)s')
logger = logging.getLogger(__name__)
